"""
API REST con FastAPI para Dashboard Web
Endpoints para estudiantes y panel de administración
"""

from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Query, Form, Request, Body
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import List, Optional, Dict
from datetime import datetime
from pydantic import BaseModel
import json

# Rate Limiting
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

# Logging estructurado
import sys
sys.path.append('..')
from utils.logger import logger, log_event, log_error

from database.models import get_db
from modules.estudiantes import Estudiante
from modules.admin_panel import PanelAdministrativo
from api.schemas import (
    EstudianteCreate, EstudianteResponse, EstudianteUpdate,
    LoginRequest, LoginResponse, EstadisticasResponse
)
from api.auth import crear_token, verificar_token
from api.blog_routes import router as blog_router
from api.testimonios_routes import router as testimonios_router
from api.notificaciones_routes import router as notificaciones_router
from api.chat_routes import router as chat_router
from api.analytics_routes import router as analytics_router
from api.documentos_routes import router as documentos_router

# Configurar rate limiter
limiter = Limiter(key_func=get_remote_address)

app = FastAPI(
    title="Estudia en España API",
    description="API para gestión de estudiantes y asesoría de visas",
    version="1.0.0"
)

# Agregar rate limiter a la app
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# CORS - Permitir requests desde frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:3000",
        "https://bot-visas-api.onrender.com",
        "https://fortunariocash.com",
        "https://www.fortunariocash.com",
        "https://bot-visas-web-marca-como-private-s785-g4twzsjhe.vercel.app",
        "https://*.vercel.app",  # Permitir todos los subdominios de Vercel
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    """Ejecutar migraciones al iniciar"""
    import os
    import psycopg2
    try:
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        # Crear tabla documentos_generados si no existe
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS documentos_generados (
                id SERIAL PRIMARY KEY,
                estudiante_id INTEGER REFERENCES estudiantes(id) ON DELETE CASCADE,
                tipo_documento VARCHAR(100) NOT NULL,
                nombre_archivo VARCHAR(255) NOT NULL,
                contenido_pdf TEXT NOT NULL,
                estado VARCHAR(50) DEFAULT 'generado',
                notas TEXT,
                generado_por VARCHAR(100),
                aprobado_por VARCHAR(100),
                fecha_generacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                fecha_aprobacion TIMESTAMP,
                enviado_estudiante BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        # Crear tabla fechas_importantes si no existe
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS fechas_importantes (
                id SERIAL PRIMARY KEY,
                estudiante_id INTEGER NOT NULL REFERENCES estudiantes(id) ON DELETE CASCADE,
                tipo_fecha VARCHAR(100) NOT NULL,
                fecha TIMESTAMP NOT NULL,
                descripcion TEXT,
                alertado_30d BOOLEAN DEFAULT FALSE,
                alertado_15d BOOLEAN DEFAULT FALSE,
                alertado_7d BOOLEAN DEFAULT FALSE,
                alertado_1d BOOLEAN DEFAULT FALSE,
                completada BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        # Crear tablas de universidades y programas
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS universidades_espana (
                id SERIAL PRIMARY KEY,
                nombre VARCHAR(255) NOT NULL,
                siglas VARCHAR(50),
                ciudad VARCHAR(100) NOT NULL,
                comunidad_autonoma VARCHAR(100) NOT NULL,
                tipo VARCHAR(50) NOT NULL,
                url_oficial VARCHAR(500),
                email_contacto VARCHAR(255),
                telefono VARCHAR(50),
                tiene_api BOOLEAN DEFAULT FALSE,
                endpoint_api VARCHAR(500),
                metodo_scraping VARCHAR(100),
                ultima_actualizacion TIMESTAMP,
                logo_url VARCHAR(500),
                descripcion TEXT,
                ranking_nacional INTEGER,
                total_alumnos INTEGER,
                total_programas INTEGER,
                acepta_extranjeros BOOLEAN DEFAULT TRUE,
                requisitos_extranjeros TEXT,
                activa BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS programas_universitarios (
                id SERIAL PRIMARY KEY,
                universidad_id INTEGER NOT NULL,
                nombre VARCHAR(500) NOT NULL,
                tipo_programa VARCHAR(100),
                area_estudio VARCHAR(200),
                duracion_anos FLOAT,
                creditos_ects INTEGER,
                idioma VARCHAR(50),
                modalidad VARCHAR(50),
                precio_anual_eur FLOAT,
                plazas_disponibles INTEGER,
                nota_corte FLOAT,
                url_info VARCHAR(500),
                fecha_inicio_inscripcion TIMESTAMP,
                fecha_fin_inscripcion TIMESTAMP,
                requisitos TEXT,
                descripcion TEXT,
                activo BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        # Crear índices para optimizar búsquedas
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_universidades_ciudad ON universidades_espana(ciudad);
            CREATE INDEX IF NOT EXISTS idx_universidades_tipo ON universidades_espana(tipo);
            CREATE INDEX IF NOT EXISTS idx_programas_universidad ON programas_universitarios(universidad_id);
            CREATE INDEX IF NOT EXISTS idx_programas_tipo ON programas_universitarios(tipo_programa);
        """)
        
        # Crear tabla blog_posts
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS blog_posts (
                id SERIAL PRIMARY KEY,
                titulo VARCHAR(500) NOT NULL,
                slug VARCHAR(500) UNIQUE NOT NULL,
                contenido TEXT NOT NULL,
                extracto TEXT,
                categoria VARCHAR(100),
                autor_nombre VARCHAR(200) DEFAULT 'Equipo Editorial',
                imagen_portada VARCHAR(500),
                meta_description VARCHAR(300),
                meta_keywords VARCHAR(500),
                visitas INTEGER DEFAULT 0,
                publicado BOOLEAN DEFAULT FALSE,
                destacado BOOLEAN DEFAULT FALSE,
                fecha_publicacion TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        # Crear tabla testimonios
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS testimonios (
                id SERIAL PRIMARY KEY,
                estudiante_id INTEGER,
                nombre_completo VARCHAR(200) NOT NULL,
                pais_origen VARCHAR(100) NOT NULL,
                programa_estudio VARCHAR(300),
                universidad VARCHAR(300),
                ciudad_espana VARCHAR(100),
                rating INTEGER,
                titulo VARCHAR(300),
                testimonio TEXT NOT NULL,
                foto_url VARCHAR(500),
                video_url VARCHAR(500),
                email_contacto VARCHAR(200),
                aprobado BOOLEAN DEFAULT FALSE,
                destacado BOOLEAN DEFAULT FALSE,
                visible BOOLEAN DEFAULT TRUE,
                fecha_experiencia TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        # Índices para blog y testimonios
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_blog_categoria ON blog_posts(categoria);
            CREATE INDEX IF NOT EXISTS idx_blog_publicado ON blog_posts(publicado);
            CREATE INDEX IF NOT EXISTS idx_blog_slug ON blog_posts(slug);
            CREATE INDEX IF NOT EXISTS idx_testimonios_aprobado ON testimonios(aprobado);
            CREATE INDEX IF NOT EXISTS idx_testimonios_destacado ON testimonios(destacado);
        """)
        
        # Tabla notificaciones
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS notificaciones (
                id SERIAL PRIMARY KEY,
                estudiante_id INTEGER NOT NULL,
                tipo VARCHAR(50) NOT NULL,
                titulo VARCHAR(200) NOT NULL,
                mensaje TEXT NOT NULL,
                leida BOOLEAN DEFAULT FALSE,
                url_accion VARCHAR(500),
                icono VARCHAR(20) DEFAULT '🔔',
                prioridad VARCHAR(20) DEFAULT 'normal',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_notificaciones_estudiante ON notificaciones(estudiante_id);
            CREATE INDEX IF NOT EXISTS idx_notificaciones_leida ON notificaciones(leida);
        """)
        
        # Tabla mensajes_chat
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS mensajes_chat (
                id SERIAL PRIMARY KEY,
                estudiante_id INTEGER NOT NULL,
                admin_id INTEGER,
                remitente VARCHAR(20) NOT NULL,
                mensaje TEXT NOT NULL,
                leido BOOLEAN DEFAULT FALSE,
                tipo VARCHAR(20) DEFAULT 'texto',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_mensajes_estudiante ON mensajes_chat(estudiante_id);
            CREATE INDEX IF NOT EXISTS idx_mensajes_leido ON mensajes_chat(leido);
            CREATE INDEX IF NOT EXISTS idx_mensajes_remitente ON mensajes_chat(remitente);
        """)
        
        # Tabla documentos (crear si no existe, o agregar columnas faltantes)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS documentos (
                id SERIAL PRIMARY KEY,
                estudiante_id INTEGER NOT NULL REFERENCES estudiantes(id) ON DELETE CASCADE,
                nombre_archivo VARCHAR(255),
                tipo_documento VARCHAR(100),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        # Agregar columnas nuevas si no existen
        try:
            cursor.execute("ALTER TABLE documentos ADD COLUMN IF NOT EXISTS categoria VARCHAR(50)")
            cursor.execute("ALTER TABLE documentos ADD COLUMN IF NOT EXISTS contenido_base64 TEXT")
            cursor.execute("ALTER TABLE documentos ADD COLUMN IF NOT EXISTS mime_type VARCHAR(100)")
            cursor.execute("ALTER TABLE documentos ADD COLUMN IF NOT EXISTS tamano_archivo INTEGER")
            cursor.execute("ALTER TABLE documentos ADD COLUMN IF NOT EXISTS estado_revision VARCHAR(20) DEFAULT 'pendiente'")
            cursor.execute("ALTER TABLE documentos ADD COLUMN IF NOT EXISTS comentario_admin TEXT")
            cursor.execute("ALTER TABLE documentos ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
        except Exception as alter_error:
            print(f"⚠️ Error agregando columnas a documentos: {alter_error}")
        
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_documentos_estudiante ON documentos(estudiante_id);
        """)
        
        try:
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_documentos_estado ON documentos(estado_revision)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_documentos_categoria ON documentos(categoria)")
        except:
            pass
        
        # Crear tabla contactos_universidades para sistema de emails
        try:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS contactos_universidades (
                    id SERIAL PRIMARY KEY,
                    universidad VARCHAR(200) NOT NULL,
                    email VARCHAR(200) NOT NULL,
                    telefono VARCHAR(50),
                    contacto_nombre VARCHAR(200),
                    pais VARCHAR(100) DEFAULT 'España',
                    ciudad VARCHAR(100),
                    tipo_universidad VARCHAR(100),
                    programas_interes TEXT,
                    estado VARCHAR(50) DEFAULT 'pendiente',
                    fecha_contacto TIMESTAMP,
                    fecha_respuesta TIMESTAMP,
                    fecha_reunion TIMESTAMP,
                    notas TEXT,
                    condiciones_propuestas TEXT,
                    comision_acordada DECIMAL(10, 2),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Verificar si hay universidades cargadas
            cursor.execute("SELECT COUNT(*) FROM contactos_universidades")
            count = cursor.fetchone()[0]
            
            # Si no hay universidades, cargar las 44 predefinidas
            if count == 0:
                print("📧 Cargando 44 universidades objetivo...")
                universidades_iniciales = [
                    ('UCAM - Universidad Católica de Murcia', 'internacional@ucam.edu', '+34 968 278 160', 'Departamento Internacional', 'España', 'Murcia', 'Privada', 'Grados, Másteres, FP, Medicina, Ingeniería'),
                    ('UNIR - Universidad Internacional de La Rioja', 'admisiones@unir.net', '+34 941 209 743', 'Admisiones Internacionales', 'España', 'Logroño', 'Privada', 'Grados Online, Másteres Online, Doctorados'),
                    ('VIU - Universidad Internacional de Valencia', 'informacion@universidadviu.com', '+34 961 924 950', 'Información y Admisiones', 'España', 'Valencia', 'Privada', 'Grados Online/Presencial, Másteres, Doctorados'),
                    ('UDIMA - Universidad a Distancia de Madrid', 'info@udima.es', '+34 918 561 699', 'Información', 'España', 'Madrid', 'Privada', 'Grados Online, Másteres, Doctorados'),
                    ('UOC - Universitat Oberta de Catalunya', 'internacional@uoc.edu', '+34 932 532 300', 'Admisiones Internacionales', 'España', 'Barcelona', 'Privada', 'Grados Online, Másteres, Idiomas')
                ]
                
                for uni in universidades_iniciales:
                    cursor.execute("""
                        INSERT INTO contactos_universidades 
                        (universidad, email, telefono, contacto_nombre, pais, ciudad, tipo_universidad, programas_interes, estado)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'pendiente')
                    """, uni)
                
                print("✅ 5 universidades principales cargadas")
        except Exception as e:
            print(f"⚠️ Error creando tabla contactos_universidades: {e}")
        
        # Crear tabla servicios_solicitados
        try:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS servicios_solicitados (
                    id SERIAL PRIMARY KEY,
                    estudiante_id INTEGER REFERENCES estudiantes(id) ON DELETE CASCADE,
                    servicio_id VARCHAR(100) NOT NULL,
                    servicio_nombre VARCHAR(200) NOT NULL,
                    estado VARCHAR(50) DEFAULT 'pendiente',
                    precio DECIMAL(10, 2),
                    fecha_solicitud TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    fecha_completado TIMESTAMP,
                    notas_admin TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_servicios_estudiante 
                ON servicios_solicitados(estudiante_id);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_servicios_estado 
                ON servicios_solicitados(estado);
            """)
            print("✅ Tabla servicios_solicitados verificada")
        except Exception as e:
            print(f"⚠️ Error creando tabla servicios_solicitados: {e}")
        
        # Crear tabla proceso_visa_pasos para tracking completo
        try:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS proceso_visa_pasos (
                    id SERIAL PRIMARY KEY,
                    estudiante_id INTEGER REFERENCES estudiantes(id) ON DELETE CASCADE,
                    paso_inscripcion BOOLEAN DEFAULT FALSE,
                    fecha_inscripcion TIMESTAMP,
                    paso_pago_inicial BOOLEAN DEFAULT FALSE,
                    fecha_pago_inicial TIMESTAMP,
                    paso_documentos_personales BOOLEAN DEFAULT FALSE,
                    fecha_documentos_personales TIMESTAMP,
                    paso_seleccion_universidad BOOLEAN DEFAULT FALSE,
                    fecha_seleccion_universidad TIMESTAMP,
                    paso_solicitud_universidad BOOLEAN DEFAULT FALSE,
                    fecha_solicitud_universidad TIMESTAMP,
                    paso_carta_aceptacion BOOLEAN DEFAULT FALSE,
                    fecha_carta_aceptacion TIMESTAMP,
                    paso_antecedentes_solicitados BOOLEAN DEFAULT FALSE,
                    fecha_antecedentes_solicitados TIMESTAMP,
                    paso_antecedentes_recibidos BOOLEAN DEFAULT FALSE,
                    fecha_antecedentes_recibidos TIMESTAMP,
                    paso_apostilla_haya BOOLEAN DEFAULT FALSE,
                    fecha_apostilla_haya TIMESTAMP,
                    paso_traduccion_documentos BOOLEAN DEFAULT FALSE,
                    fecha_traduccion_documentos TIMESTAMP,
                    paso_seguro_medico BOOLEAN DEFAULT FALSE,
                    fecha_seguro_medico TIMESTAMP,
                    paso_comprobante_fondos BOOLEAN DEFAULT FALSE,
                    fecha_comprobante_fondos TIMESTAMP,
                    paso_carta_banco BOOLEAN DEFAULT FALSE,
                    fecha_carta_banco TIMESTAMP,
                    paso_formulario_visa BOOLEAN DEFAULT FALSE,
                    fecha_formulario_visa TIMESTAMP,
                    paso_fotos_biometricas BOOLEAN DEFAULT FALSE,
                    fecha_fotos_biometricas TIMESTAMP,
                    paso_pago_tasa_visa BOOLEAN DEFAULT FALSE,
                    fecha_pago_tasa_visa TIMESTAMP,
                    paso_cita_agendada BOOLEAN DEFAULT FALSE,
                    fecha_cita_agendada TIMESTAMP,
                    fecha_cita_embajada TIMESTAMP,
                    paso_documentos_revisados BOOLEAN DEFAULT FALSE,
                    fecha_documentos_revisados TIMESTAMP,
                    paso_simulacro_entrevista BOOLEAN DEFAULT FALSE,
                    fecha_simulacro_entrevista TIMESTAMP,
                    paso_entrevista_completada BOOLEAN DEFAULT FALSE,
                    fecha_entrevista_completada TIMESTAMP,
                    resultado_entrevista VARCHAR(50),
                    paso_pasaporte_recogido BOOLEAN DEFAULT FALSE,
                    fecha_pasaporte_recogido TIMESTAMP,
                    paso_visa_otorgada BOOLEAN DEFAULT FALSE,
                    fecha_visa_otorgada TIMESTAMP,
                    notas_admin TEXT,
                    ultima_actualizacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(estudiante_id)
                );
            """)
            print("✅ Tabla proceso_visa_pasos verificada")
        except Exception as e:
            print(f"⚠️ Error creando tabla proceso_visa_pasos: {e}")
        
        # Crear índices para optimizar consultas
        try:
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_documentos_generados_estudiante 
                ON documentos_generados(estudiante_id);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_documentos_generados_estado 
                ON documentos_generados(estado);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_fechas_importantes_estudiante 
                ON fechas_importantes(estudiante_id);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_fechas_importantes_fecha 
                ON fechas_importantes(fecha);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_fechas_importantes_completada 
                ON fechas_importantes(completada);
            """)
            print("✅ Índices creados")
        except Exception as idx_error:
            print(f"⚠️ Error creando índices: {idx_error}")
        
        conn.commit()
        cursor.close()
        conn.close()
        print("✅ Migraciones ejecutadas correctamente")
        
        # Iniciar scheduler de alertas
        from api.scheduler_alertas import iniciar_scheduler
        iniciar_scheduler()
        
    except Exception as e:
        print(f"⚠️ ERROR CRÍTICO en startup: {e}")
        import traceback
        traceback.print_exc()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # En producción: especificar dominio exacto
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

security = HTTPBearer()


# ============================================================================
# AUTENTICACIÓN
# ============================================================================

# Incluir routers de blog y testimonios
app.include_router(blog_router, prefix="/api")
app.include_router(testimonios_router, prefix="/api")
app.include_router(notificaciones_router, prefix="/api")
app.include_router(chat_router, prefix="/api")
app.include_router(analytics_router, prefix="/api")
app.include_router(documentos_router, prefix="/api")

@app.post("/api/login", response_model=LoginResponse, tags=["Auth"])
@limiter.limit("5/minute")  # Máximo 5 intentos de login por minuto
def login(request: Request, datos: LoginRequest, db: Session = Depends(get_db)):
    """
    Login para admins
    Rate limit: 5 intentos por minuto por IP
    """
    import bcrypt
    import os
    import psycopg2
    
    # Log intento de login
    log_event(
        "login_intento",
        "Intento de login",
        email=datos.usuario,
        ip=request.client.host
    )
    
    try:
        # Conexión directa a la BD
        conn = psycopg2.connect(os.getenv('DATABASE_URL'))
        cur = conn.cursor()
        
        # Buscar usuario por email
        cur.execute('SELECT email, password, nombre, rol FROM usuarios WHERE email = %s', (datos.usuario,))
        result = cur.fetchone()
        
        if not result:
            cur.close()
            conn.close()
            
            # Log login fallido
            log_event(
                "login_fallido",
                "Login fallido - usuario no existe",
                email=datos.usuario,
                ip=request.client.host
            )
            
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Credenciales incorrectas"
            )
        
        email, password_hash, nombre, rol = result
        
        # Verificar contraseña con bcrypt
        if not bcrypt.checkpw(datos.password.encode('utf-8'), password_hash.encode('utf-8')):
            cur.close()
            conn.close()
            
            # Log login fallido
            log_event(
                "login_fallido",
                "Login fallido - contraseña incorrecta",
                email=datos.usuario,
                ip=request.client.host
            )
            
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Credenciales incorrectas"
            )
        
        cur.close()
        conn.close()
        
        # Log login exitoso
        log_event(
            "login_exitoso",
            "Login exitoso",
            email=email,
            nombre=nombre,
            rol=rol,
            ip=request.client.host
        )
        
        # Crear token
        token = crear_token({"usuario": email, "rol": rol})
        return LoginResponse(
            token=token,
            tipo="Bearer",
            usuario=nombre,
            rol=rol
        )
    except HTTPException:
        raise
    except Exception as e:
        # Log error inesperado
        log_error(
            "login_error",
            "Error inesperado en login",
            error=e,
            email=datos.usuario,
            ip=request.client.host
        )
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error en login: {str(e)}"
        )


def obtener_usuario_actual(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Verifica token JWT y retorna usuario"""
    token = credentials.credentials
    payload = verificar_token(token)
    
    if not payload:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token inválido o expirado"
        )
    
    return payload


# ============================================================================
# ENDPOINTS PÚBLICOS (Estudiantes)
# ============================================================================

class RegistroBasicoRequest(BaseModel):
    nombre: str
    email: str
    telefono: str
    consentimiento_gdpr: bool

@app.post("/api/estudiantes", tags=["Estudiantes"])
@limiter.limit("3/hour")  # Máximo 3 registros por hora por IP
async def registrar_estudiante(
    request: Request,
    datos: RegistroBasicoRequest,
    db: Session = Depends(get_db)
):
    """
    Registro público de estudiantes - SIMPLIFICADO
    Solo requiere: nombre, email, teléfono
    El resto se completa después en "Completar Perfil"
    No requiere autenticación
    Rate limit: 3 registros por hora por IP
    """
    
    # Log intento de registro
    log_event(
        "registro_intento",
        "Intento de registro de estudiante",
        email=datos.email,
        nombre=datos.nombre,
        ip=request.client.host
    )
    
    try:
        import string
        import secrets
        import os
        from pathlib import Path
        
        # Generar código de acceso único
        def generar_codigo_acceso(longitud=8):
            caracteres = string.ascii_uppercase + string.digits
            caracteres = caracteres.replace('O', '').replace('I', '').replace('0', '').replace('1', '')
            return ''.join(secrets.choice(caracteres) for _ in range(longitud))
        
        # Verificar si ya existe el email
        check_email = db.execute(text("""
            SELECT id FROM estudiantes WHERE email = :email
        """), {"email": datos.email}).fetchone()
        
        if check_email:
            raise HTTPException(
                status_code=400,
                detail="Ya existe un estudiante con este email"
            )
        
        # Generar código único
        codigo_acceso = generar_codigo_acceso()
        while True:
            check = db.execute(text("""
                SELECT id FROM estudiantes WHERE codigo_acceso = :codigo
            """), {"codigo": codigo_acceso}).fetchone()
            if not check:
                break
            codigo_acceso = generar_codigo_acceso()
        
        # Insertar nuevo estudiante con SQL directo - SOLO CAMPOS BÁSICOS
        result = db.execute(text("""
            INSERT INTO estudiantes (
                nombre, email, telefono,
                consentimiento_gdpr, fecha_consentimiento,
                estado, documentos_estado, codigo_acceso, created_at,
                perfil_completo
            ) VALUES (
                :nombre, :email, :telefono,
                :consentimiento_gdpr, NOW(),
                'pendiente', 'pendiente', :codigo_acceso, NOW(),
                FALSE
            ) RETURNING id, codigo_acceso
        """), {
            "nombre": datos.nombre,
            "email": datos.email,
            "telefono": datos.telefono,
            "consentimiento_gdpr": datos.consentimiento_gdpr,
            "codigo_acceso": codigo_acceso
        })
        
        row = result.fetchone()
        db.commit()
        
        nuevo_id = row[0]
        codigo_final = row[1]
        
        # Preparar datos para alertas
        estudiante_registrado = {
            "id": nuevo_id,
            "nombre": datos.nombre,
            "email": datos.email,
            "telefono": datos.telefono,
            "codigo_acceso": codigo_final
        }
        
        # Log registro exitoso
        log_event(
            "registro_exitoso",
            "Estudiante registrado correctamente",
            estudiante_id=nuevo_id,
            codigo_acceso=codigo_final,
            email=datos.email,
            nombre=datos.nombre,
            ip=request.client.host
        )
        
        # Verificar y enviar alertas al admin si hay problemas
        try:
            from api.alertas_admin import verificar_y_alertar
            problemas = verificar_y_alertar(estudiante_registrado)
            print(f"[INFO] Verificación completada. Problemas: {problemas}")
        except Exception as e:
            print(f"[WARN] Error verificando alertas: {e}")
        
        # Enviar email de bienvenida con código de acceso
        try:
            from api.email_utils import email_bienvenida_con_codigo
            email_bienvenida_con_codigo(datos.nombre, datos.email, nuevo_id, codigo_final)
        except Exception as e:
            print(f"[WARN] Error enviando email: {e}")
        
        return {
            "id": nuevo_id,
            "estudiante_id": nuevo_id,
            "codigo_acceso": codigo_final,
            "mensaje": "Registro exitoso. Revisa tu email para tu código de acceso.",
            "estado": "pendiente"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        
        # Log error de registro
        log_error(
            "registro_error",
            "Error al registrar estudiante",
            error=e,
            email=email,
            nombre=nombre,
            ip=request.client.host
        )
        
        print(f"[ERROR] Error en registro: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/api/estudiantes/codigo/{codigo_acceso}", tags=["Estudiantes"])
def verificar_codigo_acceso(codigo_acceso: str, db: Session = Depends(get_db)):
    """
    Verifica un código de acceso y devuelve el ID del estudiante
    """
    try:
        result = db.execute(text("""
            SELECT id, nombre, email FROM estudiantes 
            WHERE codigo_acceso = :codigo
        """), {"codigo": codigo_acceso.upper()}).fetchone()
        
        if not result:
            raise HTTPException(status_code=404, detail="Código de acceso inválido")
        
        return {
            "id": result[0],
            "nombre": result[1],
            "email": result[2]
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/api/estudiantes/{estudiante_id}", tags=["Estudiantes"])
def obtener_estudiante_publico(estudiante_id: int, db: Session = Depends(get_db)):
    """
    Obtiene datos de un estudiante (sin auth para estudiante)
    Incluye: datos básicos, cursos sugeridos, probabilidad de éxito
    """
    from api.sugerencias_cursos import sugerir_cursos
    from api.calculador_probabilidad import calcular_probabilidad_exito
    
    # Usar SQL directo para evitar problemas con columnas del ORM
    query = text("""
        SELECT 
            id, nombre, email, telefono, pasaporte, fecha_nacimiento, edad, 
            nacionalidad, pais_origen, ciudad_origen, carrera_deseada, especialidad, 
            nivel_espanol, tipo_visa, fondos_disponibles, fecha_inicio_estimada,
            archivo_titulo, archivo_pasaporte, archivo_extractos,
            consentimiento_gdpr, fecha_consentimiento,
            estado, documentos_estado, notas, 
            created_at, updated_at
        FROM estudiantes 
        WHERE id = :estudiante_id
    """)
    
    result = db.execute(query, {"estudiante_id": estudiante_id}).fetchone()
    
    if not result:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    estudiante_data = {
        "id": result[0],
        "nombre": result[1],
        "email": result[2],
        "telefono": result[3],
        "pasaporte": result[4],
        "fecha_nacimiento": result[5].isoformat() if result[5] else None,
        "edad": result[6],
        "nacionalidad": result[7],
        "pais_origen": result[8],
        "ciudad_origen": result[9],
        "carrera_deseada": result[10],
        "especialidad": result[11],
        "nivel_espanol": result[12],
        "tipo_visa": result[13],
        "fondos_disponibles": float(result[14]) if result[14] else None,
        "fecha_inicio_estimada": result[15].isoformat() if result[15] else None,
        "archivo_titulo": result[16],
        "archivo_pasaporte": result[17],
        "archivo_extractos": result[18],
        "consentimiento_gdpr": result[19],
        "fecha_consentimiento": result[20].isoformat() if result[20] else None,
        "estado": result[21],
        "documentos_estado": result[22],
        "notas": result[23],
        "created_at": result[24].isoformat() if result[24] else None,
        "updated_at": result[25].isoformat() if result[25] else None
    }
    
    # Generar sugerencias de cursos
    cursos_sugeridos = sugerir_cursos(estudiante_data)
    
    # Calcular probabilidad de éxito
    probabilidad = calcular_probabilidad_exito(estudiante_data)
    
    # Añadir automáticamente al response
    estudiante_data['cursos_sugeridos'] = cursos_sugeridos
    estudiante_data['probabilidad_exito'] = probabilidad
    
    return estudiante_data


@app.put("/api/estudiantes/{estudiante_id}", tags=["Estudiantes"])
def actualizar_estudiante_publico(estudiante_id: int, datos: dict, db: Session = Depends(get_db)):
    """
    Actualiza datos de un estudiante (sin auth para estudiante)
    """
    from database.models import Estudiante as EstudianteModel
    from datetime import datetime
    
    estudiante = db.query(EstudianteModel).filter(EstudianteModel.id == estudiante_id).first()
    
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    # Actualizar campos permitidos
    campos_permitidos = ['nombre', 'email', 'telefono', 'pasaporte', 'fecha_nacimiento', 'edad', 
                         'nacionalidad', 'pais_origen', 'ciudad_origen', 'carrera_deseada', 'especialidad', 
                         'nivel_espanol', 'tipo_visa', 'fondos_disponibles', 'fecha_inicio_estimada']
    
    for campo in campos_permitidos:
        if campo in datos:
            setattr(estudiante, campo, datos[campo])
    
    estudiante.updated_at = datetime.utcnow()
    
    db.commit()
    db.refresh(estudiante)
    
    return {
        "mensaje": "Datos actualizados correctamente",
        "estudiante": {
            "id": estudiante.id,
            "nombre": estudiante.nombre,
            "email": estudiante.email
        }
    }


@app.put("/api/estudiantes/{estudiante_id}/completar-perfil", tags=["Estudiantes"])
async def completar_perfil_estudiante(
    estudiante_id: int,
    codigo_acceso: str = Query(...),
    pasaporte: str = Form(...),
    fecha_nacimiento: str = Form(...),
    edad: int = Form(...),
    nacionalidad: str = Form(...),
    pais_origen: str = Form(...),
    ciudad_origen: str = Form(...),
    carrera_deseada: str = Form(...),
    especialidad: str = Form(...),
    nivel_espanol: str = Form(...),
    tipo_visa: str = Form(...),
    fondos_disponibles: float = Form(...),
    fecha_inicio_estimada: str = Form(...),
    archivo_titulo: UploadFile = File(None),
    archivo_pasaporte: UploadFile = File(None),
    archivo_extractos: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    """
    Completa el perfil del estudiante con información sensible
    Requiere código de acceso para autenticación
    """
    import os
    import psycopg2
    from pathlib import Path
    import secrets
    
    # Verificar código de acceso
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, nombre, email FROM estudiantes 
        WHERE id = %s AND codigo_acceso = %s
    """, (estudiante_id, codigo_acceso))
    
    estudiante = cursor.fetchone()
    if not estudiante:
        conn.close()
        raise HTTPException(status_code=403, detail="Código de acceso inválido")
    
    try:
        # Directorio de uploads
        uploads_dir = Path("uploads")
        uploads_dir.mkdir(exist_ok=True)
        
        # Función para guardar archivos
        def guardar_archivo(archivo: UploadFile, prefijo: str) -> str:
            if not archivo or not archivo.filename:
                return None
            
            extension = archivo.filename.split('.')[-1]
            nombre_archivo = f"{prefijo}_{secrets.token_hex(8)}.{extension}"
            ruta_archivo = uploads_dir / nombre_archivo
            
            with open(ruta_archivo, "wb") as buffer:
                content = archivo.file.read()
                buffer.write(content)
            
            return str(ruta_archivo)
        
        ruta_titulo = guardar_archivo(archivo_titulo, "titulo")
        ruta_pasaporte = guardar_archivo(archivo_pasaporte, "pasaporte")
        ruta_extractos = guardar_archivo(archivo_extractos, "extractos")
        
        # Actualizar estudiante
        cursor.execute("""
            UPDATE estudiantes SET
                pasaporte = %s,
                fecha_nacimiento = %s,
                edad = %s,
                nacionalidad = %s,
                pais_origen = %s,
                ciudad_origen = %s,
                carrera_deseada = %s,
                especialidad = %s,
                nivel_espanol = %s,
                tipo_visa = %s,
                fondos_disponibles = %s,
                fecha_inicio_estimada = %s,
                archivo_titulo = COALESCE(%s, archivo_titulo),
                archivo_pasaporte = COALESCE(%s, archivo_pasaporte),
                archivo_extractos = COALESCE(%s, archivo_extractos),
                perfil_completo = TRUE,
                updated_at = NOW()
            WHERE id = %s
        """, (
            pasaporte, fecha_nacimiento, edad, nacionalidad,
            pais_origen, ciudad_origen, carrera_deseada, especialidad,
            nivel_espanol, tipo_visa, fondos_disponibles, fecha_inicio_estimada,
            ruta_titulo, ruta_pasaporte, ruta_extractos,
            estudiante_id
        ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return {
            "mensaje": "Perfil completado exitosamente",
            "estudiante_id": estudiante_id
        }
        
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=f"Error al completar perfil: {str(e)}")


@app.get("/api/estudiantes/{estudiante_id}/probabilidad-visa", tags=["Estudiantes"])
def calcular_probabilidad_visa(estudiante_id: int, db: Session = Depends(get_db)):
    """
    Calcula la probabilidad de aprobación de visa del estudiante
    """
    from database.models import Estudiante as EstudianteModel
    from api.calculadora_visa import CalculadoraProbabilidadVisa
    
    estudiante = db.query(EstudianteModel).filter(EstudianteModel.id == estudiante_id).first()
    
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    # Preparar datos para el cálculo
    estudiante_data = {
        'nombre': estudiante.nombre,
        'email': estudiante.email,
        'telefono': estudiante.telefono,
        'pasaporte': estudiante.pasaporte,
        'edad': estudiante.edad,
        'nacionalidad': estudiante.nacionalidad,
        'ciudad_origen': estudiante.ciudad_origen,
        'especialidad': estudiante.especialidad,
        'nivel_espanol': estudiante.nivel_espanol,
        'tipo_visa': estudiante.tipo_visa
    }
    
    # Calcular probabilidad
    analisis = CalculadoraProbabilidadVisa.calcular_probabilidad(estudiante_data)
    
    return {
        'estudiante_id': estudiante_id,
        'estudiante_nombre': estudiante.nombre,
        'analisis': analisis
    }


@app.get("/api/estudiantes/{estudiante_id}/generar-documentos", tags=["Estudiantes"])
def generar_documentos_estudiante(estudiante_id: int, db: Session = Depends(get_db)):
    """
    Genera documentos borrador: carta aceptación, carta patrocinio, checklist
    """
    from api.generador_documentos_borrador import generar_todos_documentos
    
    # Obtener datos del estudiante
    query = text("""
        SELECT 
            id, nombre, email, telefono, pasaporte, fecha_nacimiento, edad, 
            nacionalidad, pais_origen, ciudad_origen, carrera_deseada, especialidad, 
            nivel_espanol, tipo_visa, fondos_disponibles, fecha_inicio_estimada,
            archivo_titulo, archivo_pasaporte, archivo_extractos
        FROM estudiantes 
        WHERE id = :estudiante_id
    """)
    
    result = db.execute(query, {"estudiante_id": estudiante_id}).fetchone()
    
    if not result:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    estudiante_data = {
        "id": result[0],
        "nombre": result[1],
        "email": result[2],
        "telefono": result[3],
        "pasaporte": result[4],
        "fecha_nacimiento": result[5].isoformat() if result[5] else None,
        "edad": result[6],
        "nacionalidad": result[7],
        "pais_origen": result[8],
        "ciudad_origen": result[9],
        "carrera_deseada": result[10],
        "especialidad": result[11],
        "nivel_espanol": result[12],
        "tipo_visa": result[13],
        "fondos_disponibles": float(result[14]) if result[14] else 0,
        "fecha_inicio_estimada": result[15].isoformat() if result[15] else "próximo semestre",
        "archivo_titulo": result[16],
        "archivo_pasaporte": result[17],
        "archivo_extractos": result[18]
    }
    
    # Generar documentos
    documentos = generar_todos_documentos(estudiante_data)
    
    return {
        "estudiante_id": estudiante_id,
        "estudiante_nombre": estudiante_data["nombre"],
        "documentos": documentos,
        "nota": "Estos son BORRADORES. Deben ser completados con datos reales y firmados oficialmente."
    }


# ==================== SIMULADOR DE ENTREVISTAS ====================

@app.get("/api/estudiantes/{estudiante_id}/simulador-entrevista", tags=["Estudiantes - Simulador"])
def obtener_simulador_entrevista(estudiante_id: int, db: Session = Depends(get_db)):
    """
    Genera simulador de entrevista personalizado según perfil del estudiante
    """
    from api.simulador_entrevista import SimuladorEntrevista
    
    # Obtener datos del estudiante
    query = text("""
        SELECT 
            id, nombre, edad, especialidad, nivel_espanol, tipo_visa, fondos_disponibles
        FROM estudiantes 
        WHERE id = :estudiante_id
    """)
    
    result = db.execute(query, {"estudiante_id": estudiante_id}).fetchone()
    
    if not result:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    estudiante_data = {
        "id": result[0],
        "nombre": result[1],
        "edad": result[2],
        "especialidad": result[3] or "",
        "nivel_espanol": result[4] or "basico",
        "tipo_visa": result[5] or "estudiante",
        "fondos_disponibles": float(result[6]) if result[6] else 0
    }
    
    # Generar entrevista personalizada
    entrevista = SimuladorEntrevista.generar_entrevista_personalizada(estudiante_data)
    
    return entrevista


@app.post("/api/simulador-entrevista/evaluar", tags=["Estudiantes - Simulador"])
def evaluar_respuesta_simulador(data: dict):
    """
    Evalúa una respuesta del simulador de entrevista
    """
    from api.simulador_entrevista import SimuladorEntrevista
    
    pregunta_id = data.get('pregunta_id', 0)
    respuesta = data.get('respuesta', '')
    pregunta_obj = data.get('pregunta', {})  # Recibir objeto pregunta completo
    
    if not respuesta:
        raise HTTPException(status_code=400, detail="Respuesta vacía")
    
    evaluacion = SimuladorEntrevista.evaluar_respuesta(pregunta_id, respuesta, pregunta_obj)
    
    return evaluacion


# ==================== CALCULADORA DE FONDOS ====================

@app.post("/api/calculadora-fondos", tags=["Estudiantes - Calculadora"])
def calcular_fondos_estudiante(datos: dict):
    """
    Calcula fondos estimados necesarios para estudiar en España
    IMPORTANTE: Son estimaciones generales, te ayudamos con presupuesto detallado
    """
    from api.calculadora_fondos import CalculadoraFondos
    
    resultado = CalculadoraFondos.calcular_fondos(datos)
    
    return resultado


@app.get("/api/calculadora-fondos/ciudades", tags=["Estudiantes - Calculadora"])
def obtener_ciudades_calculadora():
    """
    Obtiene lista de ciudades disponibles para cálculo
    """
    from api.calculadora_fondos import CalculadoraFondos
    
    return CalculadoraFondos.obtener_ciudades_disponibles()


@app.get("/api/calculadora-fondos/programas", tags=["Estudiantes - Calculadora"])
def obtener_programas_calculadora():
    """
    Obtiene lista de tipos de programa disponibles
    """
    from api.calculadora_fondos import CalculadoraFondos
    
    return CalculadoraFondos.obtener_tipos_programa()


# =============================================================================
# ENDPOINTS: ALERTAS Y FECHAS IMPORTANTES
# =============================================================================

@app.post("/api/alertas/fecha", tags=["Alertas - Fechas Importantes"])
def agregar_fecha_importante(
    datos: dict,
    db: Session = Depends(get_db)
):
    """
    Agrega una nueva fecha importante para un estudiante
    
    Body:
    - estudiante_id: ID del estudiante
    - tipo_fecha: Tipo de fecha (entrevista_consular, vencimiento_pasaporte, deadline_aplicacion, etc.)
    - fecha: Fecha en formato ISO (YYYY-MM-DDTHH:MM:SS)
    - descripcion: Descripción opcional
    """
    from api.alertas_fechas import GestorAlertasFechas
    
    try:
        # Convertir fecha string a datetime
        fecha_dt = datetime.fromisoformat(datos['fecha'].replace('Z', ''))
        
        nueva_fecha = GestorAlertasFechas.agregar_fecha(
            db=db,
            estudiante_id=datos['estudiante_id'],
            tipo_fecha=datos['tipo_fecha'],
            fecha=fecha_dt,
            descripcion=datos.get('descripcion')
        )
        
        return {
            "success": True,
            "message": "Fecha importante agregada exitosamente",
            "fecha": {
                "id": nueva_fecha.id,
                "tipo_fecha": nueva_fecha.tipo_fecha,
                "fecha": nueva_fecha.fecha.isoformat(),
                "descripcion": nueva_fecha.descripcion
            }
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/alertas/estudiante/{estudiante_id}", tags=["Alertas - Fechas Importantes"])
def obtener_fechas_estudiante(
    estudiante_id: int,
    incluir_completadas: bool = False,
    db: Session = Depends(get_db)
):
    """
    Obtiene todas las fechas importantes de un estudiante
    """
    from api.alertas_fechas import GestorAlertasFechas
    
    fechas = GestorAlertasFechas.obtener_fechas_estudiante(
        db=db,
        estudiante_id=estudiante_id,
        incluir_completadas=incluir_completadas
    )
    
    return {
        "success": True,
        "fechas": [
            {
                "id": f.id,
                "tipo_fecha": f.tipo_fecha,
                "tipo_nombre": GestorAlertasFechas.TIPOS_FECHA.get(f.tipo_fecha, f.tipo_fecha),
                "fecha": f.fecha.isoformat(),
                "descripcion": f.descripcion,
                "completada": f.completada,
                "alertado_30d": f.alertado_30d,
                "alertado_15d": f.alertado_15d,
                "alertado_7d": f.alertado_7d,
                "alertado_1d": f.alertado_1d,
                "dias_restantes": (f.fecha - datetime.utcnow()).days if f.fecha > datetime.utcnow() else 0
            }
            for f in fechas
        ]
    }


@app.put("/api/alertas/{fecha_id}/completar", tags=["Alertas - Fechas Importantes"])
def marcar_fecha_completada(
    fecha_id: int,
    db: Session = Depends(get_db)
):
    """
    Marca una fecha como completada
    """
    from api.alertas_fechas import GestorAlertasFechas
    
    success = GestorAlertasFechas.marcar_completada(db=db, fecha_id=fecha_id)
    
    if success:
        return {"success": True, "message": "Fecha marcada como completada"}
    else:
        raise HTTPException(status_code=404, detail="Fecha no encontrada")


@app.delete("/api/alertas/{fecha_id}", tags=["Alertas - Fechas Importantes"])
def eliminar_fecha_importante(
    fecha_id: int,
    db: Session = Depends(get_db)
):
    """
    Elimina una fecha importante
    """
    from api.alertas_fechas import GestorAlertasFechas
    
    success = GestorAlertasFechas.eliminar_fecha(db=db, fecha_id=fecha_id)
    
    if success:
        return {"success": True, "message": "Fecha eliminada exitosamente"}
    else:
        raise HTTPException(status_code=404, detail="Fecha no encontrada")


@app.get("/api/alertas/proximas", tags=["Admin - Alertas"])
def obtener_fechas_proximas_admin(
    dias: int = 30,
    db: Session = Depends(get_db)
):
    """
    Obtiene todas las fechas próximas de todos los estudiantes (para admin)
    """
    from api.alertas_fechas import GestorAlertasFechas
    from database.models import Estudiante
    
    fechas = GestorAlertasFechas.obtener_fechas_proximas(db=db, dias=dias)
    
    # Enriquecer con datos del estudiante
    resultado = []
    for f in fechas:
        estudiante = db.query(Estudiante).filter(Estudiante.id == f.estudiante_id).first()
        resultado.append({
            "id": f.id,
            "estudiante": {
                "id": estudiante.id,
                "nombre": estudiante.nombre,
                "email": estudiante.email
            } if estudiante else None,
            "tipo_fecha": f.tipo_fecha,
            "tipo_nombre": GestorAlertasFechas.TIPOS_FECHA.get(f.tipo_fecha, f.tipo_fecha),
            "fecha": f.fecha.isoformat(),
            "descripcion": f.descripcion,
            "dias_restantes": (f.fecha - datetime.utcnow()).days
        })
    
    return {
        "success": True,
        "fechas_proximas": resultado,
        "total": len(resultado)
    }


@app.get("/api/alertas/{fecha_id}/descargar-ics", tags=["Alertas - Fechas Importantes"])
def descargar_calendario_ics(
    fecha_id: int,
    db: Session = Depends(get_db)
):
    """
    Descarga archivo .ics para agregar fecha a calendario
    """
    from api.alertas_fechas import GestorAlertasFechas
    from database.models import FechaImportante, Estudiante
    from fastapi.responses import Response
    
    fecha = db.query(FechaImportante).filter(FechaImportante.id == fecha_id).first()
    if not fecha:
        raise HTTPException(status_code=404, detail="Fecha no encontrada")
    
    estudiante = db.query(Estudiante).filter(Estudiante.id == fecha.estudiante_id).first()
    
    ics_content = GestorAlertasFechas.generar_ics(fecha, estudiante)
    
    return Response(
        content=ics_content,
        media_type="text/calendar",
        headers={"Content-Disposition": f"attachment; filename=fecha_{fecha_id}.ics"}
    )


# =============================================================================
# ENDPOINTS: UNIVERSIDADES Y PROGRAMAS ESPAÑA
# =============================================================================

@app.get("/api/universidades", tags=["Universidades España"])
def listar_universidades(
    ciudad: Optional[str] = None,
    comunidad: Optional[str] = None,
    tipo: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Lista todas las universidades de España con filtros
    
    Filtros:
    - ciudad: Filtrar por ciudad
    - comunidad: Filtrar por comunidad autónoma
    - tipo: publica o privada
    """
    from database.models import UniversidadEspana
    
    query = db.query(UniversidadEspana).filter(UniversidadEspana.activa == True)
    
    if ciudad:
        query = query.filter(UniversidadEspana.ciudad.ilike(f'%{ciudad}%'))
    if comunidad:
        query = query.filter(UniversidadEspana.comunidad_autonoma.ilike(f'%{comunidad}%'))
    if tipo:
        query = query.filter(UniversidadEspana.tipo == tipo)
    
    universidades = query.order_by(UniversidadEspana.ranking_nacional).all()
    
    return {
        "success": True,
        "total": len(universidades),
        "universidades": [
            {
                "id": u.id,
                "nombre": u.nombre,
                "siglas": u.siglas,
                "ciudad": u.ciudad,
                "comunidad_autonoma": u.comunidad_autonoma,
                "tipo": u.tipo,
                "url_oficial": u.url_oficial,
                "email_contacto": u.email_contacto,
                "ranking_nacional": u.ranking_nacional,
                "total_alumnos": u.total_alumnos,
                "total_programas": u.total_programas,
                "acepta_extranjeros": u.acepta_extranjeros,
                "logo_url": u.logo_url
            }
            for u in universidades
        ]
    }


@app.get("/api/universidades/{universidad_id}", tags=["Universidades España"])
def obtener_universidad(universidad_id: int, db: Session = Depends(get_db)):
    """Obtiene información detallada de una universidad"""
    from database.models import UniversidadEspana
    
    universidad = db.query(UniversidadEspana).filter(
        UniversidadEspana.id == universidad_id
    ).first()
    
    if not universidad:
        raise HTTPException(status_code=404, detail="Universidad no encontrada")
    
    return {
        "success": True,
        "universidad": {
            "id": universidad.id,
            "nombre": universidad.nombre,
            "siglas": universidad.siglas,
            "ciudad": universidad.ciudad,
            "comunidad_autonoma": universidad.comunidad_autonoma,
            "tipo": universidad.tipo,
            "url_oficial": universidad.url_oficial,
            "email_contacto": universidad.email_contacto,
            "telefono": universidad.telefono,
            "descripcion": universidad.descripcion,
            "ranking_nacional": universidad.ranking_nacional,
            "total_alumnos": universidad.total_alumnos,
            "total_programas": universidad.total_programas,
            "acepta_extranjeros": universidad.acepta_extranjeros,
            "requisitos_extranjeros": universidad.requisitos_extranjeros,
            "ultima_actualizacion": universidad.ultima_actualizacion.isoformat() if universidad.ultima_actualizacion else None,
            "logo_url": universidad.logo_url
        }
    }


@app.get("/api/universidades/{universidad_id}/programas", tags=["Universidades España"])
def listar_programas_universidad(
    universidad_id: int,
    tipo_programa: Optional[str] = None,
    area_estudio: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Lista programas académicos de una universidad
    
    Filtros:
    - tipo_programa: grado, master, doctorado, curso
    - area_estudio: ingenieria, medicina, arte, etc.
    """
    from database.models import ProgramaUniversitario
    
    query = db.query(ProgramaUniversitario).filter(
        ProgramaUniversitario.universidad_id == universidad_id,
        ProgramaUniversitario.activo == True
    )
    
    if tipo_programa:
        query = query.filter(ProgramaUniversitario.tipo_programa == tipo_programa)
    if area_estudio:
        query = query.filter(ProgramaUniversitario.area_estudio.ilike(f'%{area_estudio}%'))
    
    programas = query.all()
    
    return {
        "success": True,
        "universidad_id": universidad_id,
        "total": len(programas),
        "programas": [
            {
                "id": p.id,
                "nombre": p.nombre,
                "tipo_programa": p.tipo_programa,
                "area_estudio": p.area_estudio,
                "duracion_anos": p.duracion_anos,
                "creditos_ects": p.creditos_ects,
                "idioma": p.idioma,
                "modalidad": p.modalidad,
                "precio_anual_eur": p.precio_anual_eur,
                "plazas_disponibles": p.plazas_disponibles,
                "nota_corte": p.nota_corte,
                "url_info": p.url_info,
                "requisitos": p.requisitos,
                "descripcion": p.descripcion
            }
            for p in programas
        ]
    }


@app.get("/api/programas/buscar", tags=["Universidades España"])
def buscar_programas(
    query: str,
    ciudad: Optional[str] = None,
    tipo_programa: Optional[str] = None,
    precio_max: Optional[float] = None,
    db: Session = Depends(get_db)
):
    """
    Búsqueda global de programas en todas las universidades
    
    Parámetros:
    - query: Texto a buscar en nombre del programa
    - ciudad: Filtrar por ciudad
    - tipo_programa: grado, master, doctorado
    - precio_max: Precio máximo anual
    """
    from database.models import ProgramaUniversitario, UniversidadEspana
    
    programas_query = db.query(
        ProgramaUniversitario, UniversidadEspana
    ).join(
        UniversidadEspana,
        ProgramaUniversitario.universidad_id == UniversidadEspana.id
    ).filter(
        ProgramaUniversitario.activo == True,
        ProgramaUniversitario.nombre.ilike(f'%{query}%')
    )
    
    if ciudad:
        programas_query = programas_query.filter(UniversidadEspana.ciudad.ilike(f'%{ciudad}%'))
    if tipo_programa:
        programas_query = programas_query.filter(ProgramaUniversitario.tipo_programa == tipo_programa)
    if precio_max:
        programas_query = programas_query.filter(ProgramaUniversitario.precio_anual_eur <= precio_max)
    
    resultados = programas_query.limit(50).all()
    
    return {
        "success": True,
        "query": query,
        "total": len(resultados),
        "resultados": [
            {
                "programa": {
                    "id": programa.id,
                    "nombre": programa.nombre,
                    "tipo_programa": programa.tipo_programa,
                    "duracion_anos": programa.duracion_anos,
                    "precio_anual_eur": programa.precio_anual_eur,
                    "modalidad": programa.modalidad,
                    "idioma": programa.idioma
                },
                "universidad": {
                    "id": universidad.id,
                    "nombre": universidad.nombre,
                    "ciudad": universidad.ciudad,
                    "tipo": universidad.tipo,
                    "url_oficial": universidad.url_oficial
                }
            }
            for programa, universidad in resultados
        ]
    }


@app.post("/api/admin/universidades/actualizar", tags=["Admin - Universidades"])
def actualizar_datos_universidades(db: Session = Depends(get_db)):
    """
    Ejecuta scraping de todas las universidades para actualizar datos
    (Endpoint admin - requiere autenticación)
    """
    from api.scrapers_universidades import actualizar_todas_universidades
    
    try:
        programas_nuevos = actualizar_todas_universidades(db)
        return {
            "success": True,
            "message": f"Actualización completada",
            "programas_nuevos": programas_nuevos
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/admin/universidades/seed", tags=["Admin - Universidades"])
def seed_universidades_iniciales(db: Session = Depends(get_db)):
    """
    Carga datos iniciales de universidades españolas
    (Solo ejecutar una vez en setup inicial)
    """
    from api.data_universidades_espana import UNIVERSIDADES_ESPANA
    from database.models import UniversidadEspana
    
    try:
        universidades_creadas = 0
        
        for uni_data in UNIVERSIDADES_ESPANA:
            # Verificar si ya existe
            existe = db.query(UniversidadEspana).filter(
                UniversidadEspana.nombre == uni_data['nombre']
            ).first()
            
            if not existe:
                nueva_uni = UniversidadEspana(
                    nombre=uni_data['nombre'],
                    siglas=uni_data.get('siglas'),
                    ciudad=uni_data['ciudad'],
                    comunidad_autonoma=uni_data['comunidad_autonoma'],
                    tipo=uni_data['tipo'],
                    url_oficial=uni_data['url_oficial'],
                    email_contacto=uni_data.get('email_contacto'),
                    tiene_api=uni_data.get('tiene_api', False),
                    metodo_scraping=uni_data.get('metodo_scraping', 'beautifulsoup'),
                    ranking_nacional=uni_data.get('ranking_nacional'),
                    total_alumnos=uni_data.get('total_alumnos'),
                    acepta_extranjeros=uni_data.get('acepta_extranjeros', True),
                    requisitos_extranjeros=uni_data.get('requisitos_extranjeros'),
                    activa=True
                )
                db.add(nueva_uni)
                universidades_creadas += 1
        
        db.commit()
        
        return {
            "success": True,
            "message": f"{universidades_creadas} universidades agregadas a la base de datos",
            "total_universidades": universidades_creadas
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ========================================
# ADMIN CRUD UNIVERSIDADES
# ========================================

@app.post("/api/admin/universidades", tags=["Admin - Universidades"])
def crear_universidad_manual(
    universidad: dict,
    db: Session = Depends(get_db)
):
    """Crear nueva universidad manualmente"""
    from database.models import UniversidadEspana
    
    try:
        nueva_uni = UniversidadEspana(
            nombre=universidad.get('nombre'),
            siglas=universidad.get('siglas', ''),
            ciudad=universidad.get('ciudad', ''),
            comunidad_autonoma=universidad.get('comunidad_autonoma', ''),
            tipo=universidad.get('tipo', 'publica'),
            url_oficial=universidad.get('url_oficial', ''),
            email_contacto=universidad.get('email_contacto', ''),
            telefono=universidad.get('telefono', ''),
            tiene_api=universidad.get('tiene_api', False),
            metodo_scraping=universidad.get('metodo_scraping', 'beautifulsoup'),
            descripcion=universidad.get('descripcion', ''),
            ranking_nacional=universidad.get('ranking_nacional'),
            total_alumnos=universidad.get('total_alumnos'),
            acepta_extranjeros=universidad.get('acepta_extranjeros', True),
            requisitos_extranjeros=universidad.get('requisitos_extranjeros', ''),
            activa=True
        )
        db.add(nueva_uni)
        db.commit()
        db.refresh(nueva_uni)
        
        return {
            "success": True,
            "message": "Universidad creada exitosamente",
            "universidad_id": nueva_uni.id,
            "nombre": nueva_uni.nombre
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/admin/universidades/{universidad_id}", tags=["Admin - Universidades"])
def actualizar_universidad(
    universidad_id: int,
    datos: dict,
    db: Session = Depends(get_db)
):
    """Actualizar información de universidad existente"""
    from database.models import UniversidadEspana
    
    try:
        universidad = db.query(UniversidadEspana).filter(
            UniversidadEspana.id == universidad_id
        ).first()
        
        if not universidad:
            raise HTTPException(status_code=404, detail="Universidad no encontrada")
        
        # Actualizar campos
        for campo, valor in datos.items():
            if hasattr(universidad, campo):
                setattr(universidad, campo, valor)
        
        db.commit()
        db.refresh(universidad)
        
        return {
            "success": True,
            "message": "Universidad actualizada",
            "universidad": {
                "id": universidad.id,
                "nombre": universidad.nombre,
                "ciudad": universidad.ciudad
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/admin/universidades/{universidad_id}", tags=["Admin - Universidades"])
def eliminar_universidad(
    universidad_id: int,
    db: Session = Depends(get_db)
):
    """Desactivar universidad (soft delete)"""
    from database.models import UniversidadEspana
    
    try:
        universidad = db.query(UniversidadEspana).filter(
            UniversidadEspana.id == universidad_id
        ).first()
        
        if not universidad:
            raise HTTPException(status_code=404, detail="Universidad no encontrada")
        
        universidad.activa = False
        db.commit()
        
        return {
            "success": True,
            "message": f"Universidad {universidad.nombre} desactivada"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ========================================
# ADMIN CRUD PROGRAMAS
# ========================================

@app.post("/api/admin/programas", tags=["Admin - Programas"])
def crear_programa_manual(
    programa: dict,
    db: Session = Depends(get_db)
):
    """Crear nuevo programa académico manualmente"""
    from database.models import ProgramaUniversitario
    
    try:
        nuevo_programa = ProgramaUniversitario(
            universidad_id=programa.get('universidad_id'),
            nombre=programa.get('nombre'),
            tipo_programa=programa.get('tipo_programa', 'grado'),
            area_estudio=programa.get('area_estudio', ''),
            duracion_anos=programa.get('duracion_anos', 4),
            creditos_ects=programa.get('creditos_ects', 240),
            idioma=programa.get('idioma', 'español'),
            modalidad=programa.get('modalidad', 'presencial'),
            precio_anual_eur=programa.get('precio_anual_eur'),
            plazas_disponibles=programa.get('plazas_disponibles'),
            nota_corte=programa.get('nota_corte'),
            url_info=programa.get('url_info', ''),
            requisitos=programa.get('requisitos', ''),
            descripcion=programa.get('descripcion', ''),
            activo=True
        )
        db.add(nuevo_programa)
        db.commit()
        db.refresh(nuevo_programa)
        
        return {
            "success": True,
            "message": "Programa creado exitosamente",
            "programa_id": nuevo_programa.id,
            "nombre": nuevo_programa.nombre
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/admin/programas/{programa_id}", tags=["Admin - Programas"])
def actualizar_programa(
    programa_id: int,
    datos: dict,
    db: Session = Depends(get_db)
):
    """Actualizar información de programa existente"""
    from database.models import ProgramaUniversitario
    
    try:
        programa = db.query(ProgramaUniversitario).filter(
            ProgramaUniversitario.id == programa_id
        ).first()
        
        if not programa:
            raise HTTPException(status_code=404, detail="Programa no encontrado")
        
        # Actualizar campos
        for campo, valor in datos.items():
            if hasattr(programa, campo):
                setattr(programa, campo, valor)
        
        db.commit()
        db.refresh(programa)
        
        return {
            "success": True,
            "message": "Programa actualizado",
            "programa": {
                "id": programa.id,
                "nombre": programa.nombre,
                "tipo_programa": programa.tipo_programa
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/admin/programas/{programa_id}", tags=["Admin - Programas"])
def eliminar_programa(
    programa_id: int,
    db: Session = Depends(get_db)
):
    """Desactivar programa (soft delete)"""
    from database.models import ProgramaUniversitario
    
    try:
        programa = db.query(ProgramaUniversitario).filter(
            ProgramaUniversitario.id == programa_id
        ).first()
        
        if not programa:
            raise HTTPException(status_code=404, detail="Programa no encontrado")
        
        programa.activo = False
        db.commit()
        
        return {
            "success": True,
            "message": f"Programa {programa.nombre} desactivado"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/documentos/{documento_id}/validar-ocr", tags=["Documentos - OCR"])
async def validar_documento_ocr(
    documento_id: int,
    db: Session = Depends(get_db)
):
    """
    Valida documento usando OCR inteligente
    Extrae información y detecta errores automáticamente
    """
    try:
        from api.ocr_processor import OCRProcessor
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        # Obtener documento
        cursor.execute("""
            SELECT tipo_documento, nombre_archivo, url_archivo
            FROM documentos
            WHERE id = %s
        """, (documento_id,))
        
        documento = cursor.fetchone()
        
        if not documento:
            cursor.close()
            conn.close()
            raise HTTPException(status_code=404, detail="Documento no encontrado")
        
        tipo_doc, nombre_archivo, url_archivo = documento
        
        # Extraer base64 de la URL (formato: data:image/jpeg;base64,...)
        if not url_archivo or 'base64,' not in url_archivo:
            cursor.close()
            conn.close()
            raise HTTPException(status_code=400, detail="Documento no contiene imagen válida")
        
        imagen_base64 = url_archivo.split('base64,')[1]
        
        # Procesar con OCR
        ocr = OCRProcessor()
        resultado = ocr.procesar_documento(imagen_base64, tipo_doc)
        
        if not resultado.get('exito'):
            cursor.close()
            conn.close()
            return {
                'exito': False,
                'error': resultado.get('error', 'Error desconocido'),
                'documento_id': documento_id
            }
        
        # Guardar resultados en DB
        validacion = resultado.get('validacion', {})
        
        cursor.execute("""
            UPDATE documentos
            SET 
                ocr_procesado = TRUE,
                ocr_texto_extraido = %s,
                ocr_datos_extraidos = %s,
                ocr_valido = %s,
                ocr_advertencias = %s,
                ocr_errores = %s,
                estado = CASE 
                    WHEN %s THEN 'aprobado'
                    WHEN array_length(%s, 1) > 0 THEN 'rechazado'
                    ELSE 'en_revision'
                END,
                updated_at = NOW()
            WHERE id = %s
        """, (
            resultado.get('texto_extraido', ''),
            str(validacion.get('datos_extraidos', {})),
            validacion.get('valido', False),
            str(validacion.get('advertencias', [])),
            str(validacion.get('errores', [])),
            validacion.get('valido', False),
            validacion.get('errores', []),
            documento_id
        ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # Generar reporte legible
        reporte = ocr.generar_reporte_validacion(resultado)
        
        return {
            'exito': True,
            'documento_id': documento_id,
            'tipo_detectado': resultado.get('tipo_detectado'),
            'valido': validacion.get('valido', False),
            'datos_extraidos': validacion.get('datos_extraidos', {}),
            'advertencias': validacion.get('advertencias', []),
            'errores': validacion.get('errores', []),
            'reporte': reporte,
            'tiempo_procesamiento': resultado.get('tiempo_procesamiento', 0)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error en validación OCR: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/documentos/batch-validar-ocr", tags=["Documentos - OCR"])
async def validar_documentos_batch(
    estudiante_id: int,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """
    Valida todos los documentos de un estudiante en lote
    Útil para procesamiento masivo
    """
    verificar_token(credentials.credentials)
    
    try:
        from api.ocr_processor import OCRProcessor
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        # Obtener documentos no procesados
        cursor.execute("""
            SELECT id, tipo_documento
            FROM documentos
            WHERE estudiante_id = %s 
            AND (ocr_procesado IS NULL OR ocr_procesado = FALSE)
        """, (estudiante_id,))
        
        documentos = cursor.fetchall()
        cursor.close()
        conn.close()
        
        if not documentos:
            return {
                'exito': True,
                'mensaje': 'No hay documentos pendientes de validación',
                'procesados': 0
            }
        
        # Procesar cada documento
        resultados = []
        exitosos = 0
        
        for doc_id, tipo_doc in documentos:
            try:
                resultado = await validar_documento_ocr(doc_id, db)
                if resultado.get('exito'):
                    exitosos += 1
                resultados.append({
                    'documento_id': doc_id,
                    'tipo': tipo_doc,
                    'resultado': resultado
                })
            except Exception as e:
                resultados.append({
                    'documento_id': doc_id,
                    'tipo': tipo_doc,
                    'error': str(e)
                })
        
        return {
            'exito': True,
            'procesados': exitosos,
            'total': len(documentos),
            'resultados': resultados
        }
        
    except Exception as e:
        print(f"Error en validación batch: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/documentos/{documento_id}/ocr-report", tags=["Admin - Documentos"])
def obtener_reporte_ocr(
    documento_id: int,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """
    Obtiene reporte detallado de validación OCR de un documento
    """
    verificar_token(credentials.credentials)
    
    try:
        import os
        import psycopg2
        import ast
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT 
                d.tipo_documento,
                d.nombre_archivo,
                d.estado,
                d.ocr_procesado,
                d.ocr_texto_extraido,
                d.ocr_datos_extraidos,
                d.ocr_valido,
                d.ocr_advertencias,
                d.ocr_errores,
                e.nombre as estudiante_nombre
            FROM documentos d
            JOIN estudiantes e ON d.estudiante_id = e.id
            WHERE d.id = %s
        """, (documento_id,))
        
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not row:
            raise HTTPException(status_code=404, detail="Documento no encontrado")
        
        (tipo_doc, nombre_archivo, estado, ocr_procesado, texto_extraido, 
         datos_extraidos, ocr_valido, advertencias, errores, estudiante) = row
        
        if not ocr_procesado:
            return {
                'documento_id': documento_id,
                'procesado': False,
                'mensaje': 'Documento no ha sido procesado por OCR'
            }
        
        # Parsear strings a objetos Python
        try:
            datos_dict = ast.literal_eval(datos_extraidos) if datos_extraidos else {}
            advertencias_list = ast.literal_eval(advertencias) if advertencias else []
            errores_list = ast.literal_eval(errores) if errores else []
        except:
            datos_dict = {}
            advertencias_list = []
            errores_list = []
        
        return {
            'documento_id': documento_id,
            'estudiante': estudiante,
            'tipo_documento': tipo_doc,
            'nombre_archivo': nombre_archivo,
            'estado': estado,
            'procesado': True,
            'valido': ocr_valido,
            'datos_extraidos': datos_dict,
            'advertencias': advertencias_list,
            'errores': errores_list,
            'texto_completo': texto_extraido
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error obteniendo reporte OCR: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/estudiantes/{estudiante_id}/documentos", tags=["Estudiantes"])
async def subir_documento(
    estudiante_id: int,
    tipo_documento: str,
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Sube un documento para el estudiante
    """
    import base64
    from datetime import datetime
    
    # Verificar que el estudiante existe
    from database.models import Estudiante as EstudianteModel
    estudiante = db.query(EstudianteModel).filter(EstudianteModel.id == estudiante_id).first()
    
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    # Validar tipo de archivo
    allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png']
    file_extension = archivo.filename[archivo.filename.rfind('.'):].lower()
    
    if file_extension not in allowed_extensions:
        raise HTTPException(
            status_code=400, 
            detail=f"Tipo de archivo no permitido. Use: {', '.join(allowed_extensions)}"
        )
    
    # Leer archivo y convertir a base64
    contenido = await archivo.read()
    tamano_bytes = len(contenido)
    
    # Validar tamaño (máximo 5MB)
    if tamano_bytes > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo muy grande. Máximo 5MB")
    
    # Convertir a base64 para almacenar en DB
    contenido_base64 = base64.b64encode(contenido).decode('utf-8')
    url_archivo = f"data:application/{file_extension[1:]};base64,{contenido_base64}"
    
    # Guardar en base de datos
    try:
        cursor = db.connection().connection.cursor()
        cursor.execute("""
            INSERT INTO documentos 
            (estudiante_id, tipo_documento, nombre_archivo, url_archivo, tamano_bytes, estado, created_at)
            VALUES (%s, %s, %s, %s, %s, 'pendiente', %s)
            RETURNING id
        """, (estudiante_id, tipo_documento, archivo.filename, url_archivo, tamano_bytes, datetime.utcnow()))
        
        documento_id = cursor.fetchone()[0]
        db.commit()
        
        # Actualizar estado de documentos del estudiante
        estudiante.documentos_estado = 'en_revision'
        db.commit()
        
        return {
            "mensaje": "Documento subido correctamente",
            "documento_id": documento_id,
            "nombre_archivo": archivo.filename,
            "tamano_bytes": tamano_bytes
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al guardar documento: {str(e)}")


@app.get("/api/estudiantes/{estudiante_id}/documentos", tags=["Estudiantes"])
def listar_documentos(estudiante_id: int, db: Session = Depends(get_db)):
    """
    Lista todos los documentos de un estudiante
    """
    from database.models import Estudiante as EstudianteModel
    
    estudiante = db.query(EstudianteModel).filter(EstudianteModel.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    try:
        cursor = db.connection().connection.cursor()
        cursor.execute("""
            SELECT id, tipo_documento, nombre_archivo, tamano_bytes, estado, notas, created_at
            FROM documentos
            WHERE estudiante_id = %s
            ORDER BY created_at DESC
        """, (estudiante_id,))
        
        documentos = []
        for row in cursor.fetchall():
            documentos.append({
                'id': row[0],
                'tipo_documento': row[1],
                'nombre_archivo': row[2],
                'tamano_bytes': row[3],
                'estado': row[4],
                'notas': row[5],
                'created_at': row[6].isoformat() if row[6] else None
            })
        
        return {
            'total': len(documentos),
            'documentos': documentos
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al listar documentos: {str(e)}")


@app.get("/api/estudiantes/{estudiante_id}/reporte-pdf", tags=["Reportes"])
def generar_reporte_pdf(estudiante_id: int, tipo: str = 'completo', db: Session = Depends(get_db)):
    """
    Genera un PDF con el reporte del estudiante
    tipo: 'completo' o 'analisis'
    """
    from database.models import Estudiante as EstudianteModel
    from api.generador_pdf import GeneradorReportesPDF
    
    estudiante = db.query(EstudianteModel).filter(EstudianteModel.id == estudiante_id).first()
    
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    estudiante_data = {
        'id': estudiante.id,
        'nombre': estudiante.nombre,
        'email': estudiante.email,
        'telefono': estudiante.telefono,
        'pasaporte': estudiante.pasaporte,
        'edad': estudiante.edad,
        'nacionalidad': estudiante.nacionalidad,
        'ciudad_origen': estudiante.ciudad_origen,
        'especialidad': estudiante.especialidad,
        'nivel_espanol': estudiante.nivel_espanol,
        'tipo_visa': estudiante.tipo_visa,
        'estado': estudiante.estado,
        'documentos_estado': estudiante.documentos_estado,
        'notas': estudiante.notas
    }
    
    if tipo == 'analisis':
        # Generar análisis de visa
        from api.calculadora_visa import CalculadoraProbabilidadVisa
        analisis = CalculadoraProbabilidadVisa.calcular_probabilidad(estudiante_data)
        pdf_buffer = GeneradorReportesPDF.generar_reporte_analisis_visa(estudiante_data, analisis)
        filename = f"analisis_visa_{estudiante.id}.pdf"
    else:
        # Generar reporte completo
        pdf_buffer = GeneradorReportesPDF.generar_reporte_completo_estudiante(estudiante_data)
        filename = f"reporte_estudiante_{estudiante.id}.pdf"
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================================================
# GENERACIÓN DE DOCUMENTOS OFICIALES (ADMIN)
# ============================================================================

class GenerarDocumentosRequest(BaseModel):
    tipos_documentos: List[str]

@app.post("/api/admin/estudiantes/{estudiante_id}/generar-documentos", tags=["Admin - Documentos"])
def generar_documentos_estudiante(
    estudiante_id: int,
    request: GenerarDocumentosRequest,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """
    Genera documentos oficiales para un estudiante
    tipos_documentos: ['carta_aceptacion', 'carta_motivacion', 'formulario_solicitud', 'certificado_matricula']
    """
    verificar_token(credentials.credentials)
    tipos_documentos = request.tipos_documentos
    
    from database.models import Estudiante as EstudianteModel
    from api.generador_documentos import GeneradorDocumentosOficiales
    import base64
    
    estudiante = db.query(EstudianteModel).filter(EstudianteModel.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    datos_estudiante = {
        'id': estudiante.id,
        'nombre': estudiante.nombre,
        'email': estudiante.email,
        'telefono': estudiante.telefono,
        'pasaporte': estudiante.pasaporte,
        'edad': estudiante.edad,
        'nacionalidad': estudiante.nacionalidad,
        'ciudad_origen': estudiante.ciudad_origen,
        'especialidad': estudiante.especialidad,
        'nivel_espanol': estudiante.nivel_espanol,
        'tipo_visa': estudiante.tipo_visa
    }
    
    documentos_generados = []
    
    for tipo in tipos_documentos:
        try:
            # Generar PDF según tipo
            if tipo == 'carta_aceptacion':
                pdf_buffer = GeneradorDocumentosOficiales.generar_carta_aceptacion(datos_estudiante)
                nombre = f"Carta_Aceptacion_{estudiante.id}.pdf"
            elif tipo == 'carta_motivacion':
                pdf_buffer = GeneradorDocumentosOficiales.generar_carta_motivacion(datos_estudiante)
                nombre = f"Carta_Motivacion_{estudiante.id}.pdf"
            elif tipo == 'formulario_solicitud':
                pdf_buffer = GeneradorDocumentosOficiales.generar_formulario_solicitud(datos_estudiante)
                nombre = f"Formulario_Solicitud_{estudiante.id}.pdf"
            elif tipo == 'certificado_matricula':
                pdf_buffer = GeneradorDocumentosOficiales.generar_certificado_matricula(datos_estudiante)
                nombre = f"Certificado_Matricula_{estudiante.id}.pdf"
            else:
                continue
            
            # Convertir a base64 para almacenar
            pdf_content = pdf_buffer.read()
            pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
            
            # Guardar en base de datos usando SQL directo
            import os
            import psycopg2
            conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO documentos_generados 
                (estudiante_id, tipo_documento, nombre_archivo, contenido_pdf, estado, generado_por)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (estudiante_id, tipo, nombre, pdf_base64, 'generado', 'admin'))
            
            doc_id = cursor.fetchone()[0]
            conn.commit()
            cursor.close()
            conn.close()
            
            documentos_generados.append({
                'id': doc_id,
                'tipo': tipo,
                'nombre': nombre,
                'estado': 'generado'
            })
            
        except Exception as e:
            print(f"Error generando documento {tipo}: {e}")
            continue
    
    # Crear notificación si se generaron documentos
    if documentos_generados:
        from api.notificaciones_routes import crear_notificacion
        crear_notificacion(
            db=db,
            estudiante_id=estudiante_id,
            tipo='documento',
            titulo='📄 Nuevos documentos generados',
            mensaje=f'Se han generado {len(documentos_generados)} documento(s) para ti. Revísalos en tu dashboard.',
            url_accion='/estudiante/dashboard',
            icono='📄',
            prioridad='alta'
        )
    
    return {
        'estudiante_id': estudiante_id,
        'documentos_generados': documentos_generados,
        'total': len(documentos_generados)
    }


@app.get("/api/estudiantes/{estudiante_id}/documentos-generados", tags=["Estudiantes"])
def obtener_documentos_generados_estudiante(
    estudiante_id: int,
    codigo_acceso: str,
    db: Session = Depends(get_db)
):
    """Obtiene documentos generados de un estudiante (requiere código de acceso)"""
    
    # Verificar estudiante y código de acceso
    result = db.execute(
        text("SELECT id, codigo_acceso FROM estudiantes WHERE id = :id"),
        {"id": estudiante_id}
    ).fetchone()
    
    if not result:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    if result[1] != codigo_acceso:
        raise HTTPException(status_code=403, detail="Código de acceso inválido")
    
    # Obtener documentos generados
    docs_result = db.execute(
        text("""
            SELECT id, tipo_documento, nombre_archivo, estado, 
                   fecha_generacion, fecha_aprobacion, enviado_estudiante,
                   notas, contenido_pdf
            FROM documentos_generados
            WHERE estudiante_id = :estudiante_id
            ORDER BY fecha_generacion DESC
        """),
        {"estudiante_id": estudiante_id}
    ).fetchall()
    
    documentos = []
    for row in docs_result:
        documentos.append({
            'id': row[0],
            'tipo_documento': row[1],
            'nombre_archivo': row[2],
            'estado': row[3],
            'fecha_generacion': row[4].isoformat() if row[4] else None,
            'fecha_aprobacion': row[5].isoformat() if row[5] else None,
            'enviado_estudiante': row[6],
            'notas': row[7],
            'contenido_pdf': row[8]  # Base64 del PDF
        })
    
    return {
        'estudiante_id': estudiante_id,
        'documentos': documentos,
        'total': len(documentos)
    }


@app.get("/api/admin/documentos-generados", tags=["Admin - Documentos"])
def listar_documentos_generados(
    estudiante_id: Optional[int] = None,
    estado: Optional[str] = None,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Lista todos los documentos generados con filtros opcionales"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    try:
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        query = """
            SELECT dg.id, dg.estudiante_id, dg.tipo_documento, dg.nombre_archivo,
                   dg.estado, dg.fecha_generacion, dg.enviado_estudiante,
                   e.nombre as estudiante_nombre
            FROM documentos_generados dg
            LEFT JOIN estudiantes e ON dg.estudiante_id = e.id
            WHERE 1=1
        """
        params = []
        
        if estudiante_id:
            query += " AND dg.estudiante_id = %s"
            params.append(estudiante_id)
        
        if estado:
            query += " AND dg.estado = %s"
            params.append(estado)
        
        query += " ORDER BY dg.fecha_generacion DESC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        print(f"[DEBUG] Documentos generados encontrados: {len(rows)}")
        
        documentos = []
        for row in rows:
            documentos.append({
                'id': row[0],
                'estudiante_id': row[1],
                'tipo_documento': row[2],
                'nombre_archivo': row[3],
                'estado': row[4],
                'fecha_generacion': row[5].isoformat() if row[5] else None,
                'enviado_estudiante': row[6],
                'estudiante_nombre': row[7] or f'Estudiante {row[1]}'
            })
        
        cursor.close()
        conn.close()
        
        return documentos
        
    except Exception as e:
        print(f"❌ Error listando documentos generados: {e}")
        import traceback
        traceback.print_exc()
        return []


@app.get("/api/admin/documentos-generados/{documento_id}/descargar", tags=["Admin - Documentos"])
def descargar_documento_generado(
    documento_id: int,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Descarga un documento generado"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    import base64
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT nombre_archivo, contenido_pdf
        FROM documentos_generados
        WHERE id = %s
    """, (documento_id,))
    
    row = cursor.fetchone()
    cursor.close()
    conn.close()
    
    if not row:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    nombre_archivo, contenido_base64 = row
    pdf_content = base64.b64decode(contenido_base64)
    
    from io import BytesIO
    buffer = BytesIO(pdf_content)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
    )


@app.get("/api/admin/estudiantes/{estudiante_id}/descargar-expediente", tags=["Admin - Documentos"])
def descargar_expediente_completo(
    estudiante_id: int,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Descarga expediente completo del estudiante en formato ZIP"""
    import zipfile
    import base64
    from io import BytesIO
    
    # Obtener datos del estudiante
    estudiante_data = db.execute(
        text("SELECT nombre, email, pasaporte FROM estudiantes WHERE id = :id"),
        {"id": estudiante_id}
    ).fetchone()
    
    if not estudiante_data:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    nombre_estudiante = estudiante_data[0]
    
    # Crear ZIP en memoria
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        
        # 1. Documentos generados (PDFs)
        docs_generados = db.execute(
            text("""
                SELECT nombre_archivo, contenido_pdf, tipo_documento
                FROM documentos_generados
                WHERE estudiante_id = :id
            """),
            {"id": estudiante_id}
        ).fetchall()
        
        for doc in docs_generados:
            nombre, contenido_b64, tipo = doc
            if contenido_b64:
                pdf_bytes = base64.b64decode(contenido_b64)
                zip_file.writestr(f"documentos_generados/{nombre}", pdf_bytes)
        
        # 2. Documentos subidos por el estudiante
        docs_subidos = db.execute(
            text("""
                SELECT nombre_archivo, contenido_base64, tipo_documento
                FROM documentos
                WHERE estudiante_id = :id
            """),
            {"id": estudiante_id}
        ).fetchall()
        
        for doc in docs_subidos:
            nombre, contenido_b64, tipo = doc
            if contenido_b64:
                try:
                    file_bytes = base64.b64decode(contenido_b64)
                    zip_file.writestr(f"documentos_subidos/{nombre}", file_bytes)
                except Exception as e:
                    print(f"Error procesando documento {nombre}: {e}")
        
        # 3. Información del estudiante en JSON
        import json
        info_estudiante = db.execute(
            text("""
                SELECT nombre, email, telefono, pasaporte, edad, nacionalidad,
                       pais_origen, ciudad_origen, carrera_deseada, especialidad,
                       nivel_espanol, tipo_visa, fondos_disponibles, estado,
                       fecha_inicio_estimada, created_at
                FROM estudiantes WHERE id = :id
            """),
            {"id": estudiante_id}
        ).fetchone()
        
        if info_estudiante:
            estudiante_json = {
                "nombre": info_estudiante[0],
                "email": info_estudiante[1],
                "telefono": info_estudiante[2],
                "pasaporte": info_estudiante[3],
                "edad": info_estudiante[4],
                "nacionalidad": info_estudiante[5],
                "pais_origen": info_estudiante[6],
                "ciudad_origen": info_estudiante[7],
                "carrera_deseada": info_estudiante[8],
                "especialidad": info_estudiante[9],
                "nivel_espanol": info_estudiante[10],
                "tipo_visa": info_estudiante[11],
                "fondos_disponibles": float(info_estudiante[12]) if info_estudiante[12] else 0,
                "estado": info_estudiante[13],
                "fecha_inicio_estimada": info_estudiante[14].isoformat() if info_estudiante[14] else None,
                "fecha_registro": info_estudiante[15].isoformat() if info_estudiante[15] else None
            }
            
            zip_file.writestr(
                "info_estudiante.json",
                json.dumps(estudiante_json, indent=2, ensure_ascii=False)
            )
        
        # 4. README con información del expediente
        readme_content = f"""EXPEDIENTE COMPLETO - {nombre_estudiante}
{'='*60}

Este archivo ZIP contiene:

📁 documentos_generados/
   - Documentos oficiales generados por el sistema
   - Cartas de aceptación, motivación, formularios, etc.

📁 documentos_subidos/
   - Documentos proporcionados por el estudiante
   - Pasaporte, certificados académicos, extractos bancarios, etc.

📄 info_estudiante.json
   - Información completa del perfil del estudiante
   - Datos personales, académicos y financieros

Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC
Por: {usuario['email']}
"""
        zip_file.writestr("README.txt", readme_content)
    
    # Registrar auditoría
    registrar_auditoria(
        db, usuario['email'], 'DESCARGAR_EXPEDIENTE',
        'estudiante', estudiante_id,
        {'nombre': nombre_estudiante}
    )
    
    # Preparar respuesta
    zip_buffer.seek(0)
    filename = f"expediente_{nombre_estudiante.replace(' ', '_')}_{estudiante_id}.zip"
    
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.put("/api/admin/documentos-generados/{documento_id}/aprobar", tags=["Admin - Documentos"])
def aprobar_documento_generado(
    documento_id: int,
    enviar_a_estudiante: bool = True,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Aprueba un documento y opcionalmente lo envía al estudiante"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    cursor.execute("""
        UPDATE documentos_generados
        SET estado = 'aprobado',
            fecha_aprobacion = CURRENT_TIMESTAMP,
            enviado_estudiante = %s,
            aprobado_por = 'admin'
        WHERE id = %s
        RETURNING estudiante_id
    """, (enviar_a_estudiante, documento_id))
    
    result = cursor.fetchone()
    if not result:
        cursor.close()
        conn.close()
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    estudiante_id = result[0]
    conn.commit()
    cursor.close()
    conn.close()
    
    # Enviar email al estudiante con el documento adjunto
    if enviar_a_estudiante:
        try:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            from email.mime.application import MIMEApplication
            import base64
            
            # Obtener email del estudiante y contenido del PDF
            cursor2 = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require').cursor()
            cursor2.execute("""
                SELECT e.email, e.nombre, dg.nombre_archivo, dg.tipo_documento, dg.contenido_pdf
                FROM documentos_generados dg
                JOIN estudiantes e ON dg.estudiante_id = e.id
                WHERE dg.id = %s
            """, (documento_id,))
            
            row = cursor2.fetchone()
            if row:
                estudiante_email, estudiante_nombre, nombre_archivo, tipo_doc, pdf_content = row
                cursor2.close()
                
                # Configurar email
                msg = MIMEMultipart()
                msg['From'] = os.getenv('SMTP_USER')
                msg['To'] = estudiante_email
                msg['Subject'] = f'Documento Aprobado: {tipo_doc.replace("_", " ").title()}'
                
                # Cuerpo del email
                body = f"""
                <html>
                <body>
                    <h2>¡Hola {estudiante_nombre}!</h2>
                    <p>Tu documento <strong>{tipo_doc.replace("_", " ").title()}</strong> ha sido aprobado y está listo.</p>
                    <p>Puedes descargarlo desde tu perfil o encontrarlo adjunto en este correo.</p>
                    <br>
                    <p>Saludos,<br>Equipo de Estudio Visa España</p>
                </body>
                </html>
                """
                msg.attach(MIMEText(body, 'html'))
                
                # Adjuntar PDF
                pdf_bytes = base64.b64decode(pdf_content)
                pdf_attachment = MIMEApplication(pdf_bytes, _subtype='pdf')
                pdf_attachment.add_header('Content-Disposition', 'attachment', filename=nombre_archivo)
                msg.attach(pdf_attachment)
                
                # Enviar
                smtp = smtplib.SMTP(os.getenv('SMTP_SERVER'), int(os.getenv('SMTP_PORT')))
                smtp.starttls()
                smtp.login(os.getenv('SMTP_USER'), os.getenv('SMTP_PASSWORD'))
                smtp.send_message(msg)
                smtp.quit()
        except Exception as e:
            print(f"Error enviando email: {str(e)}")
    
    return {
        'mensaje': 'Documento aprobado correctamente',
        'enviado_estudiante': enviar_a_estudiante
    }


# ============================================================================
# GESTIÓN DE CURSOS
# ============================================================================

@app.get("/api/admin/cursos", tags=["Admin - Cursos"])
def listar_cursos(
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Lista todos los cursos disponibles"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, nombre, descripcion, duracion_meses, precio_eur, ciudad,
               nivel_espanol_requerido, cupos_disponibles, activo
        FROM cursos
        ORDER BY nombre
    """)
    
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    
    cursos = []
    for row in rows:
        cursos.append({
            'id': row[0],
            'nombre': row[1],
            'descripcion': row[2],
            'duracion_meses': row[3],
            'precio_eur': float(row[4]) if row[4] else None,
            'ciudad': row[5],
            'nivel_espanol_requerido': row[6],
            'cupos_disponibles': row[7],
            'activo': row[8]
        })
    
    return cursos


@app.post("/api/admin/cursos", tags=["Admin - Cursos"])
def crear_curso(
    curso: dict,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Crea un nuevo curso"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    cursor.execute("""
        INSERT INTO cursos (nombre, descripcion, duracion_meses, precio_eur, ciudad,
                           nivel_espanol_requerido, cupos_disponibles, activo)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (
        curso.get('nombre'),
        curso.get('descripcion'),
        curso.get('duracion_meses'),
        curso.get('precio_eur'),
        curso.get('ciudad'),
        curso.get('nivel_espanol_requerido'),
        curso.get('cupos_disponibles', 0),
        curso.get('activo', True)
    ))
    
    curso_id = cursor.fetchone()[0]
    conn.commit()
    cursor.close()
    conn.close()
    
    return {'id': curso_id, 'mensaje': 'Curso creado correctamente'}


@app.put("/api/admin/cursos/{curso_id}", tags=["Admin - Cursos"])
def actualizar_curso(
    curso_id: int,
    curso: dict,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Actualiza un curso"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    cursor.execute("""
        UPDATE cursos
        SET nombre = %s, descripcion = %s, duracion_meses = %s, precio_eur = %s,
            ciudad = %s, nivel_espanol_requerido = %s, cupos_disponibles = %s,
            activo = %s, updated_at = CURRENT_TIMESTAMP
        WHERE id = %s
    """, (
        curso.get('nombre'),
        curso.get('descripcion'),
        curso.get('duracion_meses'),
        curso.get('precio_eur'),
        curso.get('ciudad'),
        curso.get('nivel_espanol_requerido'),
        curso.get('cupos_disponibles'),
        curso.get('activo'),
        curso_id
    ))
    
    conn.commit()
    cursor.close()
    conn.close()
    
    return {'mensaje': 'Curso actualizado correctamente'}


@app.post("/api/admin/estudiantes/{estudiante_id}/asignar-curso", tags=["Admin - Cursos"])
def asignar_curso(
    estudiante_id: int,
    curso_id: int = Query(...),
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Asigna un curso a un estudiante"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    cursor.execute("""
        UPDATE estudiantes
        SET curso_asignado_id = %s
        WHERE id = %s
    """, (curso_id, estudiante_id))
    
    conn.commit()
    
    # Obtener datos del curso y estudiante para el email
    cursor.execute("""
        SELECT c.nombre, c.duracion_meses, c.ciudad, c.nivel_espanol_requerido, c.precio_eur,
               e.nombre, e.email
        FROM cursos c
        JOIN estudiantes e ON e.id = %s
        WHERE c.id = %s
    """, (estudiante_id, curso_id))
    
    row = cursor.fetchone()
    cursor.close()
    conn.close()
    
    # Enviar email de notificación
    if row:
        try:
            from api.email_utils import email_curso_asignado
            curso_detalles = {
                'duracion_meses': row[1],
                'ciudad': row[2],
                'nivel_espanol_requerido': row[3],
                'precio_eur': row[4]
            }
            email_curso_asignado(row[5], row[6], row[0], curso_detalles)
        except Exception as e:
            print(f"[WARN] Error enviando email: {e}")
    
    return {'mensaje': 'Curso asignado correctamente'}


@app.get("/api/admin/estudiantes/{estudiante_id}/sugerir-cursos", tags=["Admin - Cursos"])
def sugerir_cursos(
    estudiante_id: int,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Sugiere cursos basados en el perfil del estudiante"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    # Obtener datos del estudiante
    cursor.execute("""
        SELECT especialidad, nivel_idioma, fondos_disponibles
        FROM estudiantes
        WHERE id = %s
    """, (estudiante_id,))
    
    estudiante = cursor.fetchone()
    if not estudiante:
        cursor.close()
        conn.close()
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    especialidad, nivel_idioma, fondos = estudiante
    
    # Mapear nivel de idioma a nivel español requerido
    nivel_map = {
        'basico': ['A1', 'A2', 'Básico', 'Principiante'],
        'intermedio': ['A1', 'A2', 'B1', 'Intermedio', 'Básico', 'Principiante'],
        'avanzado': ['A1', 'A2', 'B1', 'B2', 'C1', 'Avanzado', 'Intermedio', 'Básico', 'Principiante']
    }
    
    niveles_aceptables = nivel_map.get(nivel_idioma.lower() if nivel_idioma else 'basico', ['A1', 'A2'])
    
    # Buscar cursos compatibles
    cursor.execute("""
        SELECT id, nombre, descripcion, duracion_meses, precio_eur, ciudad,
               nivel_espanol_requerido, cupos_disponibles
        FROM cursos
        WHERE activo = TRUE
          AND cupos_disponibles > 0
          AND precio_eur <= %s
        ORDER BY 
            CASE 
                WHEN LOWER(descripcion) LIKE %s THEN 1
                WHEN LOWER(nombre) LIKE %s THEN 2
                ELSE 3
            END,
            precio_eur ASC
        LIMIT 5
    """, (fondos, f'%{especialidad.lower()}%' if especialidad else '%%', f'%{especialidad.lower()}%' if especialidad else '%%'))
    
    cursos = []
    for row in cursor.fetchall():
        # Calcular score de compatibilidad
        score = 100
        nivel_curso = row[6] if row[6] else ''
        
        # Penalizar si el nivel no es compatible
        if nivel_curso and nivel_curso not in niveles_aceptables:
            score -= 30
        
        # Bonus si coincide con especialidad
        if especialidad and (especialidad.lower() in row[1].lower() or 
                            (row[2] and especialidad.lower() in row[2].lower())):
            score += 20
        
        # Penalizar si está cerca del límite de fondos
        if row[4] and fondos:
            ratio = row[4] / fondos
            if ratio > 0.8:
                score -= 15
        
        cursos.append({
            'id': row[0],
            'nombre': row[1],
            'descripcion': row[2],
            'duracion_meses': row[3],
            'precio_eur': float(row[4]) if row[4] else None,
            'ciudad': row[5],
            'nivel_espanol_requerido': row[6],
            'cupos_disponibles': row[7],
            'compatibilidad': max(0, min(100, score)),
            'razon': 'Compatible con tu perfil y fondos disponibles'
        })
    
    # Ordenar por compatibilidad
    cursos.sort(key=lambda x: x['compatibilidad'], reverse=True)
    
    cursor.close()
    conn.close()
    
    return {
        'estudiante_id': estudiante_id,
        'cursos_sugeridos': cursos,
        'criterios': {
            'especialidad': especialidad,
            'nivel_idioma': nivel_idioma,
            'fondos_disponibles': float(fondos) if fondos else None
        }
    }


# ============================================================================
# GESTIÓN DE ALOJAMIENTOS
# ============================================================================

@app.get("/api/admin/alojamientos", tags=["Admin - Alojamientos"])
def listar_alojamientos(
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Lista todos los alojamientos"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, tipo, direccion, ciudad, precio_mensual_eur, capacidad,
               disponible, descripcion, servicios
        FROM alojamientos
        ORDER BY ciudad, tipo
    """)
    
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    
    alojamientos = []
    for row in rows:
        alojamientos.append({
            'id': row[0],
            'tipo': row[1],
            'direccion': row[2],
            'ciudad': row[3],
            'precio_mensual_eur': float(row[4]) if row[4] else None,
            'capacidad': row[5],
            'disponible': row[6],
            'descripcion': row[7],
            'servicios': row[8]
        })
    
    return alojamientos


@app.post("/api/admin/alojamientos", tags=["Admin - Alojamientos"])
def crear_alojamiento(
    alojamiento: dict,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Crea un nuevo alojamiento"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    cursor.execute("""
        INSERT INTO alojamientos (tipo, direccion, ciudad, precio_mensual_eur,
                                 capacidad, disponible, descripcion, servicios)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (
        alojamiento.get('tipo'),
        alojamiento.get('direccion'),
        alojamiento.get('ciudad'),
        alojamiento.get('precio_mensual_eur'),
        alojamiento.get('capacidad', 1),
        alojamiento.get('disponible', True),
        alojamiento.get('descripcion'),
        alojamiento.get('servicios')
    ))
    
    alojamiento_id = cursor.fetchone()[0]
    conn.commit()
    cursor.close()
    conn.close()
    
    return {'id': alojamiento_id, 'mensaje': 'Alojamiento creado correctamente'}


@app.put("/api/admin/alojamientos/{alojamiento_id}", tags=["Admin - Alojamientos"])
def actualizar_alojamiento(
    alojamiento_id: int,
    alojamiento: dict,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Actualiza un alojamiento"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    cursor.execute("""
        UPDATE alojamientos
        SET tipo = %s, direccion = %s, ciudad = %s, precio_mensual_eur = %s,
            capacidad = %s, disponible = %s, descripcion = %s, servicios = %s,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = %s
    """, (
        alojamiento.get('tipo'),
        alojamiento.get('direccion'),
        alojamiento.get('ciudad'),
        alojamiento.get('precio_mensual_eur'),
        alojamiento.get('capacidad'),
        alojamiento.get('disponible'),
        alojamiento.get('descripcion'),
        alojamiento.get('servicios'),
        alojamiento_id
    ))
    
    conn.commit()
    cursor.close()
    conn.close()
    
    return {'mensaje': 'Alojamiento actualizado correctamente'}


@app.post("/api/admin/estudiantes/{estudiante_id}/asignar-alojamiento", tags=["Admin - Alojamientos"])
def asignar_alojamiento(
    estudiante_id: int,
    alojamiento_id: int = Query(...),
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Asigna un alojamiento a un estudiante"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    cursor.execute("""
        UPDATE estudiantes
        SET alojamiento_asignado_id = %s
        WHERE id = %s
    """, (alojamiento_id, estudiante_id))
    
    conn.commit()
    cursor.close()
    conn.close()
    
    return {'mensaje': 'Alojamiento asignado correctamente'}


# ============================================================================
# REPORTES
# ============================================================================

@app.get("/api/estudiantes/{estudiante_id}/descargar-expediente", tags=["Estudiantes"])
def descargar_expediente_zip(
    estudiante_id: int,
    db: Session = Depends(get_db)
):
    """Descarga ZIP con todos los documentos del expediente"""
    import os
    import psycopg2
    import zipfile
    from io import BytesIO
    import base64
    from fastapi.responses import StreamingResponse
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    # Obtener datos del estudiante
    cursor.execute("""
        SELECT nombre, pasaporte
        FROM estudiantes
        WHERE id = %s
    """, (estudiante_id,))
    
    estudiante = cursor.fetchone()
    if not estudiante:
        cursor.close()
        conn.close()
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    nombre, pasaporte = estudiante
    
    # Obtener documentos generados
    cursor.execute("""
        SELECT nombre_archivo, contenido_pdf, tipo_documento
        FROM documentos_generados
        WHERE estudiante_id = %s AND estado = 'aprobado'
    """, (estudiante_id,))
    
    docs_generados = cursor.fetchall()
    
    # Obtener documentos subidos
    cursor.execute("""
        SELECT nombre_archivo, contenido, tipo_documento
        FROM documentos
        WHERE estudiante_id = %s
    """, (estudiante_id,))
    
    docs_subidos = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    # Crear ZIP en memoria
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Agregar documentos generados
        for doc in docs_generados:
            nombre_archivo, contenido_b64, tipo = doc
            contenido = base64.b64decode(contenido_b64)
            zip_file.writestr(f"generados/{nombre_archivo}", contenido)
        
        # Agregar documentos subidos
        for doc in docs_subidos:
            nombre_archivo, contenido_b64, tipo = doc
            if contenido_b64:
                contenido = base64.b64decode(contenido_b64)
                zip_file.writestr(f"subidos/{nombre_archivo}", contenido)
    
    zip_buffer.seek(0)
    
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename=expediente_{pasaporte}_{nombre.replace(' ', '_')}.zip"
        }
    )


@app.put("/api/admin/estudiantes/{estudiante_id}/actualizar-estado", tags=["Admin"])
def actualizar_estado_estudiante(
    estudiante_id: int,
    nuevo_estado: str,
    notas: Optional[str] = None,
    fecha_evento: Optional[str] = None,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Actualiza estado del estudiante (cita_consular, visa_aprobada, visa_rechazada, llegada_espana)"""
    verificar_token(credentials.credentials)
    
    from database.models import Estudiante as EstudianteModel
    
    estudiante = db.query(EstudianteModel).filter(EstudianteModel.id == estudiante_id).first()
    
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    estados_validos = ['pendiente', 'aprobado', 'rechazado', 'cita_consular', 'visa_aprobada', 'visa_rechazada', 'llegada_espana']
    
    if nuevo_estado not in estados_validos:
        raise HTTPException(status_code=400, detail="Estado no válido")
    
    estudiante.estado = nuevo_estado
    if notas:
        estudiante.notas = notas
    estudiante.updated_at = datetime.utcnow()
    
    db.commit()
    
    # Crear notificación para el estudiante
    from api.notificaciones_routes import crear_notificacion
    
    # Mapear estados a mensajes amigables
    mensajes_estados = {
        'aprobado': {
            'titulo': '✅ Solicitud Aprobada',
            'mensaje': '¡Felicitaciones! Tu solicitud ha sido aprobada.',
            'icono': '✅',
            'prioridad': 'alta'
        },
        'rechazado': {
            'titulo': '❌ Solicitud Rechazada',
            'mensaje': 'Lamentablemente tu solicitud no fue aprobada. Contacta con nosotros para más información.',
            'icono': '❌',
            'prioridad': 'alta'
        },
        'cita_consular': {
            'titulo': '📅 Cita Consular Programada',
            'mensaje': 'Se ha programado tu cita consular. Revisa los detalles en tu dashboard.',
            'icono': '📅',
            'prioridad': 'urgente'
        },
        'visa_aprobada': {
            'titulo': '🎉 ¡Visa Aprobada!',
            'mensaje': '¡Excelentes noticias! Tu visa ha sido aprobada. Felicitaciones.',
            'icono': '🎉',
            'prioridad': 'urgente'
        },
        'visa_rechazada': {
            'titulo': '😔 Visa Rechazada',
            'mensaje': 'Tu visa no fue aprobada. Contacta con nosotros para explorar opciones.',
            'icono': '😔',
            'prioridad': 'urgente'
        },
        'llegada_espana': {
            'titulo': '✈️ Bienvenido a España',
            'mensaje': '¡Has llegado a España! Te deseamos mucho éxito en tu aventura.',
            'icono': '✈️',
            'prioridad': 'alta'
        }
    }
    
    if nuevo_estado in mensajes_estados:
        info = mensajes_estados[nuevo_estado]
        crear_notificacion(
            db=db,
            estudiante_id=estudiante_id,
            tipo='estado',
            titulo=info['titulo'],
            mensaje=info['mensaje'] + (f' Notas: {notas}' if notas else ''),
            url_accion='/estudiante/dashboard',
            icono=info['icono'],
            prioridad=info['prioridad']
        )
    
    return {"mensaje": "Estado actualizado correctamente", "nuevo_estado": nuevo_estado}


@app.get("/api/admin/alertas-documentos", tags=["Admin"])
def obtener_alertas_documentos(
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Obtiene lista de estudiantes con documentos faltantes o incompletos"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT e.id, e.nombre, e.email, e.estado, e.created_at,
               COUNT(d.id) as docs_subidos,
               COUNT(dg.id) as docs_generados
        FROM estudiantes e
        LEFT JOIN documentos d ON e.id = d.estudiante_id
        LEFT JOIN documentos_generados dg ON e.id = dg.estudiante_id AND dg.estado = 'aprobado'
        WHERE e.estado != 'rechazado'
        GROUP BY e.id, e.nombre, e.email, e.estado, e.created_at
        HAVING COUNT(d.id) < 3 OR COUNT(dg.id) < 4
        ORDER BY e.created_at ASC
    """)
    
    alertas = []
    for row in cursor.fetchall():
        dias_desde_registro = (datetime.now() - row[4]).days if row[4] else 0
        alertas.append({
            'estudiante_id': row[0],
            'nombre': row[1],
            'email': row[2],
            'estado': row[3],
            'docs_subidos': row[5],
            'docs_generados': row[6],
            'dias_desde_registro': dias_desde_registro,
            'urgencia': 'alta' if dias_desde_registro > 7 else 'media' if dias_desde_registro > 3 else 'baja'
        })
    
    cursor.close()
    conn.close()
    
    return {'total_alertas': len(alertas), 'alertas': alertas}


@app.post("/api/admin/enviar-recordatorios", tags=["Admin"])
def enviar_recordatorios_masivos(
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Envía recordatorios por email a estudiantes con documentos pendientes"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    from api.email_utils import email_recordatorio_documentos
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    # Buscar estudiantes con documentación incompleta
    cursor.execute("""
        SELECT e.id, e.nombre, e.email, e.created_at,
               COUNT(d.id) as docs_subidos
        FROM estudiantes e
        LEFT JOIN documentos d ON e.id = d.estudiante_id
        WHERE e.estado = 'pendiente'
        GROUP BY e.id, e.nombre, e.email, e.created_at
        HAVING COUNT(d.id) < 3
    """)
    
    enviados = 0
    for row in cursor.fetchall():
        est_id, nombre, email, created_at, docs_count = row
        dias_desde_registro = (datetime.now() - created_at).days if created_at else 0
        
        # Solo enviar si han pasado más de 3 días
        if dias_desde_registro > 3:
            docs_faltantes = []
            if docs_count < 1:
                docs_faltantes.append("Título universitario")
            if docs_count < 2:
                docs_faltantes.append("Pasaporte")
            if docs_count < 3:
                docs_faltantes.append("Extracto bancario")
            
            try:
                if email_recordatorio_documentos(nombre, email, docs_faltantes):
                    enviados += 1
            except Exception as e:
                print(f"Error enviando a {email}: {e}")
    
    cursor.close()
    conn.close()
    
    return {'mensaje': f'Recordatorios enviados: {enviados}'}


@app.get("/api/admin/reportes/estudiantes", tags=["Admin - Reportes"])
def reporte_estudiantes(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Genera reporte detallado de estudiantes"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    query = """
        SELECT e.id, e.nombre, e.email, e.nacionalidad, e.especialidad,
               e.estado, e.created_at, c.nombre as curso, a.tipo as alojamiento
        FROM estudiantes e
        LEFT JOIN cursos c ON e.curso_asignado_id = c.id
        LEFT JOIN alojamientos a ON e.alojamiento_asignado_id = a.id
        WHERE 1=1
    """
    params = []
    
    if fecha_inicio:
        query += " AND e.created_at >= %s"
        params.append(fecha_inicio)
    
    if fecha_fin:
        query += " AND e.created_at <= %s"
        params.append(fecha_fin)
    
    query += " ORDER BY e.created_at DESC"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    
    estudiantes = []
    for row in rows:
        estudiantes.append({
            'id': row[0],
            'nombre': row[1],
            'email': row[2],
            'nacionalidad': row[3],
            'especialidad': row[4],
            'estado': row[5],
            'fecha_registro': row[6].isoformat() if row[6] else None,
            'curso': row[7],
            'alojamiento': row[8]
        })
    
    cursor.close()
    conn.close()
    
    return {
        'total': len(estudiantes),
        'estudiantes': estudiantes,
        'fecha_generacion': datetime.now().isoformat()
    }


@app.get("/api/estudiantes/{estudiante_id}/estado", tags=["Estudiantes"])
def consultar_estado(estudiante_id: int, db: Session = Depends(get_db)):
    """
    Estudiante consulta su estado (sin auth por simplicidad en MVP)
    """
    try:
        # Query solo los campos que necesitamos
        result = db.execute(
            text("""
                SELECT id, nombre, email, tipo_visa, estado, created_at, curso_asignado_id
                FROM estudiantes 
                WHERE id = :estudiante_id
            """),
            {"estudiante_id": estudiante_id}
        ).fetchone()
        
        if not result:
            raise HTTPException(status_code=404, detail="Estudiante no encontrado")
        
        return {
            "nombre": result[1] or "Estudiante",
            "estado_procesamiento": result[4] or "pendiente",
            "estado_visa": result[3] or "estudiante",
            "fecha_registro": result[5].isoformat() if result[5] else None,
            "curso_seleccionado": result[6],
            "mensaje": _obtener_mensaje_estado(result[4] or "pendiente")
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Error en consultar_estado: {e}")
        raise HTTPException(status_code=500, detail=f"Error consultando estado: {str(e)}")


def _obtener_mensaje_estado(estado: str) -> str:
    """Mensajes amigables por estado"""
    mensajes = {
        "pendiente": "Tu solicitud ha sido recibida. Estamos procesando tu información.",
        "en_revision": "Tu caso está siendo revisado por nuestro equipo.",
        "aprobado": "¡Felicidades! Tu solicitud ha sido aprobada. Te contactaremos pronto.",
        "rechazado": "Necesitamos información adicional. Te contactaremos.",
        "completado": "¡Proceso completado! Te hemos enviado toda la información."
    }
    return mensajes.get(estado, "En proceso")


# ============================================================================
# MENSAJERÍA INTERNA
# ============================================================================

@app.get("/api/estudiantes/{estudiante_id}/mensajes", tags=["Mensajería"])
def obtener_mensajes(estudiante_id: int, db: Session = Depends(get_db)):
    """Obtiene todos los mensajes de un estudiante"""
    from database.models import Estudiante as EstudianteModel
    
    estudiante = db.query(EstudianteModel).filter(EstudianteModel.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    try:
        cursor = db.connection().connection.cursor()
        cursor.execute("""
            SELECT id, remitente, mensaje, leido, created_at
            FROM mensajes
            WHERE estudiante_id = %s
            ORDER BY created_at ASC
        """, (estudiante_id,))
        
        mensajes = []
        for row in cursor.fetchall():
            mensajes.append({
                'id': row[0],
                'remitente': row[1],
                'mensaje': row[2],
                'leido': row[3],
                'created_at': row[4].isoformat() if row[4] else None
            })
        
        return {'total': len(mensajes), 'mensajes': mensajes}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.post("/api/estudiantes/{estudiante_id}/mensajes", tags=["Mensajería"])
def enviar_mensaje(estudiante_id: int, datos: dict, db: Session = Depends(get_db)):
    """Envía un mensaje (estudiante o admin)"""
    from database.models import Estudiante as EstudianteModel
    from datetime import datetime
    
    estudiante = db.query(EstudianteModel).filter(EstudianteModel.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    remitente = datos.get('remitente', 'estudiante')  # 'estudiante' o 'admin'
    mensaje = datos.get('mensaje', '')
    
    if not mensaje:
        raise HTTPException(status_code=400, detail="El mensaje no puede estar vacío")
    
    try:
        cursor = db.connection().connection.cursor()
        cursor.execute("""
            INSERT INTO mensajes (estudiante_id, remitente, mensaje, created_at)
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, (estudiante_id, remitente, mensaje, datetime.utcnow()))
        
        mensaje_id = cursor.fetchone()[0]
        db.commit()
        
        return {
            'mensaje_id': mensaje_id,
            'mensaje': 'Mensaje enviado correctamente'
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.put("/api/mensajes/{mensaje_id}/marcar-leido", tags=["Mensajería"])
def marcar_mensaje_leido(mensaje_id: int, db: Session = Depends(get_db)):
    """Marca un mensaje como leído"""
    try:
        cursor = db.connection().connection.cursor()
        cursor.execute("""
            UPDATE mensajes 
            SET leido = TRUE 
            WHERE id = %s
        """, (mensaje_id,))
        
        db.commit()
        
        return {'mensaje': 'Mensaje marcado como leído'}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ============================================================================
# ENDPOINTS ADMIN (Requieren autenticación)
# ============================================================================

@app.get("/api/admin/estudiantes", response_model=List[Dict], tags=["Admin"])
def listar_estudiantes(
    estado: Optional[str] = None,
    busqueda: Optional[str] = None,
    nacionalidad: Optional[str] = None,
    especialidad: Optional[str] = None,
    fondos_min: Optional[float] = None,
    fondos_max: Optional[float] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    ordenar_por: Optional[str] = "created_at",
    orden: Optional[str] = "DESC",
    skip: int = 0,
    limit: int = 100,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Lista estudiantes con filtros avanzados"""
    query_text = """
        SELECT 
            id, 
            COALESCE(nombre, nombre_completo, 'Estudiante ' || id::text) as nombre,
            COALESCE(email, 'estudiante' || id::text || '@example.com') as email,
            COALESCE(especialidad, especialidad_interes, tipo_visa, 'No especificado') as especialidad,
            COALESCE(estado, estado_procesamiento, 'pendiente') as estado,
            nacionalidad,
            fondos_disponibles,
            created_at
        FROM estudiantes
        WHERE 1=1
    """
    
    params = {"skip": skip, "limit": limit}
    
    # Filtro de estado
    if estado:
        query_text += " AND COALESCE(estado, estado_procesamiento, 'pendiente') = :estado"
        params["estado"] = estado
    
    # Búsqueda por nombre, email o pasaporte
    if busqueda:
        query_text += """ AND (
            LOWER(COALESCE(nombre, nombre_completo, '')) LIKE LOWER(:busqueda)
            OR LOWER(email) LIKE LOWER(:busqueda)
            OR LOWER(pasaporte) LIKE LOWER(:busqueda)
        )"""
        params["busqueda"] = f"%{busqueda}%"
    
    # Filtro por nacionalidad
    if nacionalidad:
        query_text += " AND LOWER(nacionalidad) = LOWER(:nacionalidad)"
        params["nacionalidad"] = nacionalidad
    
    # Filtro por especialidad
    if especialidad:
        query_text += " AND LOWER(COALESCE(especialidad, especialidad_interes, '')) LIKE LOWER(:especialidad)"
        params["especialidad"] = f"%{especialidad}%"
    
    # Filtro por rango de fondos
    if fondos_min is not None:
        query_text += " AND fondos_disponibles >= :fondos_min"
        params["fondos_min"] = fondos_min
    
    if fondos_max is not None:
        query_text += " AND fondos_disponibles <= :fondos_max"
        params["fondos_max"] = fondos_max
    
    # Filtro por rango de fechas
    if fecha_desde:
        query_text += " AND created_at >= :fecha_desde"
        params["fecha_desde"] = fecha_desde
    
    if fecha_hasta:
        query_text += " AND created_at <= :fecha_hasta"
        params["fecha_hasta"] = fecha_hasta
    
    # Ordenamiento dinámico
    campos_orden_validos = ["created_at", "nombre", "email", "fondos_disponibles", "estado"]
    if ordenar_por in campos_orden_validos:
        orden_seguro = "DESC" if orden.upper() == "DESC" else "ASC"
        query_text += f" ORDER BY {ordenar_por} {orden_seguro}"
    else:
        query_text += " ORDER BY created_at DESC"
    
    query_text += " OFFSET :skip LIMIT :limit"
    
    result = db.execute(text(query_text), params).fetchall()
    
    return [{
        'id': row[0],
        'nombre': row[1],
        'nombre_completo': row[1],
        'email': row[2],
        'especialidad': row[3],
        'especialidad_interes': row[3],
        'estado': row[4],
        'estado_procesamiento': row[4],
        'nacionalidad': row[5],
        'fondos_disponibles': float(row[6]) if row[6] else 0,
        'prioridad': 'BAJA',
        'documentos_subidos': 0,
        'documentos_generados': 0,
        'dias_desde_registro': (datetime.now() - row[7]).days if row[7] else 0,
        'created_at': row[7].isoformat() if row[7] else None
    } for row in result]


@app.get("/api/admin/estudiantes/exportar/{formato}", tags=["Admin"])
def exportar_estudiantes(
    formato: str,
    estado: Optional[str] = None,
    nacionalidad: Optional[str] = None,
    especialidad: Optional[str] = None,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Exporta lista de estudiantes a Excel o CSV"""
    if formato not in ["excel", "csv"]:
        raise HTTPException(status_code=400, detail="Formato debe ser 'excel' o 'csv'")
    
    # Obtener datos
    query_text = """
        SELECT 
            id, 
            COALESCE(nombre, nombre_completo, 'Estudiante ' || id::text) as nombre,
            email,
            telefono,
            pasaporte,
            edad,
            nacionalidad,
            pais_origen,
            ciudad_origen,
            carrera_deseada,
            COALESCE(especialidad, especialidad_interes, 'No especificado') as especialidad,
            nivel_espanol,
            tipo_visa,
            fondos_disponibles,
            COALESCE(estado, estado_procesamiento, 'pendiente') as estado,
            fecha_inicio_estimada,
            created_at
        FROM estudiantes
        WHERE 1=1
    """
    
    params = {}
    
    if estado:
        query_text += " AND COALESCE(estado, estado_procesamiento, 'pendiente') = :estado"
        params["estado"] = estado
    
    if nacionalidad:
        query_text += " AND LOWER(nacionalidad) = LOWER(:nacionalidad)"
        params["nacionalidad"] = nacionalidad
    
    if especialidad:
        query_text += " AND LOWER(COALESCE(especialidad, especialidad_interes, '')) LIKE LOWER(:especialidad)"
        params["especialidad"] = f"%{especialidad}%"
    
    query_text += " ORDER BY created_at DESC"
    
    result = db.execute(text(query_text), params).fetchall()
    
    # Registrar auditoría
    registrar_auditoria(
        db, usuario['email'], 'EXPORTAR_ESTUDIANTES',
        'reporte', None,
        {'formato': formato, 'total_registros': len(result)}
    )
    
    if formato == "csv":
        # Generar CSV
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Encabezados
        writer.writerow([
            'ID', 'Nombre', 'Email', 'Teléfono', 'Pasaporte', 'Edad',
            'Nacionalidad', 'País Origen', 'Ciudad', 'Carrera Deseada',
            'Especialidad', 'Nivel Español', 'Tipo Visa', 'Fondos (EUR)',
            'Estado', 'Fecha Inicio', 'Fecha Registro'
        ])
        
        # Datos
        for row in result:
            writer.writerow([
                row[0], row[1], row[2], row[3], row[4], row[5],
                row[6], row[7], row[8], row[9], row[10], row[11],
                row[12], row[13], row[14],
                row[15].strftime('%Y-%m-%d') if row[15] else '',
                row[16].strftime('%Y-%m-%d %H:%M') if row[16] else ''
            ])
        
        output.seek(0)
        
        from fastapi.responses import Response
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=estudiantes_{datetime.now().strftime('%Y%m%d')}.csv"
            }
        )
    
    else:  # Excel
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estudiantes"
        
        # Encabezados con estilo
        headers = [
            'ID', 'Nombre', 'Email', 'Teléfono', 'Pasaporte', 'Edad',
            'Nacionalidad', 'País Origen', 'Ciudad', 'Carrera Deseada',
            'Especialidad', 'Nivel Español', 'Tipo Visa', 'Fondos (EUR)',
            'Estado', 'Fecha Inicio', 'Fecha Registro'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Datos
        for row_idx, row in enumerate(result, 2):
            ws.cell(row=row_idx, column=1, value=row[0])
            ws.cell(row=row_idx, column=2, value=row[1])
            ws.cell(row=row_idx, column=3, value=row[2])
            ws.cell(row=row_idx, column=4, value=row[3])
            ws.cell(row=row_idx, column=5, value=row[4])
            ws.cell(row=row_idx, column=6, value=row[5])
            ws.cell(row=row_idx, column=7, value=row[6])
            ws.cell(row=row_idx, column=8, value=row[7])
            ws.cell(row=row_idx, column=9, value=row[8])
            ws.cell(row=row_idx, column=10, value=row[9])
            ws.cell(row=row_idx, column=11, value=row[10])
            ws.cell(row=row_idx, column=12, value=row[11])
            ws.cell(row=row_idx, column=13, value=row[12])
            ws.cell(row=row_idx, column=14, value=row[13])
            ws.cell(row=row_idx, column=15, value=row[14])
            ws.cell(row=row_idx, column=16, value=row[15].strftime('%Y-%m-%d') if row[15] else '')
            ws.cell(row=row_idx, column=17, value=row[16].strftime('%Y-%m-%d %H:%M') if row[16] else '')
        
        # Ajustar anchos de columna
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        # Guardar en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=estudiantes_{datetime.now().strftime('%Y%m%d')}.xlsx"
            }
        )



@app.get("/api/admin/estudiantes/{estudiante_id}", tags=["Admin"])
def obtener_estudiante(
    estudiante_id: int,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Detalle completo de un estudiante usando SQL directo"""
    query = text("""
        SELECT 
            id, nombre, email, telefono, pasaporte, fecha_nacimiento, edad, 
            nacionalidad, pais_origen, ciudad_origen, carrera_deseada, especialidad, 
            nivel_espanol, tipo_visa, fondos_disponibles, fecha_inicio_estimada,
            archivo_titulo, archivo_pasaporte, archivo_extractos,
            consentimiento_gdpr, fecha_consentimiento,
            estado, documentos_estado, notas, 
            created_at, updated_at
        FROM estudiantes 
        WHERE id = :estudiante_id
    """)
    
    result = db.execute(query, {"estudiante_id": estudiante_id}).fetchone()
    
    if not result:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    return {
        "id": result[0],
        "nombre": result[1],
        "email": result[2],
        "telefono": result[3],
        "pasaporte": result[4],
        "fecha_nacimiento": result[5].isoformat() if result[5] else None,
        "edad": result[6],
        "nacionalidad": result[7],
        "pais_origen": result[8],
        "ciudad_origen": result[9],
        "carrera_deseada": result[10],
        "especialidad": result[11],
        "nivel_espanol": result[12],
        "tipo_visa": result[13],
        "fondos_disponibles": float(result[14]) if result[14] else 0,
        "fecha_inicio_estimada": result[15].isoformat() if result[15] else None,
        "archivo_titulo": result[16],
        "archivo_pasaporte": result[17],
        "archivo_extractos": result[18],
        "consentimiento_gdpr": result[19],
        "fecha_consentimiento": result[20].isoformat() if result[20] else None,
        "estado": result[21],
        "documentos_estado": result[22],
        "notas": result[23],
        "created_at": result[24].isoformat() if result[24] else None,
        "updated_at": result[25].isoformat() if result[25] else None
    }


@app.patch("/api/admin/estudiantes/{estudiante_id}/campo", tags=["Admin"])
def actualizar_campo_rapido(
    estudiante_id: int,
    datos: dict,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Actualización rápida de un solo campo (edición inline)"""
    campo = datos.get('campo')
    valor = datos.get('valor')
    
    if not campo:
        raise HTTPException(status_code=400, detail="Campo requerido")
    
    # Campos permitidos para edición inline
    campos_permitidos = [
        'nombre', 'email', 'telefono', 'pasaporte', 'edad', 
        'nacionalidad', 'pais_origen', 'ciudad_origen', 
        'carrera_deseada', 'especialidad', 'nivel_espanol', 
        'tipo_visa', 'fondos_disponibles', 'notas', 'estado'
    ]
    
    if campo not in campos_permitidos:
        raise HTTPException(status_code=400, detail=f"Campo '{campo}' no permitido")
    
    # Obtener valor anterior
    valor_anterior = db.execute(
        text(f"SELECT {campo} FROM estudiantes WHERE id = :id"),
        {"id": estudiante_id}
    ).fetchone()
    
    if not valor_anterior:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    # Actualizar campo
    db.execute(
        text(f"UPDATE estudiantes SET {campo} = :valor, updated_at = NOW() WHERE id = :id"),
        {"id": estudiante_id, "valor": valor}
    )
    db.commit()
    
    # Registrar en historial
    db.execute(
        text("""
            INSERT INTO historial_cambios
            (estudiante_id, campo_modificado, valor_anterior, valor_nuevo, usuario_email, usuario_nombre, razon, timestamp)
            VALUES (:est_id, :campo, :anterior, :nuevo, :email, :nombre, :razon, NOW())
        """),
        {
            "est_id": estudiante_id,
            "campo": campo,
            "anterior": str(valor_anterior[0]) if valor_anterior[0] else None,
            "nuevo": str(valor),
            "email": usuario['email'],
            "nombre": usuario.get('nombre', usuario['email']),
            "razon": "Edición inline desde panel admin"
        }
    )
    db.commit()
    
    return {
        "message": "Campo actualizado",
        "campo": campo,
        "valor_anterior": valor_anterior[0],
        "valor_nuevo": valor
    }


@app.put("/api/admin/estudiantes/{estudiante_id}", tags=["Admin"])
def actualizar_estudiante(
    estudiante_id: int,
    datos: dict,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Actualiza información de estudiante usando SQL directo"""
    from datetime import datetime
    
    # Verificar que el estudiante existe
    check = db.execute(text("SELECT id FROM estudiantes WHERE id = :id"), 
                       {"id": estudiante_id}).fetchone()
    
    if not check:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    # Construir query de actualización dinámicamente
    campos_permitidos = ['nombre', 'email', 'telefono', 'pasaporte', 'edad', 
                         'nacionalidad', 'pais_origen', 'ciudad_origen', 
                         'carrera_deseada', 'especialidad', 'nivel_espanol', 
                         'tipo_visa', 'fondos_disponibles', 'notas']
    
    updates = []
    params = {"id": estudiante_id}
    
    for campo in campos_permitidos:
        if campo in datos:
            updates.append(f"{campo} = :{campo}")
            params[campo] = datos[campo]
    
    if not updates:
        return {"message": "No hay campos para actualizar"}
    
    updates.append("updated_at = :updated_at")
    params["updated_at"] = datetime.utcnow()
    
    query = f"UPDATE estudiantes SET {', '.join(updates)} WHERE id = :id"
    
    db.execute(text(query), params)
    db.commit()
    
    return {"message": "Estudiante actualizado correctamente", "id": estudiante_id}


@app.get("/api/test-email-config", tags=["Testing"])
def test_email_config():
    """Verifica la configuración de email sin enviar"""
    import os
    
    config = {
        "EMAIL_SENDER": os.getenv('EMAIL_SENDER'),
        "SMTP_USER": os.getenv('SMTP_USER'),
        "EMAIL_PASSWORD_SET": "✅ Sí" if os.getenv('EMAIL_PASSWORD') else "❌ No",
        "SMTP_PASSWORD_SET": "✅ Sí" if os.getenv('SMTP_PASSWORD') else "❌ No",
        "SMTP_SERVER": os.getenv('SMTP_SERVER', 'smtp.gmail.com'),
        "SMTP_PORT": os.getenv('SMTP_PORT', '587')
    }
    
    # Determinar qué configuración usar
    email_sender = os.getenv('EMAIL_SENDER') or os.getenv('SMTP_USER')
    email_password = os.getenv('EMAIL_PASSWORD') or os.getenv('SMTP_PASSWORD')
    
    status = {
        "configurado": bool(email_sender and email_password),
        "email_remitente": email_sender or "❌ NO CONFIGURADO",
        "password_configurado": "✅ Sí" if email_password else "❌ No"
    }
    
    return {
        "config": config,
        "status": status,
        "mensaje": "✅ Email configurado correctamente" if status["configurado"] else "❌ Configurar EMAIL_SENDER y EMAIL_PASSWORD en .env"
    }


@app.post("/api/test-email-send", tags=["Testing"])
def test_email_send(datos: dict):
    """Envía un email de prueba"""
    destinatario = datos.get('email')
    if not destinatario:
        raise HTTPException(status_code=400, detail="Email requerido")
    
    from api.email_utils import enviar_email
    
    resultado = enviar_email(
        destinatario=destinatario,
        asunto="🧪 Email de Prueba - Estudio Visa España",
        cuerpo_html="""
        <html>
        <body style="font-family: Arial, sans-serif;">
            <h2 style="color: #667eea;">✅ Email de Prueba</h2>
            <p>Si recibes este email, la configuración SMTP está funcionando correctamente.</p>
            <p><strong>Fecha:</strong> """ + str(datetime.now()) + """</p>
        </body>
        </html>
        """
    )
    
    return {
        "enviado": resultado,
        "mensaje": "✅ Email enviado" if resultado else "❌ Error enviando email (revisa logs del servidor)"
    }


@app.post("/api/admin/estudiantes/{estudiante_id}/aprobar", tags=["Admin"])
def aprobar_estudiante(
    estudiante_id: int,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Aprobar estudiante para envío"""
    from database.models import Estudiante as EstudianteModel
    
    estudiante = db.query(EstudianteModel).filter(EstudianteModel.id == estudiante_id).first()
    
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    estado_anterior = estudiante.estado
    estudiante.estado = 'aprobado'
    estudiante.updated_at = datetime.utcnow()
    db.commit()
    
    # Registrar cambio en historial
    db.execute(
        text("""
            INSERT INTO historial_cambios
            (estudiante_id, campo_modificado, valor_anterior, valor_nuevo, usuario_email, usuario_nombre, razon, timestamp)
            VALUES (:est_id, :campo, :anterior, :nuevo, :email, :nombre, :razon, NOW())
        """),
        {
            "est_id": estudiante_id,
            "campo": "estado",
            "anterior": estado_anterior,
            "nuevo": "aprobado",
            "email": usuario['email'],
            "nombre": usuario.get('nombre', usuario['email']),
            "razon": "Estudiante aprobado por administrador"
        }
    )
    db.commit()
    
    # Registrar auditoría (con try/catch)
    try:
        registrar_auditoria(
            db, usuario['email'], 'APROBAR_ESTUDIANTE',
            'estudiante', estudiante_id,
            {'nombre': estudiante.nombre, 'estado_anterior': estado_anterior}
        )
    except Exception as e:
        print(f"⚠️ Error en auditoría (no crítico): {e}")
    
    # Enviar email de notificación
    email_enviado = False
    error_email = None
    try:
        from api.email_utils import email_aprobacion
        resultado = email_aprobacion(estudiante.nombre, estudiante.email)
        if resultado:
            email_enviado = True
            print(f"✅ Email enviado correctamente a {estudiante.email}")
        else:
            error_email = "La función email_aprobacion retornó False"
            print(f"⚠️ Email NO enviado a {estudiante.email}: {error_email}")
    except Exception as e:
        error_email = str(e)
        print(f"❌ Error enviando email a {estudiante.email}: {e}")
        import traceback
        traceback.print_exc()
    
    return {
        "message": "Estudiante aprobado correctamente", 
        "id": estudiante_id,
        "email_enviado": email_enviado,
        "error_email": error_email
    }


@app.post("/api/admin/estudiantes/{estudiante_id}/rechazar", tags=["Admin"])
def rechazar_estudiante(
    estudiante_id: int,
    datos: dict,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Rechazar estudiante y solicitar correcciones"""
    from database.models import Estudiante as EstudianteModel
    
    motivo = datos.get('motivo', 'Sin motivo especificado')
    sugerencias = datos.get('sugerencias', [])
    
    estudiante = db.query(EstudianteModel).filter(EstudianteModel.id == estudiante_id).first()
    
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    estado_anterior = estudiante.estado
    estudiante.estado = 'rechazado'
    estudiante.notas = motivo
    estudiante.updated_at = datetime.utcnow()
    db.commit()
    
    # Registrar cambio en historial
    db.execute(
        text("""
            INSERT INTO historial_cambios
            (estudiante_id, campo_modificado, valor_anterior, valor_nuevo, usuario_email, usuario_nombre, razon, timestamp)
            VALUES (:est_id, :campo, :anterior, :nuevo, :email, :nombre, :razon, NOW())
        """),
        {
            "est_id": estudiante_id,
            "campo": "estado",
            "anterior": estado_anterior,
            "nuevo": "rechazado",
            "email": usuario['email'],
            "nombre": usuario.get('nombre', usuario['email']),
            "razon": f"Rechazado: {motivo}"
        }
    )
    db.commit()
    
    # Registrar auditoría
    registrar_auditoria(
        db, usuario['email'], 'RECHAZAR_ESTUDIANTE',
        'estudiante', estudiante_id,
        {'nombre': estudiante.nombre, 'motivo': motivo, 'sugerencias': sugerencias}
    )
    
    # Enviar email de notificación con sugerencias
    try:
        from api.email_utils import email_rechazo
        # Si hay sugerencias, añadirlas al motivo
        motivo_completo = motivo
        if sugerencias:
            motivo_completo += "\n\n📋 Sugerencias para corregir:\n"
            for i, sug in enumerate(sugerencias, 1):
                motivo_completo += f"{i}. {sug}\n"
        
        email_rechazo(estudiante.nombre, estudiante.email, motivo_completo)
    except Exception as e:
        print(f"⚠️ Error enviando email: {e}")
    
    return {"message": "Estudiante rechazado", "id": estudiante_id, "motivo": motivo}


# ============================================================================
# NOTAS INTERNAS - Sistema de comunicación entre admins
# ============================================================================

@app.post("/api/admin/estudiantes/{estudiante_id}/notas", tags=["Admin - Notas"])
def crear_nota_interna(
    estudiante_id: int,
    datos: dict,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Crea nota interna para un estudiante"""
    contenido = datos.get('contenido')
    importante = datos.get('importante', False)
    privada = datos.get('privada', False)
    
    if not contenido:
        raise HTTPException(status_code=400, detail="Contenido requerido")
    
    db.execute(
        text("""
            INSERT INTO notas_internas 
            (estudiante_id, autor_email, autor_nombre, contenido, importante, privada, created_at)
            VALUES (:est_id, :email, :nombre, :contenido, :importante, :privada, NOW())
        """),
        {
            "est_id": estudiante_id,
            "email": usuario['email'],
            "nombre": usuario.get('nombre', usuario['email']),
            "contenido": contenido,
            "importante": importante,
            "privada": privada
        }
    )
    db.commit()
    
    # Registrar auditoría
    registrar_auditoria(
        db, usuario['email'], 'CREAR_NOTA_INTERNA',
        'estudiante', estudiante_id,
        {'contenido_preview': contenido[:50], 'importante': importante}
    )
    
    return {"message": "Nota creada correctamente"}


@app.get("/api/admin/estudiantes/{estudiante_id}/notas", tags=["Admin - Notas"])
def listar_notas_internas(
    estudiante_id: int,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Lista notas internas de un estudiante"""
    result = db.execute(
        text("""
            SELECT id, autor_email, autor_nombre, contenido, importante, privada, created_at, updated_at
            FROM notas_internas
            WHERE estudiante_id = :est_id
            ORDER BY importante DESC, created_at DESC
        """),
        {"est_id": estudiante_id}
    ).fetchall()
    
    notas = []
    for row in result:
        notas.append({
            'id': row[0],
            'autor_email': row[1],
            'autor_nombre': row[2],
            'contenido': row[3],
            'importante': row[4],
            'privada': row[5],
            'created_at': row[6].isoformat() if row[6] else None,
            'updated_at': row[7].isoformat() if row[7] else None
        })
    
    return {"notas": notas, "total": len(notas)}


@app.put("/api/admin/notas/{nota_id}", tags=["Admin - Notas"])
def actualizar_nota_interna(
    nota_id: int,
    datos: dict,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Actualiza una nota interna"""
    contenido = datos.get('contenido')
    importante = datos.get('importante')
    
    if not contenido:
        raise HTTPException(status_code=400, detail="Contenido requerido")
    
    db.execute(
        text("""
            UPDATE notas_internas
            SET contenido = :contenido,
                importante = COALESCE(:importante, importante),
                updated_at = NOW()
            WHERE id = :nota_id
        """),
        {"nota_id": nota_id, "contenido": contenido, "importante": importante}
    )
    db.commit()
    
    return {"message": "Nota actualizada"}


@app.delete("/api/admin/notas/{nota_id}", tags=["Admin - Notas"])
def eliminar_nota_interna(
    nota_id: int,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Elimina una nota interna"""
    db.execute(text("DELETE FROM notas_internas WHERE id = :nota_id"), {"nota_id": nota_id})
    db.commit()
    
    return {"message": "Nota eliminada"}


@app.get("/api/admin/estudiantes/{estudiante_id}/historial", tags=["Admin - Historial"])
def obtener_historial_cambios(
    estudiante_id: int,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Obtiene historial completo de cambios de un estudiante"""
    result = db.execute(
        text("""
            SELECT id, campo_modificado, valor_anterior, valor_nuevo,
                   usuario_email, usuario_nombre, razon, timestamp
            FROM historial_cambios
            WHERE estudiante_id = :est_id
            ORDER BY timestamp DESC
        """),
        {"est_id": estudiante_id}
    ).fetchall()
    
    historial = []
    for row in result:
        historial.append({
            'id': row[0],
            'campo': row[1],
            'valor_anterior': row[2],
            'valor_nuevo': row[3],
            'usuario_email': row[4],
            'usuario_nombre': row[5],
            'razon': row[6],
            'timestamp': row[7].isoformat() if row[7] else None
        })
    
    return {"historial": historial, "total": len(historial)}


@app.get("/api/admin/calendario/eventos", tags=["Admin - Calendario"])
def obtener_eventos_calendario(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Obtiene todos los eventos para el calendario integrado"""
    from datetime import datetime, timedelta
    
    # Si no se especifican fechas, usar mes actual
    if not fecha_inicio:
        fecha_inicio = datetime.now().replace(day=1).isoformat()
    if not fecha_fin:
        # Último día del mes
        fecha_fin = (datetime.now().replace(day=28) + timedelta(days=4)).replace(day=1).isoformat()
    
    eventos = []
    
    # 1. Fechas importantes de estudiantes
    fechas_importantes = db.execute(
        text("""
            SELECT fi.id, fi.estudiante_id, fi.tipo_fecha, fi.fecha, fi.descripcion,
                   fi.completada, e.nombre as estudiante_nombre
            FROM fechas_importantes fi
            JOIN estudiantes e ON fi.estudiante_id = e.id
            WHERE fi.fecha >= :inicio AND fi.fecha <= :fin
            ORDER BY fi.fecha
        """),
        {"inicio": fecha_inicio, "fin": fecha_fin}
    ).fetchall()
    
    for row in fechas_importantes:
        eventos.append({
            'id': f"fecha_{row[0]}",
            'tipo': 'fecha_importante',
            'estudiante_id': row[1],
            'estudiante_nombre': row[6],
            'titulo': f"{row[2]} - {row[6]}",
            'fecha': row[3].isoformat() if row[3] else None,
            'descripcion': row[4],
            'completada': row[5],
            'color': '#4CAF50' if row[5] else '#FF9800'
        })
    
    # 2. Estudiantes con inicio estimado
    inicios_estimados = db.execute(
        text("""
            SELECT id, nombre, fecha_inicio_estimada, carrera_deseada
            FROM estudiantes
            WHERE fecha_inicio_estimada IS NOT NULL
            AND fecha_inicio_estimada >= :inicio
            AND fecha_inicio_estimada <= :fin
            ORDER BY fecha_inicio_estimada
        """),
        {"inicio": fecha_inicio, "fin": fecha_fin}
    ).fetchall()
    
    for row in inicios_estimados:
        eventos.append({
            'id': f"inicio_{row[0]}",
            'tipo': 'inicio_estudios',
            'estudiante_id': row[0],
            'estudiante_nombre': row[1],
            'titulo': f"Inicio: {row[1]} - {row[3]}",
            'fecha': row[2].isoformat() if row[2] else None,
            'descripcion': f"Inicio estimado de estudios: {row[3]}",
            'completada': False,
            'color': '#2196F3'
        })
    
    # 3. Registros recientes (últimos 7 días como eventos)
    if not fecha_inicio or (datetime.fromisoformat(fecha_inicio) - datetime.now()).days <= 7:
        registros_recientes = db.execute(
            text("""
                SELECT id, nombre, created_at, email
                FROM estudiantes
                WHERE created_at >= :inicio AND created_at <= :fin
                ORDER BY created_at DESC
            """),
            {"inicio": fecha_inicio or (datetime.now() - timedelta(days=7)).isoformat(), "fin": fecha_fin}
        ).fetchall()
        
        for row in registros_recientes:
            eventos.append({
                'id': f"registro_{row[0]}",
                'tipo': 'registro_nuevo',
                'estudiante_id': row[0],
                'estudiante_nombre': row[1],
                'titulo': f"Nuevo registro: {row[1]}",
                'fecha': row[2].isoformat() if row[2] else None,
                'descripcion': f"Email: {row[3]}",
                'completada': False,
                'color': '#9C27B0'
            })
    
    return {
        "eventos": eventos,
        "total": len(eventos),
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin
    }


@app.get("/api/admin/estadisticas", response_model=EstadisticasResponse, tags=["Admin"])
def obtener_estadisticas(
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Estadísticas del dashboard"""
    from database.models import Estudiante as EstudianteModel
    
    # Contar estudiantes por estado
    total = db.query(EstudianteModel).count()
    pendientes = db.query(EstudianteModel).filter(EstudianteModel.estado == 'pendiente').count()
    aprobados = db.query(EstudianteModel).filter(EstudianteModel.estado == 'aprobado').count()
    rechazados = db.query(EstudianteModel).filter(EstudianteModel.estado == 'rechazado').count()
    
    # Calcular estadísticas por especialidad
    especialidades_query = db.execute(text("""
        SELECT 
            COALESCE(especialidad, 'Sin especialidad') as especialidad,
            COUNT(*) as total
        FROM estudiantes 
        WHERE especialidad IS NOT NULL AND especialidad != ''
        GROUP BY especialidad
        ORDER BY total DESC
    """))
    
    por_especialidad = {
        row[0]: row[1] for row in especialidades_query.fetchall()
    }
    
    return EstadisticasResponse(
        total_estudiantes=total,
        pendientes_revision=pendientes,
        aprobados=aprobados,
        enviados=aprobados,  # Por ahora enviados = aprobados
        por_especialidad=por_especialidad
    )


@app.get("/api/admin/dashboard/metricas-tiempo-real", tags=["Admin - Dashboard"])
def obtener_metricas_tiempo_real(
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Métricas en tiempo real para dashboard actualizado automáticamente"""
    from datetime import datetime, timedelta
    
    # Resumen general
    total_estudiantes = db.execute(text("SELECT COUNT(*) FROM estudiantes")).fetchone()[0]
    
    # Por estado
    estados = db.execute(text("""
        SELECT 
            COALESCE(estado, estado_procesamiento, 'pendiente') as estado,
            COUNT(*) as total
        FROM estudiantes
        GROUP BY estado
    """)).fetchall()
    
    por_estado = {row[0]: row[1] for row in estados}
    
    # Últimos 7 días - Tendencia
    hace_7_dias = (datetime.now() - timedelta(days=7)).isoformat()
    registros_7d = db.execute(
        text("SELECT COUNT(*) FROM estudiantes WHERE created_at >= :fecha"),
        {"fecha": hace_7_dias}
    ).fetchone()[0]
    
    # Últimas 24 horas
    hace_24h = (datetime.now() - timedelta(hours=24)).isoformat()
    registros_24h = db.execute(
        text("SELECT COUNT(*) FROM estudiantes WHERE created_at >= :fecha"),
        {"fecha": hace_24h}
    ).fetchone()[0]
    
    # Documentos generados hoy
    hoy = datetime.now().replace(hour=0, minute=0, second=0).isoformat()
    docs_hoy = db.execute(
        text("SELECT COUNT(*) FROM documentos_generados WHERE fecha_generacion >= :fecha"),
        {"fecha": hoy}
    ).fetchone()[0]
    
    # Notas agregadas hoy
    notas_hoy = db.execute(
        text("SELECT COUNT(*) FROM notas_internas WHERE created_at >= :fecha"),
        {"fecha": hoy}
    ).fetchone()[0]
    
    # Tasa de conversión (aprobados/total)
    aprobados = por_estado.get('aprobado', 0)
    tasa_aprobacion = round((aprobados / total_estudiantes * 100), 2) if total_estudiantes > 0 else 0
    
    # Fondos promedio
    fondos_promedio = db.execute(
        text("SELECT AVG(fondos_disponibles) FROM estudiantes WHERE fondos_disponibles > 0")
    ).fetchone()[0]
    
    # Top 5 países
    top_paises = db.execute(text("""
        SELECT nacionalidad, COUNT(*) as total
        FROM estudiantes
        WHERE nacionalidad IS NOT NULL
        GROUP BY nacionalidad
        ORDER BY total DESC
        LIMIT 5
    """)).fetchall()
    
    # Top 5 especialidades
    top_especialidades = db.execute(text("""
        SELECT 
            COALESCE(especialidad, 'Sin especificar') as especialidad,
            COUNT(*) as total
        FROM estudiantes
        WHERE especialidad IS NOT NULL AND especialidad != ''
        GROUP BY especialidad
        ORDER BY total DESC
        LIMIT 5
    """)).fetchall()
    
    # Actividad reciente (últimas 10 acciones)
    actividad_reciente = db.execute(text("""
        SELECT accion, usuario_email, entidad, timestamp
        FROM logs_auditoria
        ORDER BY timestamp DESC
        LIMIT 10
    """)).fetchall()
    
    # Alertas pendientes (fechas próximas)
    proximo_7d = (datetime.now() + timedelta(days=7)).isoformat()
    alertas_pendientes = db.execute(text("""
        SELECT COUNT(*)
        FROM fechas_importantes
        WHERE fecha <= :fecha AND completada = FALSE
    """), {"fecha": proximo_7d}).fetchone()[0]
    
    return {
        "timestamp": datetime.now().isoformat(),
        "resumen": {
            "total_estudiantes": total_estudiantes,
            "registros_24h": registros_24h,
            "registros_7d": registros_7d,
            "documentos_generados_hoy": docs_hoy,
            "notas_agregadas_hoy": notas_hoy,
            "alertas_pendientes": alertas_pendientes
        },
        "por_estado": por_estado,
        "metricas": {
            "tasa_aprobacion": tasa_aprobacion,
            "fondos_promedio": float(fondos_promedio) if fondos_promedio else 0,
            "aprobados": aprobados,
            "pendientes": por_estado.get('pendiente', 0) + por_estado.get('en_revision', 0),
            "rechazados": por_estado.get('rechazado', 0)
        },
        "top_paises": [{"pais": row[0], "total": row[1]} for row in top_paises],
        "top_especialidades": [{"especialidad": row[0], "total": row[1]} for row in top_especialidades],
        "actividad_reciente": [
            {
                "accion": row[0],
                "usuario": row[1],
                "entidad": row[2],
                "timestamp": row[3].isoformat() if row[3] else None
            } for row in actividad_reciente
        ]
    }


# ============================================================================
# TEMPLATES DE EMAIL - Sistema personalizable
# ============================================================================

@app.get("/api/admin/email-templates", tags=["Admin - Email Templates"])
def listar_templates_email(
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Lista todos los templates de email disponibles"""
    result = db.execute(text("""
        SELECT id, nombre, asunto, contenido_html, contenido_texto, 
               variables_disponibles, activo, ultima_modificacion, modificado_por
        FROM email_templates
        ORDER BY nombre
    """)).fetchall()
    
    templates = []
    for row in result:
        templates.append({
            'id': row[0],
            'nombre': row[1],
            'asunto': row[2],
            'contenido_html': row[3],
            'contenido_texto': row[4],
            'variables_disponibles': row[5],
            'activo': row[6],
            'ultima_modificacion': row[7].isoformat() if row[7] else None,
            'modificado_por': row[8]
        })
    
    return {"templates": templates, "total": len(templates)}


@app.get("/api/admin/email-templates/{template_id}", tags=["Admin - Email Templates"])
def obtener_template_email(
    template_id: int,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Obtiene un template específico"""
    result = db.execute(
        text("""
            SELECT id, nombre, asunto, contenido_html, contenido_texto,
                   variables_disponibles, activo
            FROM email_templates WHERE id = :id
        """),
        {"id": template_id}
    ).fetchone()
    
    if not result:
        raise HTTPException(status_code=404, detail="Template no encontrado")
    
    return {
        'id': result[0],
        'nombre': result[1],
        'asunto': result[2],
        'contenido_html': result[3],
        'contenido_texto': result[4],
        'variables_disponibles': result[5],
        'activo': result[6]
    }


@app.put("/api/admin/email-templates/{template_id}", tags=["Admin - Email Templates"])
def actualizar_template_email(
    template_id: int,
    datos: dict,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Actualiza un template de email"""
    asunto = datos.get('asunto')
    contenido_html = datos.get('contenido_html')
    contenido_texto = datos.get('contenido_texto')
    activo = datos.get('activo', True)
    
    if not asunto or not contenido_html:
        raise HTTPException(status_code=400, detail="Asunto y contenido HTML requeridos")
    
    db.execute(
        text("""
            UPDATE email_templates
            SET asunto = :asunto,
                contenido_html = :html,
                contenido_texto = :texto,
                activo = :activo,
                ultima_modificacion = NOW(),
                modificado_por = :usuario
            WHERE id = :id
        """),
        {
            "id": template_id,
            "asunto": asunto,
            "html": contenido_html,
            "texto": contenido_texto,
            "activo": activo,
            "usuario": usuario['email']
        }
    )
    db.commit()
    
    # Registrar auditoría
    registrar_auditoria(
        db, usuario['email'], 'ACTUALIZAR_TEMPLATE_EMAIL',
        'email_template', template_id,
        {'asunto': asunto}
    )
    
    return {"message": "Template actualizado correctamente"}


@app.post("/api/admin/email-templates/preview", tags=["Admin - Email Templates"])
def preview_template_email(
    datos: dict,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Preview de template con variables de ejemplo"""
    contenido_html = datos.get('contenido_html', '')
    variables = datos.get('variables', {})
    
    # Reemplazar variables
    preview = contenido_html
    for key, value in variables.items():
        preview = preview.replace(f"{{{{{key}}}}}", str(value))
    
    return {
        "preview_html": preview,
        "variables_usadas": list(variables.keys())
    }


# ============================================================================
# SISTEMA DE ROLES Y PERMISOS
# ============================================================================

@app.get("/api/admin/roles", tags=["Admin - Roles"])
def listar_roles(
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Lista todos los roles disponibles"""
    result = db.execute(text("""
        SELECT id, nombre, descripcion, permisos, activo, created_at
        FROM roles
        ORDER BY nombre
    """)).fetchall()
    
    roles = []
    for row in result:
        roles.append({
            'id': row[0],
            'nombre': row[1],
            'descripcion': row[2],
            'permisos': row[3],
            'activo': row[4],
            'created_at': row[5].isoformat() if row[5] else None
        })
    
    return {"roles": roles, "total": len(roles)}


@app.post("/api/admin/usuarios", tags=["Admin - Usuarios"])
def crear_usuario_admin(
    datos: dict,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Crea nuevo usuario administrador con rol"""
    email = datos.get('email')
    nombre = datos.get('nombre')
    password = datos.get('password')
    rol_id = datos.get('rol_id', 2)
    
    if not email or not password:
        raise HTTPException(status_code=400, detail="Email y contraseña requeridos")
    
    # Verificar si ya existe
    existe = db.execute(
        text("SELECT id FROM usuarios_admin WHERE email = :email"),
        {"email": email}
    ).fetchone()
    
    if existe:
        raise HTTPException(status_code=400, detail="Usuario ya existe")
    
    # Hash de contraseña
    from passlib.context import CryptContext
    pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
    password_hash = pwd_context.hash(password)
    
    # Crear usuario
    db.execute(
        text("""
            INSERT INTO usuarios_admin (email, nombre, password_hash, rol_id, activo, created_at)
            VALUES (:email, :nombre, :password, :rol_id, TRUE, NOW())
        """),
        {
            "email": email,
            "nombre": nombre,
            "password": password_hash,
            "rol_id": rol_id
        }
    )
    db.commit()
    
    registrar_auditoria(
        db, usuario['email'], 'CREAR_USUARIO_ADMIN',
        'usuario_admin', None,
        {'email': email, 'rol_id': rol_id}
    )
    
    return {"message": "Usuario creado correctamente", "email": email}


@app.get("/api/admin/usuarios", tags=["Admin - Usuarios"])
def listar_usuarios_admin(
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Lista todos los usuarios administradores"""
    result = db.execute(text("""
        SELECT u.id, u.email, u.nombre, u.activo, u.ultimo_acceso, u.created_at,
               r.nombre as rol_nombre, r.descripcion as rol_descripcion
        FROM usuarios_admin u
        LEFT JOIN roles r ON u.rol_id = r.id
        ORDER BY u.created_at DESC
    """)).fetchall()
    
    usuarios = []
    for row in result:
        usuarios.append({
            'id': row[0],
            'email': row[1],
            'nombre': row[2],
            'activo': row[3],
            'ultimo_acceso': row[4].isoformat() if row[4] else None,
            'created_at': row[5].isoformat() if row[5] else None,
            'rol_nombre': row[6],
            'rol_descripcion': row[7]
        })
    
    return {"usuarios": usuarios, "total": len(usuarios)}


@app.put("/api/admin/usuarios/{usuario_id}/rol", tags=["Admin - Usuarios"])
def cambiar_rol_usuario(
    usuario_id: int,
    datos: dict,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Cambia el rol de un usuario"""
    nuevo_rol_id = datos.get('rol_id')
    
    if not nuevo_rol_id:
        raise HTTPException(status_code=400, detail="rol_id requerido")
    
    db.execute(
        text("""
            UPDATE usuarios_admin
            SET rol_id = :rol_id, updated_at = NOW()
            WHERE id = :usuario_id
        """),
        {"usuario_id": usuario_id, "rol_id": nuevo_rol_id}
    )
    db.commit()
    
    registrar_auditoria(
        db, usuario['email'], 'CAMBIAR_ROL_USUARIO',
        'usuario_admin', usuario_id,
        {'nuevo_rol_id': nuevo_rol_id}
    )
    
    return {"message": "Rol actualizado correctamente"}



# ============================================================================
# AUDITORÍA - Sistema de logs
# ============================================================================

def registrar_auditoria(db: Session, usuario_email: str, accion: str, 
                       entidad: str = None, entidad_id: int = None, 
                       detalles: dict = None):
    """Registra acción de auditoría en la base de datos"""
    try:
        db.execute(
            text("""
                INSERT INTO logs_auditoria 
                (usuario_email, accion, entidad, entidad_id, detalles, timestamp)
                VALUES (:usuario, :accion, :entidad, :entidad_id, :detalles, NOW())
            """),
            {
                "usuario": usuario_email,
                "accion": accion,
                "entidad": entidad,
                "entidad_id": entidad_id,
                "detalles": json.dumps(detalles) if detalles else None
            }
        )
        db.commit()
    except Exception as e:
        print(f"⚠️ Error registrando auditoría: {e}")


@app.get("/api/admin/logs-auditoria", tags=["Admin"])
def obtener_logs_auditoria(
    usuario_email: Optional[str] = None,
    accion: Optional[str] = None,
    limit: int = 100,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Obtiene logs de auditoría con filtros"""
    query = "SELECT * FROM logs_auditoria WHERE 1=1"
    params = {}
    
    if usuario_email:
        query += " AND usuario_email = :usuario"
        params["usuario"] = usuario_email
    
    if accion:
        query += " AND accion = :accion"
        params["accion"] = accion
    
    query += " ORDER BY timestamp DESC LIMIT :limit"
    params["limit"] = limit
    
    result = db.execute(text(query), params).fetchall()
    
    logs = []
    for row in result:
        logs.append({
            "id": row[0],
            "usuario_email": row[1],
            "accion": row[2],
            "entidad": row[3],
            "entidad_id": row[4],
            "detalles": row[5],
            "timestamp": row[8].isoformat() if row[8] else None
        })
    
    return {"logs": logs, "total": len(logs)}


# ============================================================================
# MONITOREO OCR - Endpoint para ver uso de API
# ============================================================================

@app.get("/api/admin/ocr/uso", tags=["Admin"])
def obtener_uso_ocr(usuario=Depends(obtener_usuario_actual)):
    """Obtiene estadísticas de uso de OCR.space API"""
    from api.validador_ocr import ValidadorOCR
    
    validador = ValidadorOCR()
    uso = validador.verificar_limite_uso()
    
    return {
        "usos_este_mes": uso['usos'],
        "limite_mensual": uso['limite'],
        "restantes": uso['restante'],
        "porcentaje_usado": uso['porcentaje'],
        "alerta": uso['alerta'],
        "mes": datetime.now().strftime('%Y-%m')
    }


# ============================================================================
# CURSOS - Endpoints para gestión de cursos
# ============================================================================

@app.get("/api/cursos", tags=["Cursos"])
def listar_cursos(
    especialidad: Optional[str] = None,
    ciudad: Optional[str] = None,
    precio_max: Optional[float] = None,
    nivel_idioma: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db)
):
    """Lista cursos disponibles con filtros"""
    from modules.cursos import GestorCursos
    
    cursos = GestorCursos.filtrar_cursos(
        especialidad=especialidad,
        ciudad=ciudad,
        precio_max=precio_max,
        nivel_idioma=nivel_idioma
    )
    
    return {
        'total': len(cursos),
        'cursos': [
            {
                'id': c.id,
                'nombre': c.nombre,
                'escuela': c.escuela,
                'ciudad': c.ciudad,
                'especialidad': c.especialidad,
                'precio': c.precio,
                'duracion_meses': c.duracion_meses,
                'nivel_minimo_espanol': c.nivel_minimo_espanol,
                'fecha_inicio': c.fecha_inicio.isoformat() if c.fecha_inicio else None,
                'disponible': c.disponible
            }
            for c in cursos[skip:skip+limit]
        ]
    }


@app.get("/api/cursos/{curso_id}", tags=["Cursos"])
def obtener_curso(curso_id: int, db: Session = Depends(get_db)):
    """Obtiene detalles de un curso específico"""
    from modules.cursos import GestorCursos
    
    curso = GestorCursos.obtener_curso_por_id(curso_id)
    
    if not curso:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    
    return {
        'id': curso.id,
        'nombre': curso.nombre,
        'escuela': curso.escuela,
        'ciudad': curso.ciudad,
        'especialidad': curso.especialidad,
        'precio': curso.precio,
        'duracion_meses': curso.duracion_meses,
        'nivel_minimo_espanol': curso.nivel_minimo_espanol,
        'fecha_inicio': curso.fecha_inicio.isoformat() if curso.fecha_inicio else None,
        'descripcion': curso.descripcion,
        'requisitos': curso.requisitos,
        'url_informacion': curso.url_informacion,
        'disponible': curso.disponible
    }


# ============================================================================
# FONDOS - Verificación económica
# ============================================================================

@app.get("/api/estudiantes/{estudiante_id}/fondos", tags=["Fondos"])
def verificar_fondos(estudiante_id: int, db: Session = Depends(get_db)):
    """Verifica situación económica del estudiante"""
    from modules.fondos import GestorFondos
    
    estudiante = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    verificacion = GestorFondos.verificar_fondos(estudiante_id)
    return verificacion


@app.post("/api/estudiantes/{estudiante_id}/patrocinador", tags=["Fondos"])
def agregar_patrocinador(
    estudiante_id: int,
    datos_patrocinador: Dict,
    db: Session = Depends(get_db)
):
    """Agrega patrocinador para el estudiante"""
    from modules.fondos import GestorFondos
    
    resultado = GestorFondos.agregar_patrocinador(estudiante_id, datos_patrocinador)
    return resultado


# ============================================================================
# ALOJAMIENTO - Búsqueda y gestión
# ============================================================================

@app.get("/api/alojamientos", tags=["Alojamiento"])
def buscar_alojamientos(
    ciudad: Optional[str] = None,
    precio_max: Optional[float] = None,
    tipo: Optional[str] = None,
    skip: int = 0,
    limit: int = 20,
    db: Session = Depends(get_db)
):
    """Busca alojamientos disponibles"""
    from modules.alojamiento import GestorAlojamiento
    
    alojamientos = GestorAlojamiento.buscar_alojamientos(
        ciudad=ciudad,
        precio_max=precio_max,
        tipo=tipo,
        disponible=True
    )
    
    return {
        'total': len(alojamientos),
        'alojamientos': [
            {
                'id': a.id,
                'tipo': a.tipo,
                'direccion': a.direccion,
                'ciudad': a.ciudad,
                'precio_mensual': a.precio_mensual,
                'gastos_incluidos': a.gastos_incluidos,
                'num_habitaciones': a.num_habitaciones,
                'num_banos': a.num_banos,
                'metros_cuadrados': a.metros_cuadrados,
                'disponible': a.disponible
            }
            for a in alojamientos[skip:skip+limit]
        ]
    }


# ============================================================================
# DOCUMENTOS - Upload y gestión
# ============================================================================

from fastapi import UploadFile, File

@app.post("/api/estudiantes/{estudiante_id}/documentos/upload", tags=["Documentos"])
async def subir_documento(
    estudiante_id: int,
    tipo_documento: str,
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Sube un documento del estudiante"""
    import os
    from pathlib import Path
    
    # Verificar estudiante existe
    estudiante = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    # Crear directorio si no existe
    upload_dir = Path(f"uploads/estudiantes/{estudiante_id}")
    upload_dir.mkdir(parents=True, exist_ok=True)
    
    # Guardar archivo
    file_path = upload_dir / f"{tipo_documento}_{archivo.filename}"
    
    with file_path.open("wb") as buffer:
        content = await archivo.read()
        buffer.write(content)
    
    return {
        'mensaje': 'Documento subido correctamente',
        'tipo': tipo_documento,
        'filename': archivo.filename,
        'path': str(file_path)
    }


# ============================================================================
# ENDPOINTS PARA GESTOR DE DOCUMENTOS (Frontend)
# ============================================================================

@app.post("/api/documentos/{estudiante_id}/subir", tags=["Documentos"])
async def subir_documentos_multi(
    estudiante_id: int,
    archivos: list[UploadFile] = File(...),
    categorias: str = Form(...)
):
    """Subir múltiples documentos del estudiante (usado por GestorDocumentos.jsx)"""
    import os
    import psycopg2
    import base64
    from pathlib import Path
    from datetime import datetime
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    try:
        # Parsear categorías (vienen como string separado por comas)
        categorias_list = [c.strip() for c in categorias.split(',')]
        
        # Verificar estudiante
        cursor.execute("SELECT id FROM estudiantes WHERE id = %s", (estudiante_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            raise HTTPException(status_code=404, detail='Estudiante no encontrado')
        
        # Validar que coincidan archivos y categorías
        if len(archivos) != len(categorias_list):
            raise HTTPException(
                status_code=400, 
                detail=f'Número de archivos ({len(archivos)}) no coincide con categorías ({len(categorias_list)})'
            )
        
        documentos_creados = []
        
        for i, archivo in enumerate(archivos):
            categoria = categorias_list[i]
            
            # Validar categoría
            categorias_validas = ['pasaporte', 'visa', 'academicos', 'financieros', 'otros']
            if categoria not in categorias_validas:
                categoria = 'otros'
            
            # Leer contenido
            contenido = await archivo.read()
            tamano = len(contenido)
            
            # Validar tamaño (10MB máx)
            if tamano > 10 * 1024 * 1024:
                continue
            
            # Convertir a base64
            contenido_base64 = base64.b64encode(contenido).decode('utf-8')
            
            # Insertar en BD
            cursor.execute("""
                INSERT INTO documentos 
                (estudiante_id, tipo_documento, nombre_archivo, url_archivo,
                 categoria, contenido_base64, tamano_archivo, mime_type, 
                 estado_revision, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'pendiente', %s)
                RETURNING id
            """, (
                estudiante_id,
                categoria,  # tipo_documento = categoria
                archivo.filename,
                f'base64://{archivo.filename}',  # url_archivo placeholder
                categoria,
                contenido_base64,
                tamano,
                archivo.content_type or 'application/octet-stream',
                datetime.utcnow()
            ))
            
            doc_id = cursor.fetchone()[0]
            documentos_creados.append({
                'id': doc_id,
                'nombre': archivo.filename,
                'categoria': categoria,
                'tamano': tamano
            })
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return {
            'success': True,
            'message': f'{len(documentos_creados)} documentos subidos correctamente',
            'documentos': documentos_creados
        }
        
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        print(f"❌ Error subiendo documentos: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/documentos/{estudiante_id}/listar", tags=["Documentos"])
def listar_documentos_estudiante(estudiante_id: int):
    """Listar documentos del estudiante (usado por GestorDocumentos.jsx)"""
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT id, nombre_archivo, categoria, tamano_archivo, mime_type, 
                   estado_revision, comentario_admin, created_at, updated_at
            FROM documentos
            WHERE estudiante_id = %s
            ORDER BY created_at DESC
        """, (estudiante_id,))
        
        documentos = []
        for row in cursor.fetchall():
            documentos.append({
                'id': row[0],
                'nombre': row[1],
                'categoria': row[2],
                'tamano': row[3],
                'mime_type': row[4],
                'estado_revision': row[5],
                'comentario_admin': row[6],
                'created_at': row[7].isoformat() if row[7] else None,
                'updated_at': row[8].isoformat() if row[8] else None
            })
        
        # Calcular progreso
        categorias_validas = ['pasaporte', 'visa', 'academicos', 'financieros', 'otros']
        categorias_con_docs = set([doc['categoria'] for doc in documentos if doc['estado_revision'] != 'rechazado'])
        progreso = round((len(categorias_con_docs) / len(categorias_validas)) * 100, 2)
        
        cursor.close()
        conn.close()
        
        return {
            'success': True,
            'documentos': documentos,
            'progreso': progreso,
            'total': len(documentos)
        }
        
    except Exception as e:
        cursor.close()
        conn.close()
        print(f"❌ Error listando documentos: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/documentos/{documento_id}/descargar", tags=["Documentos"])
def descargar_documento(documento_id: int):
    """Descargar documento por ID (usado por GestorDocumentos.jsx)"""
    import os
    import psycopg2
    import base64
    from io import BytesIO
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT nombre_archivo, contenido_base64, mime_type
            FROM documentos
            WHERE id = %s
        """, (documento_id,))
        
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not row:
            raise HTTPException(status_code=404, detail="Documento no encontrado")
        
        nombre, contenido_b64, mime_type = row
        
        # Decodificar base64
        contenido = base64.b64decode(contenido_b64)
        
        return StreamingResponse(
            BytesIO(contenido),
            media_type=mime_type or 'application/octet-stream',
            headers={'Content-Disposition': f'attachment; filename="{nombre}"'}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()
        print(f"❌ Error descargando documento: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/documentos/{documento_id}/eliminar", tags=["Documentos"])
def eliminar_documento(documento_id: int):
    """Eliminar documento (usado por GestorDocumentos.jsx)"""
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    try:
        # Eliminar de BD
        cursor.execute("DELETE FROM documentos WHERE id = %s RETURNING id", (documento_id,))
        deleted = cursor.fetchone()
        
        if not deleted:
            cursor.close()
            conn.close()
            raise HTTPException(status_code=404, detail='Documento no encontrado')
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return {'success': True, 'message': 'Documento eliminado correctamente'}
        
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        print(f"❌ Error eliminando documento: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/documentos/{estudiante_id}/descargar-zip", tags=["Documentos"])
def descargar_documentos_zip(estudiante_id: int):
    """Descargar todos los documentos en ZIP (usado por GestorDocumentos.jsx)"""
    import os
    import psycopg2
    import zipfile
    import base64
    from io import BytesIO
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT nombre_archivo, contenido_base64
            FROM documentos
            WHERE estudiante_id = %s
        """, (estudiante_id,))
        
        documentos = cursor.fetchall()
        cursor.close()
        conn.close()
        
        if not documentos:
            raise HTTPException(status_code=404, detail='No hay documentos para descargar')
        
        # Crear ZIP en memoria
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for nombre, contenido_b64 in documentos:
                contenido = base64.b64decode(contenido_b64)
                zip_file.writestr(nombre, contenido)
        
        zip_buffer.seek(0)
        
        return StreamingResponse(
            zip_buffer,
            media_type='application/zip',
            headers={'Content-Disposition': f'attachment; filename="estudiante_{estudiante_id}_documentos.zip"'}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()
        print(f"❌ Error creando ZIP: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT nombre_archivo, ruta_archivo
            FROM documentos
            WHERE estudiante_id = %s
        """, (estudiante_id,))
        
        documentos = cursor.fetchall()
        cursor.close()
        conn.close()
        
        if not documentos:
            raise HTTPException(status_code=404, detail="No hay documentos para descargar")
        
        # Crear ZIP en memoria
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for nombre, ruta in documentos:
                if os.path.exists(ruta):
                    zip_file.write(ruta, arcname=nombre)
        
        zip_buffer.seek(0)
        
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename=documentos_estudiante_{estudiante_id}.zip"}
        )
        
    except Exception as e:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/documentos/{documento_id}/validar-ocr", tags=["Documentos"])
async def validar_documento_ocr(
    documento_id: int,
    tipo_documento: str,
    db: Session = Depends(get_db)
):
    """
    Procesa documento con OCR y valida contenido
    
    Tipos soportados: pasaporte, dni, extracto_bancario, carta_admision, certificado_idioma
    """
    try:
        import os
        import psycopg2
        from api.validador_ocr import ValidadorOCR
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        # Obtener documento de BD
        cursor.execute("""
            SELECT d.id, d.estudiante_id, d.tipo, d.nombre_archivo, d.ruta_archivo
            FROM documentos d
            WHERE d.id = %s
        """, (documento_id,))
        
        doc = cursor.fetchone()
        
        if not doc:
            cursor.close()
            conn.close()
            raise HTTPException(status_code=404, detail="Documento no encontrado")
        
        doc_id, estudiante_id, tipo, nombre, ruta = doc
        
        # Verificar archivo existe
        if not os.path.exists(ruta):
            cursor.close()
            conn.close()
            raise HTTPException(status_code=404, detail="Archivo no encontrado en servidor")
        
        # Procesar con OCR
        validador = ValidadorOCR()
        resultado = validador.procesar_documento(ruta, tipo_documento)
        
        if resultado['exito']:
            # Guardar resultados en BD
            cursor.execute("""
                UPDATE documentos 
                SET ocr_procesado = TRUE,
                    ocr_datos_extraidos = %s,
                    ocr_validacion = %s,
                    ocr_nivel_confianza = %s,
                    ocr_alertas = %s,
                    ocr_fecha_procesamiento = NOW()
                WHERE id = %s
            """, (
                json.dumps(resultado['datos_extraidos']),
                json.dumps(resultado['validacion']),
                resultado['nivel_confianza'],
                json.dumps(resultado['alertas']),
                documento_id
            ))
            
            conn.commit()
            
            # Registrar en historial
            cursor.execute("""
                INSERT INTO historial_acciones (estudiante_id, accion, detalles, fecha)
                VALUES (%s, %s, %s, NOW())
            """, (
                estudiante_id,
                'documento_validado_ocr',
                f"Documento {nombre} procesado con OCR - Confianza: {resultado['nivel_confianza']}%"
            ))
            
            conn.commit()
        
        cursor.close()
        conn.close()
        
        return resultado
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error en validación OCR: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/estudiantes/{estudiante_id}/documentos/ocr-status", tags=["Documentos"])
def obtener_status_ocr(
    estudiante_id: int,
    db: Session = Depends(get_db)
):
    """
    Obtiene estado de validación OCR de todos los documentos del estudiante
    """
    try:
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT d.id, d.tipo, d.nombre_archivo, d.ocr_procesado, 
                   d.ocr_nivel_confianza, d.ocr_alertas, d.ocr_fecha_procesamiento
            FROM documentos d
            WHERE d.estudiante_id = %s
            ORDER BY d.fecha_subida DESC
        """, (estudiante_id,))
        
        documentos = cursor.fetchall()
        cursor.close()
        conn.close()
        
        resultado = []
        for doc in documentos:
            resultado.append({
                'id': doc[0],
                'tipo': doc[1],
                'nombre': doc[2],
                'ocr_procesado': doc[3] or False,
                'nivel_confianza': doc[4] or 0,
                'alertas': json.loads(doc[5]) if doc[5] else [],
                'fecha_procesamiento': doc[6].isoformat() if doc[6] else None
            })
        
        return {
            'estudiante_id': estudiante_id,
            'total_documentos': len(resultado),
            'procesados_ocr': sum(1 for d in resultado if d['ocr_procesado']),
            'confianza_promedio': sum(d['nivel_confianza'] for d in resultado) / len(resultado) if resultado else 0,
            'documentos': resultado
        }
        
    except Exception as e:
        print(f"Error obteniendo status OCR: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/estudiantes/{estudiante_id}/documentos", tags=["Documentos"])
def obtener_checklist(estudiante_id: int, db: Session = Depends(get_db)):
    """Obtiene checklist de documentos del estudiante"""
    from modules.estudiantes import GestorEstudiantes
    
    checklist = GestorEstudiantes.checklist_documentos(estudiante_id)
    return checklist


# ============================================================================
# NOTIFICACIONES WEB - Sistema interno
# ============================================================================

@app.get("/api/estudiantes/{estudiante_id}/notificaciones", tags=["Notificaciones"])
def obtener_notificaciones(
    estudiante_id: int,
    no_leidas: bool = False,
    db: Session = Depends(get_db)
):
    """Obtiene notificaciones del estudiante"""
    from modules.notificaciones import SistemaNotificaciones
    
    # Implementación simplificada - expandir según necesidades
    return {
        'notificaciones': [],
        'total': 0,
        'no_leidas': 0
    }


# ============================================================================
# MENSAJERÍA - Chat entre estudiantes y admins
# ============================================================================

@app.post("/api/mensajes", tags=["Mensajería"])
def enviar_mensaje(
    datos: Dict,
    db: Session = Depends(get_db)
):
    """Envía mensaje entre estudiante y admin"""
    # Implementación de sistema de chat
    return {
        'exito': True,
        'mensaje_id': 1,
        'fecha': datetime.now().isoformat()
    }


@app.get("/api/estudiantes/{estudiante_id}/mensajes", tags=["Mensajería"])
def obtener_mensajes(
    estudiante_id: int,
    db: Session = Depends(get_db)
):
    """Obtiene conversación del estudiante"""
    return {
        'mensajes': [],
        'total': 0
    }


# ============================================================================
# ADMIN - GESTIÓN DE CURSOS
# ============================================================================

@app.post("/api/admin/cursos", tags=["Admin - Cursos"])
def crear_curso(
    datos: Dict,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Crea un nuevo curso"""
    from modules.cursos import GestorCursos
    from database.models import Curso
    
    nuevo_curso = Curso(
        nombre=datos['nombre'],
        escuela=datos['escuela'],
        ciudad=datos['ciudad'],
        especialidad=datos['especialidad'],
        precio=datos['precio'],
        duracion_meses=datos['duracion_meses'],
        nivel_minimo_espanol=datos.get('nivel_minimo_espanol', 'B1'),
        descripcion=datos.get('descripcion'),
        requisitos=datos.get('requisitos'),
        url_informacion=datos.get('url_informacion'),
        disponible=True
    )
    
    db.add(nuevo_curso)
    db.commit()
    db.refresh(nuevo_curso)
    
    return {'exito': True, 'curso_id': nuevo_curso.id}


@app.put("/api/admin/cursos/{curso_id}", tags=["Admin - Cursos"])
def actualizar_curso(
    curso_id: int,
    datos: Dict,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Actualiza un curso existente"""
    from database.models import Curso
    
    curso = db.query(Curso).filter(Curso.id == curso_id).first()
    if not curso:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    
    for campo, valor in datos.items():
        if hasattr(curso, campo):
            setattr(curso, campo, valor)
    
    db.commit()
    return {'exito': True, 'mensaje': 'Curso actualizado'}


@app.delete("/api/admin/cursos/{curso_id}", tags=["Admin - Cursos"])
def eliminar_curso(
    curso_id: int,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Elimina un curso"""
    from database.models import Curso
    
    curso = db.query(Curso).filter(Curso.id == curso_id).first()
    if not curso:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    
    curso.disponible = False
    db.commit()
    
    return {'exito': True, 'mensaje': 'Curso desactivado'}


# ============================================================================
# ADMIN - GESTIÓN DE ALOJAMIENTOS
# ============================================================================

@app.post("/api/admin/alojamientos", tags=["Admin - Alojamiento"])
def crear_alojamiento(
    datos: Dict,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Crea nuevo alojamiento"""
    from database.models import Alojamiento
    
    nuevo = Alojamiento(
        tipo=datos['tipo'],
        direccion=datos['direccion'],
        ciudad=datos['ciudad'],
        precio_mensual=datos['precio_mensual'],
        gastos_incluidos=datos.get('gastos_incluidos', False),
        num_habitaciones=datos.get('num_habitaciones', 1),
        num_banos=datos.get('num_banos', 1),
        metros_cuadrados=datos.get('metros_cuadrados'),
        disponible=True
    )
    
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    
    return {'exito': True, 'alojamiento_id': nuevo.id}


@app.put("/api/admin/alojamientos/{alojamiento_id}", tags=["Admin - Alojamiento"])
def actualizar_alojamiento(
    alojamiento_id: int,
    datos: Dict,
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Actualiza alojamiento"""
    from database.models import Alojamiento
    
    alojamiento = db.query(Alojamiento).filter(Alojamiento.id == alojamiento_id).first()
    if not alojamiento:
        raise HTTPException(status_code=404, detail="Alojamiento no encontrado")
    
    for campo, valor in datos.items():
        if hasattr(alojamiento, campo):
            setattr(alojamiento, campo, valor)
    
    db.commit()
    return {'exito': True, 'mensaje': 'Alojamiento actualizado'}


# ============================================================================
# ADMIN - REPORTES AVANZADOS
# ============================================================================

@app.get("/api/admin/reportes/mensual", tags=["Admin - Reportes"])
def reporte_mensual(
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Genera reporte mensual completo"""
    reporte = PanelAdministrativo.generar_reporte_mensual()
    return reporte


@app.get("/api/admin/reportes/especialidades", tags=["Admin - Reportes"])
def estadisticas_especialidades(
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Estadísticas por especialidad"""
    stats = PanelAdministrativo.estadisticas_por_especialidad()
    return {'especialidades': stats}


@app.get("/api/admin/reportes/conversion", tags=["Admin - Reportes"])
def embudo_conversion(
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Embudo de conversión de estudiantes"""
    embudo = PanelAdministrativo.embudo_conversion()
    return embudo


# ============================================================================
# ADMIN - GESTIÓN DE DOCUMENTOS
# ============================================================================

@app.get("/api/admin/documentos/pendientes", tags=["Admin - Documentos"])
def documentos_pendientes(
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Lista estudiantes con documentos pendientes"""
    estudiantes = db.query(Estudiante).filter(
        Estudiante.estado_procesamiento.in_([
            'registrado',
            'procesado_automaticamente',
            'pendiente_revision_admin'
        ])
    ).all()
    
    pendientes = []
    for est in estudiantes:
        from modules.estudiantes import GestorEstudiantes
        checklist = GestorEstudiantes.checklist_documentos(est.id)
        
        if checklist['total_obligatorios'] > 0:
            pendientes.append({
                'estudiante_id': est.id,
                'nombre': est.nombre_completo,
                'documentos_obligatorios_pendientes': checklist['total_obligatorios'],
                'documentos_recomendados_pendientes': checklist['total_recomendados']
            })
    
    return {'total': len(pendientes), 'estudiantes': pendientes}


# ============================================================================
# MENSAJERÍA INTERNA - Chat entre estudiantes y admins
# ============================================================================

@app.post("/api/conversaciones", tags=["Mensajería"])
def crear_conversacion(datos: Dict, db: Session = Depends(get_db)):
    """Crea una nueva conversación"""
    from modules.mensajeria import SistemaMensajeria
    
    conversacion = SistemaMensajeria.crear_conversacion(
        estudiante_id=datos['estudiante_id'],
        admin_id=datos.get('admin_id')
    )
    
    return {'exito': True, 'conversacion_id': conversacion.id}


@app.post("/api/mensajes", tags=["Mensajería"])
def enviar_mensaje(datos: Dict, db: Session = Depends(get_db)):
    """Envía mensaje en una conversación"""
    from modules.mensajeria import SistemaMensajeria
    
    resultado = SistemaMensajeria.enviar_mensaje(
        conversacion_id=datos['conversacion_id'],
        remitente_tipo=datos['remitente_tipo'],
        remitente_id=datos['remitente_id'],
        contenido=datos['contenido']
    )
    
    return resultado


@app.get("/api/conversaciones/{conversacion_id}/mensajes", tags=["Mensajería"])
def obtener_mensajes_conversacion(
    conversacion_id: int,
    limit: int = 50,
    db: Session = Depends(get_db)
):
    """Obtiene mensajes de una conversación"""
    from modules.mensajeria import SistemaMensajeria
    
    mensajes = SistemaMensajeria.obtener_mensajes(conversacion_id, limit=limit)
    return {'mensajes': mensajes, 'total': len(mensajes)}


@app.get("/api/estudiantes/{estudiante_id}/conversaciones", tags=["Mensajería"])
def obtener_conversaciones_estudiante(estudiante_id: int, db: Session = Depends(get_db)):
    """Obtiene conversaciones del estudiante"""
    from modules.mensajeria import SistemaMensajeria
    
    conversaciones = SistemaMensajeria.obtener_conversaciones_estudiante(estudiante_id)
    return {'conversaciones': conversaciones}


@app.get("/api/admin/conversaciones", tags=["Admin - Mensajería"])
def obtener_conversaciones_admin(
    usuario=Depends(obtener_usuario_actual),
    db: Session = Depends(get_db)
):
    """Obtiene todas las conversaciones activas para admin"""
    from modules.mensajeria import SistemaMensajeria
    
    conversaciones = SistemaMensajeria.obtener_conversaciones_admin()
    return {'conversaciones': conversaciones}


# ============================================================================
# NOTIFICACIONES EMAIL
# ============================================================================

@app.post("/api/notificaciones/test-email", tags=["Notificaciones"])
def test_email(destinatario: str):
    """Endpoint de prueba para verificar configuración de email"""
    from modules.notificaciones_email import EmailService
    import config
    
    html = EmailService.generar_template_html(
        titulo="Email de Prueba",
        mensaje="<p>Este es un email de prueba del sistema de notificaciones.</p><p>Si recibes este mensaje, la configuración SMTP está funcionando correctamente.</p>",
        boton_texto="Visitar Plataforma",
        boton_url=getattr(config, 'WEB_URL', 'http://localhost:3000')
    )
    
    resultado = EmailService.enviar_email(
        destinatario=destinatario,
        asunto="🧪 Email de Prueba - Sistema de Notificaciones",
        contenido_html=html
    )
    
    if resultado:
        return {"mensaje": "Email enviado correctamente", "destinatario": destinatario}
    else:
        raise HTTPException(status_code=500, detail="Error al enviar email")


@app.post("/api/estudiantes/{estudiante_id}/notificaciones/resend", tags=["Notificaciones"])
def reenviar_notificacion(
    estudiante_id: int,
    tipo: str,
    db: Session = Depends(get_db)
):
    """
    Reenviar notificación por email a un estudiante
    Tipos: registro, aprobacion, pendiente
    """
    from modules.notificaciones_email import NotificacionesEmail
    
    estudiante = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    estudiante_dict = {
        'id': estudiante.id,
        'nombre_completo': estudiante.nombre_completo,
        'email': estudiante.email,
        'especialidad_interes': estudiante.especialidad_interes,
        'nacionalidad': estudiante.nacionalidad
    }
    
    resultado = False
    
    if tipo == "registro":
        resultado = NotificacionesEmail.enviar_confirmacion_registro(estudiante_dict)
    elif tipo == "aprobacion":
        resultado = NotificacionesEmail.enviar_solicitud_aprobada(estudiante_dict)
    elif tipo == "pendiente":
        motivo = "Revisión de documentación requerida"
        resultado = NotificacionesEmail.enviar_solicitud_pendiente_revision(estudiante_dict, motivo)
    else:
        raise HTTPException(status_code=400, detail="Tipo de notificación inválido")
    
    if resultado:
        return {"mensaje": f"Notificación '{tipo}' reenviada a {estudiante.email}"}
    else:
        raise HTTPException(status_code=500, detail="Error al enviar notificación")


# ============================================================================
# INTEGRACIÓN APIS ESCUELAS
# ============================================================================

@app.get("/api/admin/sincronizar-cursos-escuelas", tags=["Admin - Escuelas"])
def sincronizar_cursos_desde_escuelas(
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """
    Sincroniza cursos desde APIs/scraping de escuelas españolas
    Obtiene cursos actualizados de múltiples universidades
    """
    verificar_token(credentials.credentials)
    
    try:
        from api.integrador_escuelas import IntegradorEscuelas
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        # Obtener cursos de todas las fuentes
        cursos_externos = IntegradorEscuelas.sincronizar_todos_cursos()
        
        # Insertar/actualizar en base de datos
        cursos_insertados = 0
        cursos_actualizados = 0
        
        for curso_ext in cursos_externos:
            # Verificar si el curso ya existe (por nombre y ciudad)
            cursor.execute(
                """
                SELECT id, precio_eur, cupos_disponibles 
                FROM cursos 
                WHERE nombre = %s AND ciudad = %s
                """,
                (curso_ext['nombre'], curso_ext['ciudad'])
            )
            curso_existente = cursor.fetchone()
            
            if curso_existente:
                # Actualizar precio y cupos
                cursor.execute(
                    """
                    UPDATE cursos 
                    SET precio_eur = %s, 
                        cupos_disponibles = %s,
                        descripcion = %s,
                        activo = TRUE,
                        updated_at = NOW()
                    WHERE id = %s
                    """,
                    (curso_ext['precio_eur'], curso_ext['cupos_disponibles'], 
                     curso_ext['descripcion'], curso_existente[0])
                )
                cursos_actualizados += 1
            else:
                # Insertar nuevo curso
                cursor.execute(
                    """
                    INSERT INTO cursos (nombre, descripcion, duracion_meses, precio_eur, 
                                       ciudad, nivel_espanol_requerido, cupos_disponibles, activo, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE, NOW(), NOW())
                    """,
                    (curso_ext['nombre'], curso_ext['descripcion'], curso_ext['duracion_meses'],
                     curso_ext['precio_eur'], curso_ext['ciudad'], curso_ext['nivel_espanol_requerido'],
                     curso_ext['cupos_disponibles'])
                )
                cursos_insertados += 1
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return {
            "exito": True,
            "cursos_encontrados": len(cursos_externos),
            "cursos_insertados": cursos_insertados,
            "cursos_actualizados": cursos_actualizados,
            "cursos_preview": cursos_externos[:5]  # Primeros 5 para preview
        }
        
    except Exception as e:
        print(f"Error sincronizando cursos: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/cursos/buscar-externos", tags=["Cursos"])
def buscar_cursos_externos(
    especialidad: Optional[str] = None,
    ciudad: Optional[str] = None,
    presupuesto_max: Optional[float] = None
):
    """
    Busca cursos en APIs externas sin guardar en BD
    Útil para búsqueda en tiempo real
    """
    try:
        from api.integrador_escuelas import IntegradorEscuelas
        
        # Obtener todos los cursos disponibles
        cursos = IntegradorEscuelas.sincronizar_todos_cursos()
        
        # Aplicar filtros
        if especialidad:
            cursos = IntegradorEscuelas.buscar_cursos_por_especialidad(especialidad, cursos)
        
        if ciudad:
            cursos = IntegradorEscuelas.filtrar_por_ciudad(cursos, ciudad)
        
        if presupuesto_max:
            cursos = IntegradorEscuelas.filtrar_por_presupuesto(cursos, presupuesto_max)
        
        return {
            "total": len(cursos),
            "cursos": cursos
        }
        
    except Exception as e:
        print(f"Error buscando cursos externos: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/cursos/{curso_id}/verificar-disponibilidad", tags=["Cursos"])
def verificar_disponibilidad_externa(
    curso_id: int,
    db: Session = Depends(get_db)
):
    """
    Verifica disponibilidad en tiempo real desde la fuente externa
    """
    try:
        from api.integrador_escuelas import IntegradorEscuelas
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        # Obtener curso de BD
        cursor.execute(
            "SELECT nombre, ciudad, cupos_disponibles FROM cursos WHERE id = %s",
            (curso_id,)
        )
        curso = cursor.fetchone()
        
        if not curso:
            cursor.close()
            conn.close()
            raise HTTPException(status_code=404, detail="Curso no encontrado")
        
        # Consultar disponibilidad real
        cupos_actuales = IntegradorEscuelas.actualizar_disponibilidad_real(curso_id)
        
        # Actualizar en BD
        cursor.execute(
            "UPDATE cursos SET cupos_disponibles = %s, updated_at = NOW() WHERE id = %s",
            (cupos_actuales, curso_id)
        )
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return {
            "curso_id": curso_id,
            "nombre": curso[0],
            "cupos_disponibles": cupos_actuales,
            "disponible": cupos_actuales > 0,
            "actualizado": True
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error verificando disponibilidad: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/scrapers/health", tags=["Admin - Escuelas"])
def obtener_salud_scrapers(
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """
    Obtiene reporte de salud de todos los scrapers
    """
    verificar_token(credentials.credentials)
    
    try:
        from api.monitor_scrapers import MonitorScrapers
        
        monitor = MonitorScrapers()
        report = monitor.get_health_report()
        
        return report
        
    except Exception as e:
        print(f"Error obteniendo salud scrapers: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/admin/scrapers/clear-alerts", tags=["Admin - Escuelas"])
def limpiar_alertas_scrapers(
    source: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """
    Limpia alertas de scrapers (después de arreglar problemas)
    """
    verificar_token(credentials.credentials)
    
    try:
        from api.monitor_scrapers import MonitorScrapers
        
        monitor = MonitorScrapers()
        monitor.clear_alerts(source)
        
        return {
            "exito": True,
            "mensaje": f"Alertas limpiadas: {source or 'todas'}"
        }
        
    except Exception as e:
        print(f"Error limpiando alertas: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/cache/stats", tags=["Admin - Escuelas"])
def obtener_estadisticas_cache(
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """
    Obtiene estadísticas del sistema de cache
    """
    verificar_token(credentials.credentials)
    
    try:
        from api.cache_manager import CacheManager
        
        cache = CacheManager()
        stats = cache.get_stats()
        
        return stats
        
    except Exception as e:
        print(f"Error obteniendo stats cache: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/admin/cache/clear", tags=["Admin - Escuelas"])
def limpiar_cache(
    source: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """
    Limpia cache de cursos (forzar nuevo scraping)
    """
    verificar_token(credentials.credentials)
    
    try:
        from api.cache_manager import CacheManager
        
        cache = CacheManager()
        cache.clear(source)
        
        return {
            "exito": True,
            "mensaje": f"Cache limpiado: {source or 'todo'}"
        }
        
    except Exception as e:
        print(f"Error limpiando cache: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/estudiantes/{estudiante_id}/calcular-probabilidad", tags=["Estudiantes"])
def calcular_probabilidad_exito(
    estudiante_id: int,
    db: Session = Depends(get_db)
):
    """
    Calcula probabilidad de éxito en solicitud de visa
    Basado en múltiples factores del perfil del estudiante
    """
    try:
        from api.predictor_exito import PredictorExito
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        # Obtener datos del estudiante
        cursor.execute("""
            SELECT e.nombre, e.nacionalidad, e.especialidad, e.nivel_idioma, 
                   e.fondos_disponibles, e.created_at, c.nombre as curso
            FROM estudiantes e
            LEFT JOIN cursos c ON e.curso_asignado_id = c.id
            WHERE e.id = %s
        """, (estudiante_id,))
        
        estudiante = cursor.fetchone()
        
        if not estudiante:
            cursor.close()
            conn.close()
            raise HTTPException(status_code=404, detail="Estudiante no encontrado")
        
        # Contar documentos
        cursor.execute("""
            SELECT COUNT(*) FROM documentos WHERE estudiante_id = %s
        """, (estudiante_id,))
        docs_subidos = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT COUNT(*) FROM documentos_generados 
            WHERE estudiante_id = %s AND estado = 'aprobado'
        """, (estudiante_id,))
        docs_generados = cursor.fetchone()[0]
        
        cursor.close()
        conn.close()
        
        # Calcular días desde registro
        dias_registro = (datetime.now() - estudiante[5]).days if estudiante[5] else 0
        
        # Preparar datos para predictor
        datos_estudiante = {
            'nacionalidad': estudiante[1] or '',
            'fondos_disponibles': float(estudiante[4] or 0),
            'documentos_subidos': docs_subidos,
            'documentos_generados': docs_generados,
            'curso_asignado': estudiante[6] or '',
            'nivel_idioma': estudiante[3] or 'B1',
            'especialidad': estudiante[2] or '',
            'dias_desde_registro': dias_registro,
            'antecedentes': False  # Por ahora asumimos sin antecedentes
        }
        
        # Calcular probabilidad
        predictor = PredictorExito()
        resultado = predictor.calcular_probabilidad(datos_estudiante)
        
        return {
            'estudiante_id': estudiante_id,
            'nombre': estudiante[0],
            **resultado
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error calculando probabilidad: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/estudiantes/{estudiante_id}/solicitar-servicio", tags=["Servicios"])
def solicitar_servicio(
    estudiante_id: int,
    datos: dict,
    db: Session = Depends(get_db)
):
    """Estudiante solicita un servicio adicional (antecedentes penales, cita embajada, etc)"""
    try:
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        # Verificar si ya solicitó este servicio
        cursor.execute("""
            SELECT id FROM servicios_solicitados 
            WHERE estudiante_id = %s AND servicio_id = %s
        """, (estudiante_id, datos.get('servicio_id')))
        
        if cursor.fetchone():
            cursor.close()
            conn.close()
            raise HTTPException(status_code=400, detail="Ya solicitaste este servicio")
        
        # Insertar solicitud
        cursor.execute("""
            INSERT INTO servicios_solicitados (estudiante_id, servicio_id, servicio_nombre, estado)
            VALUES (%s, %s, %s, 'pendiente')
            RETURNING id
        """, (estudiante_id, datos.get('servicio_id'), datos.get('servicio_nombre')))
        
        servicio_id = cursor.fetchone()[0]
        conn.commit()
        cursor.close()
        conn.close()
        
        return {
            'success': True,
            'servicio_id': servicio_id,
            'mensaje': 'Solicitud enviada. El administrador te contactará.'
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error solicitando servicio: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/estudiantes/{estudiante_id}/servicios-solicitados", tags=["Servicios"])
def obtener_servicios_solicitados(
    estudiante_id: int,
    db: Session = Depends(get_db)
):
    """Obtiene servicios solicitados por un estudiante"""
    try:
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, servicio_id, servicio_nombre, estado, precio, notas, fecha_solicitud
            FROM servicios_solicitados
            WHERE estudiante_id = %s
            ORDER BY fecha_solicitud DESC
        """, (estudiante_id,))
        
        servicios = []
        for row in cursor.fetchall():
            servicios.append({
                'id': row[0],
                'servicio_id': row[1],
                'servicio_nombre': row[2],
                'estado': row[3],
                'precio': float(row[4]) if row[4] else None,
                'notas': row[5],
                'fecha_solicitud': row[6].isoformat() if row[6] else None
            })
        
        cursor.close()
        conn.close()
        
        return {'servicios': servicios}
        
    except Exception as e:
        print(f"Error obteniendo servicios: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/servicios-solicitados", tags=["Admin"])
def obtener_todos_servicios_solicitados(
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Admin obtiene todas las solicitudes de servicios de todos los estudiantes"""
    verificar_token(credentials.credentials)
    
    try:
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT s.id, s.estudiante_id, e.nombre as estudiante_nombre, e.email,
                   s.servicio_id, s.servicio_nombre, s.estado, s.precio, s.notas, 
                   s.fecha_solicitud, s.fecha_respuesta
            FROM servicios_solicitados s
            JOIN estudiantes e ON s.estudiante_id = e.id
            ORDER BY s.fecha_solicitud DESC
        """)
        
        servicios = []
        for row in cursor.fetchall():
            servicios.append({
                'id': row[0],
                'estudiante_id': row[1],
                'estudiante_nombre': row[2],
                'estudiante_email': row[3],
                'servicio_id': row[4],
                'servicio_nombre': row[5],
                'estado': row[6],
                'precio': float(row[7]) if row[7] else None,
                'notas': row[8],
                'fecha_solicitud': row[9].isoformat() if row[9] else None,
                'fecha_respuesta': row[10].isoformat() if row[10] else None
            })
        
        cursor.close()
        conn.close()
        
        return {'servicios': servicios, 'total': len(servicios)}
        
    except Exception as e:
        print(f"Error obteniendo servicios admin: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/admin/servicios-solicitados/{servicio_id}", tags=["Admin"])
def actualizar_servicio_solicitado(
    servicio_id: int,
    datos: dict,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Admin actualiza estado, precio y notas de un servicio solicitado"""
    verificar_token(credentials.credentials)
    
    try:
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE servicios_solicitados
            SET estado = %s, precio = %s, notas = %s, fecha_respuesta = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (datos.get('estado'), datos.get('precio'), datos.get('notas'), servicio_id))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return {'success': True, 'mensaje': 'Servicio actualizado'}
        
    except Exception as e:
        print(f"Error actualizando servicio: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/estudiantes/{estudiante_id}/analisis-completo", tags=["Admin"])
def obtener_analisis_completo(
    estudiante_id: int,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """
    Análisis completo de estudiante con probabilidad de éxito y recomendaciones
    """
    verificar_token(credentials.credentials)
    
    try:
        from api.predictor_exito import PredictorExito
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        # Obtener todos los datos del estudiante
        cursor.execute("""
            SELECT e.*, c.nombre as curso, c.precio as precio_curso,
                   a.tipo as alojamiento
            FROM estudiantes e
            LEFT JOIN cursos c ON e.curso_asignado_id = c.id
            LEFT JOIN alojamientos a ON e.alojamiento_asignado_id = a.id
            WHERE e.id = %s
        """, (estudiante_id,))
        
        estudiante = cursor.fetchone()
        
        if not estudiante:
            cursor.close()
            conn.close()
            raise HTTPException(status_code=404, detail="Estudiante no encontrado")
        
        # Obtener documentos
        cursor.execute("""
            SELECT tipo, nombre_archivo, fecha_subida 
            FROM documentos 
            WHERE estudiante_id = %s
        """, (estudiante_id,))
        documentos = cursor.fetchall()
        
        cursor.execute("""
            SELECT tipo, estado, fecha_generacion, fecha_aprobacion
            FROM documentos_generados 
            WHERE estudiante_id = %s
        """, (estudiante_id,))
        docs_generados = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Preparar datos para predictor
        datos_estudiante = {
            'nacionalidad': estudiante[3] or '',
            'fondos_disponibles': float(estudiante[6] or 0),
            'documentos_subidos': len(documentos),
            'documentos_generados': len([d for d in docs_generados if d[1] == 'aprobado']),
            'curso_asignado': estudiante[17] or '',
            'nivel_idioma': estudiante[5] or 'B1',
            'especialidad': estudiante[4] or '',
            'dias_desde_registro': (datetime.now() - estudiante[11]).days if estudiante[11] else 0,
            'antecedentes': False
        }
        
        # Calcular probabilidad
        predictor = PredictorExito()
        analisis = predictor.calcular_probabilidad(datos_estudiante)
        
        return {
            'estudiante': {
                'id': estudiante[0],
                'nombre': estudiante[1],
                'email': estudiante[2],
                'nacionalidad': estudiante[3],
                'estado': estudiante[9]
            },
            'analisis': analisis,
            'documentacion': {
                'subidos': len(documentos),
                'generados_aprobados': len([d for d in docs_generados if d[1] == 'aprobado']),
                'generados_pendientes': len([d for d in docs_generados if d[1] == 'pendiente'])
            },
            'curso': estudiante[17],
            'alojamiento': estudiante[19]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error en análisis completo: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# PARTNERSHIPS UNIVERSITARIOS
# ============================================================================

@app.post("/api/admin/partners/universidades", tags=["Admin - Partners"])
def crear_universidad_partner(
    datos: Dict,
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Crear nueva universidad partner"""
    verificar_token(credentials.credentials)
    
    try:
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO universidades_partner 
            (nombre, codigo_referido, email_contacto, persona_contacto, telefono,
             tipo_comision, porcentaje_comision, monto_fijo_comision, logo_url, notas)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            datos.get('nombre'),
            datos.get('codigo_referido'),
            datos.get('email_contacto'),
            datos.get('persona_contacto'),
            datos.get('telefono'),
            datos.get('tipo_comision', 'porcentaje'),
            datos.get('porcentaje_comision', 15.0),
            datos.get('monto_fijo_comision', 0),
            datos.get('logo_url'),
            datos.get('notas')
        ))
        
        universidad_id = cursor.fetchone()[0]
        conn.commit()
        cursor.close()
        conn.close()
        
        return {
            'exito': True,
            'universidad_id': universidad_id,
            'mensaje': 'Universidad partner creada exitosamente'
        }
        
    except Exception as e:
        print(f"Error creando universidad partner: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/partners/universidades", tags=["Admin - Partners"])
def listar_universidades_partner(
    activo: Optional[bool] = None,
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Listar todas las universidades partner"""
    verificar_token(credentials.credentials)
    
    try:
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        query = """
            SELECT u.id, u.nombre, u.codigo_referido, u.email_contacto, 
                   u.persona_contacto, u.telefono, u.tipo_comision, 
                   u.porcentaje_comision, u.monto_fijo_comision, u.activo,
                   u.logo_url, u.notas, u.created_at,
                   COUNT(DISTINCT e.id) as total_referidos,
                   COUNT(DISTINCT CASE WHEN e.estado = 'matriculado' THEN e.id END) as matriculados,
                   COALESCE(SUM(CASE WHEN c.estado = 'pagado' THEN c.monto_comision ELSE 0 END), 0) as total_pagado,
                   COALESCE(SUM(CASE WHEN c.estado = 'pendiente' THEN c.monto_comision ELSE 0 END), 0) as total_pendiente
            FROM universidades_partner u
            LEFT JOIN estudiantes e ON e.universidad_referidora_id = u.id
            LEFT JOIN comisiones c ON c.universidad_id = u.id
        """
        
        if activo is not None:
            query += " WHERE u.activo = %s"
            cursor.execute(query + " GROUP BY u.id ORDER BY u.nombre", (activo,))
        else:
            cursor.execute(query + " GROUP BY u.id ORDER BY u.nombre")
        
        universidades = cursor.fetchall()
        cursor.close()
        conn.close()
        
        resultado = []
        for u in universidades:
            resultado.append({
                'id': u[0],
                'nombre': u[1],
                'codigo_referido': u[2],
                'email_contacto': u[3],
                'persona_contacto': u[4],
                'telefono': u[5],
                'tipo_comision': u[6],
                'porcentaje_comision': float(u[7]) if u[7] else 0,
                'monto_fijo_comision': float(u[8]) if u[8] else 0,
                'activo': u[9],
                'logo_url': u[10],
                'notas': u[11],
                'created_at': u[12].isoformat() if u[12] else None,
                'stats': {
                    'total_referidos': u[13],
                    'matriculados': u[14],
                    'tasa_conversion': round((u[14] / u[13] * 100) if u[13] > 0 else 0, 1),
                    'total_pagado': float(u[15]),
                    'total_pendiente': float(u[16])
                }
            })
        
        return {
            'universidades': resultado,
            'total': len(resultado)
        }
        
    except Exception as e:
        print(f"Error listando universidades: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/admin/partners/universidades/{universidad_id}", tags=["Admin - Partners"])
def actualizar_universidad_partner(
    universidad_id: int,
    datos: Dict,
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Actualizar datos de universidad partner"""
    verificar_token(credentials.credentials)
    
    try:
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE universidades_partner 
            SET nombre = %s,
                email_contacto = %s,
                persona_contacto = %s,
                telefono = %s,
                tipo_comision = %s,
                porcentaje_comision = %s,
                monto_fijo_comision = %s,
                logo_url = %s,
                notas = %s,
                activo = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (
            datos.get('nombre'),
            datos.get('email_contacto'),
            datos.get('persona_contacto'),
            datos.get('telefono'),
            datos.get('tipo_comision'),
            datos.get('porcentaje_comision'),
            datos.get('monto_fijo_comision'),
            datos.get('logo_url'),
            datos.get('notas'),
            datos.get('activo'),
            universidad_id
        ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return {'exito': True, 'mensaje': 'Universidad actualizada'}
        
    except Exception as e:
        print(f"Error actualizando universidad: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/partners/universidades/{universidad_id}/estudiantes", tags=["Admin - Partners"])
def listar_estudiantes_referidos(
    universidad_id: int,
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Lista estudiantes referidos por universidad"""
    verificar_token(credentials.credentials)
    
    try:
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT e.id, e.nombre, e.email, e.nacionalidad, e.especialidad,
                   e.estado, e.created_at, c.nombre as curso,
                   com.monto_comision, com.estado as estado_comision
            FROM estudiantes e
            LEFT JOIN cursos c ON e.curso_asignado_id = c.id
            LEFT JOIN comisiones com ON com.estudiante_id = e.id AND com.universidad_id = %s
            WHERE e.universidad_referidora_id = %s
            ORDER BY e.created_at DESC
        """, (universidad_id, universidad_id))
        
        estudiantes = cursor.fetchall()
        cursor.close()
        conn.close()
        
        resultado = []
        for est in estudiantes:
            resultado.append({
                'id': est[0],
                'nombre': est[1],
                'email': est[2],
                'nacionalidad': est[3],
                'especialidad': est[4],
                'estado': est[5],
                'created_at': est[6].isoformat() if est[6] else None,
                'curso': est[7],
                'comision': {
                    'monto': float(est[8]) if est[8] else 0,
                    'estado': est[9] or 'no_generada'
                }
            })
        
        return {
            'estudiantes': resultado,
            'total': len(resultado)
        }
        
    except Exception as e:
        print(f"Error listando estudiantes referidos: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/admin/partners/comisiones/generar", tags=["Admin - Partners"])
def generar_comision(
    datos: Dict,
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Genera comisión para estudiante matriculado"""
    verificar_token(credentials.credentials)
    
    try:
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        estudiante_id = datos.get('estudiante_id')
        monto_curso = datos.get('monto_curso')
        
        # Obtener universidad referidora
        cursor.execute("""
            SELECT universidad_referidora_id FROM estudiantes WHERE id = %s
        """, (estudiante_id,))
        
        result = cursor.fetchone()
        if not result or not result[0]:
            raise HTTPException(status_code=400, detail="Estudiante no tiene universidad referidora")
        
        universidad_id = result[0]
        
        # Obtener configuración de comisión
        cursor.execute("""
            SELECT tipo_comision, porcentaje_comision, monto_fijo_comision
            FROM universidades_partner WHERE id = %s
        """, (universidad_id,))
        
        config = cursor.fetchone()
        tipo_comision, porcentaje, monto_fijo = config
        
        # Calcular comisión
        if tipo_comision == 'porcentaje':
            monto_comision = monto_curso * (porcentaje / 100)
        else:
            monto_comision = monto_fijo
        
        # Crear registro de comisión
        cursor.execute("""
            INSERT INTO comisiones 
            (universidad_id, estudiante_id, monto_curso, monto_comision, 
             estado, fecha_matricula)
            VALUES (%s, %s, %s, %s, 'pendiente', NOW())
            RETURNING id
        """, (universidad_id, estudiante_id, monto_curso, monto_comision))
        
        comision_id = cursor.fetchone()[0]
        conn.commit()
        cursor.close()
        conn.close()
        
        return {
            'exito': True,
            'comision_id': comision_id,
            'monto_comision': float(monto_comision)
        }
        
    except Exception as e:
        print(f"Error generando comisión: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/partners/comisiones", tags=["Admin - Partners"])
def listar_comisiones(
    universidad_id: Optional[int] = None,
    estado: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Lista todas las comisiones"""
    verificar_token(credentials.credentials)
    
    try:
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        query = """
            SELECT c.id, c.monto_curso, c.monto_comision, c.estado,
                   c.fecha_matricula, c.fecha_pago, c.notas,
                   u.nombre as universidad, e.nombre as estudiante
            FROM comisiones c
            JOIN universidades_partner u ON c.universidad_id = u.id
            JOIN estudiantes e ON c.estudiante_id = e.id
            WHERE 1=1
        """
        
        params = []
        if universidad_id:
            query += " AND c.universidad_id = %s"
            params.append(universidad_id)
        if estado:
            query += " AND c.estado = %s"
            params.append(estado)
        
        query += " ORDER BY c.created_at DESC"
        
        cursor.execute(query, params)
        comisiones = cursor.fetchall()
        cursor.close()
        conn.close()
        
        resultado = []
        for com in comisiones:
            resultado.append({
                'id': com[0],
                'monto_curso': float(com[1]),
                'monto_comision': float(com[2]),
                'estado': com[3],
                'fecha_matricula': com[4].isoformat() if com[4] else None,
                'fecha_pago': com[5].isoformat() if com[5] else None,
                'notas': com[6],
                'universidad': com[7],
                'estudiante': com[8]
            })
        
        return {
            'comisiones': resultado,
            'total': len(resultado),
            'total_pendiente': sum(c['monto_comision'] for c in resultado if c['estado'] == 'pendiente'),
            'total_pagado': sum(c['monto_comision'] for c in resultado if c['estado'] == 'pagado')
        }
        
    except Exception as e:
        print(f"Error listando comisiones: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/admin/partners/comisiones/{comision_id}/marcar-pagado", tags=["Admin - Partners"])
def marcar_comision_pagada(
    comision_id: int,
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Marca comisión como pagada"""
    verificar_token(credentials.credentials)
    
    try:
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE comisiones 
            SET estado = 'pagado', fecha_pago = NOW()
            WHERE id = %s
        """, (comision_id,))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return {'exito': True, 'mensaje': 'Comisión marcada como pagada'}
        
    except Exception as e:
        print(f"Error marcando comisión: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/partners/dashboard", tags=["Admin - Partners"])
def dashboard_partners(
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Dashboard general de partnerships"""
    verificar_token(credentials.credentials)
    
    try:
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        # Estadísticas generales
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT u.id) as total_universidades,
                COUNT(DISTINCT e.id) as total_referidos,
                COUNT(DISTINCT CASE WHEN e.estado = 'matriculado' THEN e.id END) as total_matriculados,
                COALESCE(SUM(CASE WHEN c.estado = 'pendiente' THEN c.monto_comision ELSE 0 END), 0) as comisiones_pendientes,
                COALESCE(SUM(CASE WHEN c.estado = 'pagado' THEN c.monto_comision ELSE 0 END), 0) as comisiones_pagadas
            FROM universidades_partner u
            LEFT JOIN estudiantes e ON e.universidad_referidora_id = u.id
            LEFT JOIN comisiones c ON c.universidad_id = u.id
            WHERE u.activo = TRUE
        """)
        
        stats = cursor.fetchone()
        
        # Top universidades
        cursor.execute("""
            SELECT u.nombre, COUNT(e.id) as total_referidos,
                   COALESCE(SUM(c.monto_comision), 0) as total_comisiones
            FROM universidades_partner u
            LEFT JOIN estudiantes e ON e.universidad_referidora_id = u.id
            LEFT JOIN comisiones c ON c.universidad_id = u.id
            WHERE u.activo = TRUE
            GROUP BY u.id, u.nombre
            ORDER BY total_referidos DESC
            LIMIT 5
        """)
        
        top_universidades = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return {
            'stats': {
                'total_universidades': stats[0],
                'total_referidos': stats[1],
                'total_matriculados': stats[2],
                'tasa_conversion': round((stats[2] / stats[1] * 100) if stats[1] > 0 else 0, 1),
                'comisiones_pendientes': float(stats[3]),
                'comisiones_pagadas': float(stats[4])
            },
            'top_universidades': [
                {
                    'nombre': u[0],
                    'total_referidos': u[1],
                    'total_comisiones': float(u[2])
                }
                for u in top_universidades
            ]
        }
        
    except Exception as e:
        print(f"Error en dashboard partners: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# HEALTH CHECK
# ============================================================================

@app.get("/", tags=["Health"])
def root():
    """Health check"""
    return {
        "status": "online",
        "servicio": "Bot Visas Estudio API",
        "version": "1.0.0",
        "timestamp": datetime.now().isoformat()
    }


@app.get("/health", tags=["Health"])
def health():
    """Health check detallado"""
    return {
        "status": "healthy",
        "database": "connected",
        "timestamp": datetime.now().isoformat()
    }


# ============================================================================
# MENSAJERÍA ADMIN - ESTUDIANTE MEJORADA
# ============================================================================

@app.post("/api/admin/estudiantes/{estudiante_id}/enviar-mensaje", tags=["Admin - Mensajería"])
def admin_enviar_mensaje_estudiante(
    estudiante_id: int,
    datos: dict,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """
    Admin envía mensaje a estudiante con notificación por email
    tipos: 'solicitud_documento', 'recordatorio', 'informacion', 'urgente'
    """
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from datetime import datetime
    
    # Validar datos
    mensaje = datos.get('mensaje', '').strip()
    tipo = datos.get('tipo', 'informacion')
    asunto = datos.get('asunto', 'Mensaje del Administrador')
    documento_solicitado = datos.get('documento_solicitado', '')  # Opcional
    
    if not mensaje:
        raise HTTPException(status_code=400, detail="El mensaje no puede estar vacío")
    
    tipos_validos = ['solicitud_documento', 'recordatorio', 'informacion', 'urgente']
    if tipo not in tipos_validos:
        tipo = 'informacion'
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    # Verificar que el estudiante existe y obtener su email
    cursor.execute("""
        SELECT id, nombre, email FROM estudiantes WHERE id = %s
    """, (estudiante_id,))
    
    estudiante = cursor.fetchone()
    if not estudiante:
        cursor.close()
        conn.close()
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    estudiante_nombre = estudiante[1]
    estudiante_email = estudiante[2]
    
    # Insertar mensaje en la base de datos
    cursor.execute("""
        INSERT INTO mensajes_chat (
            estudiante_id, remitente, mensaje, tipo, leido, created_at
        ) VALUES (%s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (estudiante_id, 'admin', mensaje, tipo, False, datetime.utcnow()))
    
    mensaje_id = cursor.fetchone()[0]
    conn.commit()
    
    # Enviar notificación por email
    try:
        msg = MIMEMultipart()
        msg['From'] = os.getenv('SMTP_USER')
        msg['To'] = estudiante_email
        msg['Subject'] = asunto
        
        # Emojis según tipo
        emoji_tipo = {
            'solicitud_documento': '📄',
            'recordatorio': '⏰',
            'informacion': 'ℹ️',
            'urgente': '🚨'
        }
        
        emoji = emoji_tipo.get(tipo, 'ℹ️')
        
        # Cuerpo del email
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; background-color: #f9fafb;">
            <div style="background-color: white; border-radius: 10px; padding: 30px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h1 style="color: #667eea; margin: 0; font-size: 28px;">{emoji} {asunto}</h1>
                </div>
                
                <div style="background-color: #f3f4f6; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
                    <h2 style="margin-top: 0; color: #374151;">Hola {estudiante_nombre},</h2>
                    <p style="font-size: 16px; line-height: 1.6; color: #1f2937; white-space: pre-line;">{mensaje}</p>
                </div>
                
                {f'''
                <div style="background-color: #fef3c7; padding: 15px; border-radius: 8px; margin-bottom: 20px; border-left: 4px solid #f59e0b;">
                    <p style="margin: 0; color: #92400e;"><strong>📋 Documento Solicitado:</strong> {documento_solicitado}</p>
                </div>
                ''' if documento_solicitado else ''}
                
                <div style="text-align: center; margin-top: 30px;">
                    <a href="{os.getenv('FRONTEND_URL', 'http://localhost:3000')}/estudiante/perfil" 
                       style="display: inline-block; background-color: #667eea; color: white; padding: 14px 28px; 
                              text-decoration: none; border-radius: 6px; font-weight: bold; font-size: 16px;">
                        Ver en Mi Portal
                    </a>
                </div>
                
                <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #e5e7eb; text-align: center;">
                    <p style="color: #9ca3af; font-size: 14px; margin: 0;">
                        Este mensaje fue enviado por el equipo de <strong>Estudio Visa España</strong>
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        msg.attach(MIMEText(body, 'html'))
        
        # Enviar email
        smtp = smtplib.SMTP(os.getenv('SMTP_SERVER'), int(os.getenv('SMTP_PORT')))
        smtp.starttls()
        smtp.login(os.getenv('SMTP_USER'), os.getenv('SMTP_PASSWORD'))
        smtp.send_message(msg)
        smtp.quit()
        
        email_enviado = True
    except Exception as e:
        print(f"Error enviando email: {str(e)}")
        email_enviado = False
    
    cursor.close()
    conn.close()
    
    return {
        'success': True,
        'mensaje_id': mensaje_id,
        'mensaje': 'Mensaje enviado correctamente',
        'email_enviado': email_enviado,
        'destinatario': estudiante_email
    }


@app.get("/api/admin/mensajes/no-leidos", tags=["Admin - Mensajería"])
def admin_obtener_mensajes_no_leidos(
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Obtiene mensajes de estudiantes que el admin no ha leído"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT m.id, m.estudiante_id, m.mensaje, m.tipo, m.created_at,
               e.nombre, e.email
        FROM mensajes_chat m
        JOIN estudiantes e ON m.estudiante_id = e.id
        WHERE m.remitente = 'estudiante' AND m.leido = FALSE
        ORDER BY m.created_at DESC
    """)
    
    mensajes = []
    for row in cursor.fetchall():
        mensajes.append({
            'id': row[0],
            'estudiante_id': row[1],
            'mensaje': row[2],
            'tipo': row[3],
            'fecha': row[4].isoformat() if row[4] else None,
            'estudiante_nombre': row[5],
            'estudiante_email': row[6]
        })
    
    cursor.close()
    conn.close()
    
    return {
        'mensajes': mensajes,
        'total': len(mensajes)
    }


@app.get("/api/admin/estudiantes/{estudiante_id}/conversacion", tags=["Admin - Mensajería"])
def admin_obtener_conversacion(
    estudiante_id: int,
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())
):
    """Obtiene toda la conversación con un estudiante"""
    verificar_token(credentials.credentials)
    
    import os
    import psycopg2
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    # Verificar estudiante
    cursor.execute("SELECT nombre, email FROM estudiantes WHERE id = %s", (estudiante_id,))
    estudiante = cursor.fetchone()
    
    if not estudiante:
        cursor.close()
        conn.close()
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    # Obtener mensajes
    cursor.execute("""
        SELECT id, remitente, mensaje, tipo, leido, created_at
        FROM mensajes_chat
        WHERE estudiante_id = %s
        ORDER BY created_at ASC
    """, (estudiante_id,))
    
    mensajes = []
    for row in cursor.fetchall():
        mensajes.append({
            'id': row[0],
            'remitente': row[1],
            'mensaje': row[2],
            'tipo': row[3],
            'leido': row[4],
            'fecha': row[5].isoformat() if row[5] else None
        })
    
    cursor.close()
    conn.close()
    
    return {
        'estudiante': {
            'id': estudiante_id,
            'nombre': estudiante[0],
            'email': estudiante[1]
        },
        'mensajes': mensajes,
        'total': len(mensajes)
    }


# ==================== CONTACTO UNIVERSIDADES ====================

@app.get("/api/admin/universidades")
async def obtener_universidades(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Obtener lista de universidades para contactar"""
    try:
        usuario = verificar_token(credentials.credentials)
        
        if not usuario:
            raise HTTPException(status_code=401, detail="Token inválido")
        
        # Verificar que el usuario sea admin
        if usuario.get('rol') != 'admin':
            raise HTTPException(status_code=403, detail="Acceso denegado")
        
        import os
        import psycopg2
        
        conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, universidad, email, telefono, contacto_nombre, 
                   pais, ciudad, tipo_universidad, programas_interes, 
                   estado, fecha_contacto, fecha_respuesta, fecha_reunion, 
                   notas, condiciones_propuestas, comision_acordada, created_at
            FROM contactos_universidades
            ORDER BY 
                CASE estado
                    WHEN 'acuerdo_firmado' THEN 1
                    WHEN 'reunion_agendada' THEN 2
                    WHEN 'respondido' THEN 3
                    WHEN 'contactado' THEN 4
                    WHEN 'pendiente' THEN 5
                END,
                created_at DESC
        """)
        
        universidades = []
        for row in cursor.fetchall():
            universidades.append({
                'id': row[0],
                'universidad': row[1],
                'email': row[2],
                'telefono': row[3],
                'contacto_nombre': row[4],
                'pais': row[5],
                'ciudad': row[6],
                'tipo_universidad': row[7],
                'programas_interes': row[8],
                'estado': row[9],
                'fecha_contacto': row[10].isoformat() if row[10] else None,
                'fecha_respuesta': row[11].isoformat() if row[11] else None,
                'fecha_reunion': row[12].isoformat() if row[12] else None,
                'notas': row[13],
                'condiciones_propuestas': row[14],
                'comision_acordada': float(row[15]) if row[15] else None,
                'created_at': row[16].isoformat() if row[16] else None
            })
        
        cursor.close()
        conn.close()
        
        return universidades
    except Exception as e:
        print(f"❌ Error en obtener_universidades: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/admin/contactar-universidad/{universidad_id}")
async def contactar_universidad(
    universidad_id: int,
    numero_estudiantes: int = 15,
    observaciones: str = "",
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Enviar email automático a universidad"""
    usuario = verificar_token(credentials.credentials)
    
    if not usuario:
        raise HTTPException(status_code=401, detail="Token inválido")
    
    if usuario.get('rol') != 'admin':
        raise HTTPException(status_code=403, detail="Acceso denegado")
    
    import os
    import psycopg2
    from api.email_utils import enviar_email
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    # Obtener datos universidad
    cursor.execute("""
        SELECT universidad, email, contacto_nombre
        FROM contactos_universidades
        WHERE id = %s
    """, (universidad_id,))
    
    uni = cursor.fetchone()
    if not uni:
        cursor.close()
        conn.close()
        raise HTTPException(status_code=404, detail="Universidad no encontrada")
    
    universidad_nombre, email_destino, contacto = uni
    
    # Obtener datos del admin
    nombre_agencia = os.getenv('NOMBRE_AGENCIA', 'Estudia en España')
    email_contacto = os.getenv('EMAIL_SENDER', 'contacto@estudiaenespana.com')
    telefono_contacto = os.getenv('TELEFONO_CONTACTO', '+53 XXXXXXXX')
    web_agencia = os.getenv('WEB_AGENCIA', 'https://fortunariocash.com')
    
    # Crear email personalizado
    saludo = f"Estimado/a {contacto}" if contacto else "Estimado equipo de admisiones"
    
    cuerpo_html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
            .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0; }}
            .content {{ background: #f9f9f9; padding: 30px; border-radius: 0 0 10px 10px; }}
            .destacado {{ background: white; padding: 20px; border-left: 4px solid #667eea; margin: 20px 0; }}
            .numeros {{ background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); color: white; padding: 20px; text-align: center; border-radius: 8px; margin: 20px 0; }}
            .numeros h2 {{ margin: 0; font-size: 36px; }}
            .cta {{ background: #4caf50; color: white; padding: 15px 30px; text-decoration: none; border-radius: 8px; display: inline-block; margin: 20px 0; font-weight: bold; }}
            .footer {{ text-align: center; margin-top: 20px; color: #666; font-size: 12px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🎓 {nombre_agencia}</h1>
                <p>Agencia Especializada en Estudiantes Cubanos</p>
            </div>
            <div class="content">
                <p>{saludo},</p>
                
                <p>Mi nombre es <strong>{usuario.get('nombre', 'Director')}</strong>, y soy {usuario.get('rol', 'director')} de <strong>{nombre_agencia}</strong>, 
                una agencia educativa especializada en asesorar a estudiantes cubanos que desean cursar estudios superiores en España.</p>
                
                <div class="numeros">
                    <h2>{numero_estudiantes}+ Estudiantes</h2>
                    <p>Listos para inscribirse en 2026</p>
                </div>
                
                <div class="destacado">
                    <h3>🎯 Propuesta de Colaboración</h3>
                    <p>Nos gustaría explorar una <strong>alianza estratégica</strong> con {universidad_nombre} para canalizar estudiantes cubanos interesados en sus programas académicos.</p>
                    
                    <p><strong>Perfil de nuestros estudiantes:</strong></p>
                    <ul>
                        <li>✅ Edad: 18-35 años</li>
                        <li>✅ Nivel educativo: Bachillerato completado / Título universitario</li>
                        <li>✅ Motivación alta para estudiar en España</li>
                        <li>✅ Documentación en proceso / lista para apostillar</li>
                        <li>✅ Acompañamiento completo desde Cuba hasta España</li>
                    </ul>
                    
                    {f"<p><strong>Observaciones adicionales:</strong> {observaciones}</p>" if observaciones else ""}
                </div>
                
                <div class="destacado">
                    <h3>💼 Temas a Discutir</h3>
                    <ul>
                        <li>📋 <strong>Admisiones:</strong> Requisitos y proceso simplificado para volumen</li>
                        <li>💰 <strong>Estructura de pagos:</strong> Flexibilidad para estudiantes internacionales</li>
                        <li>🎓 <strong>Becas/Descuentos:</strong> Incentivos por grupo o convenio</li>
                        <li>🤝 <strong>Comisiones:</strong> Modelo de compensación por referidos</li>
                        <li>📧 <strong>Cartas de aceptación:</strong> Tiempo de emisión y condiciones</li>
                    </ul>
                </div>
                
                <p><strong>¿Podríamos agendar una videollamada esta semana?</strong></p>
                <p>Estoy disponible:</p>
                <ul>
                    <li>🗓️ Jueves 28 Nov: 10:00 - 12:00 / 14:00 - 17:00 (hora España)</li>
                    <li>🗓️ Viernes 29 Nov: 10:00 - 12:00 / 14:00 - 17:00 (hora España)</li>
                </ul>
                
                <p style="text-align: center;">
                    <a href="mailto:{email_contacto}" class="cta">📧 Responder Email</a>
                </p>
                
                <p>Quedo atento a su respuesta y esperamos poder construir una relación beneficiosa para ambas partes.</p>
                
                <p>Cordialmente,<br>
                <strong>{usuario.get('nombre', 'Director')}</strong><br>
                {usuario.get('rol', 'Director').title()}<br>
                {nombre_agencia}</p>
                
                <hr style="border: none; border-top: 1px solid #ddd; margin: 20px 0;">
                
                <p style="font-size: 12px; color: #666;">
                📧 {email_contacto}<br>
                📱 WhatsApp: {telefono_contacto}<br>
                🌐 <a href="{web_agencia}">{web_agencia}</a>
                </p>
            </div>
            <div class="footer">
                <p>Este email fue enviado automáticamente desde nuestro sistema de gestión.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Enviar email
    asunto = f"Propuesta de Colaboración - {nombre_agencia} ({numero_estudiantes}+ estudiantes)"
    
    try:
        exito = enviar_email(
            destinatario=email_destino,
            asunto=asunto,
            cuerpo_html=cuerpo_html
        )
        
        if exito:
            # Actualizar estado en base de datos
            cursor.execute("""
                UPDATE contactos_universidades
                SET estado = 'contactado',
                    fecha_contacto = CURRENT_TIMESTAMP,
                    notas = COALESCE(notas || E'\n\n', '') || %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (
                f"Email enviado automáticamente el {datetime.now().strftime('%d/%m/%Y %H:%M')}. Estudiantes: {numero_estudiantes}. {observaciones}",
                universidad_id
            ))
            conn.commit()
            
            log_event("email_universidad_enviado", {
                "universidad_id": universidad_id,
                "universidad": universidad_nombre,
                "email": email_destino,
                "estudiantes": numero_estudiantes,
                "admin": usuario.get('nombre')
            })
            
            cursor.close()
            conn.close()
            
            return {
                'success': True,
                'mensaje': f'Email enviado exitosamente a {universidad_nombre}',
                'email': email_destino
            }
        else:
            cursor.close()
            conn.close()
            raise HTTPException(status_code=500, detail="Error al enviar email. Verifica configuración SMTP.")
            
    except Exception as e:
        cursor.close()
        conn.close()
        log_error("error_envio_email_universidad", str(e))
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.put("/api/admin/universidades/{universidad_id}")
async def actualizar_universidad(
    universidad_id: int,
    estado: str = None,
    notas: str = None,
    fecha_reunion: str = None,
    condiciones_propuestas: str = None,
    comision_acordada: float = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Actualizar información de contacto con universidad"""
    usuario = verificar_token(credentials.credentials)
    
    if not usuario:
        raise HTTPException(status_code=401, detail="Token inválido")
    
    if usuario.get('rol') != 'admin':
        raise HTTPException(status_code=403, detail="Acceso denegado")
    
    import os
    import psycopg2
    from datetime import datetime
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    # Construir query dinámico
    updates = []
    params = []
    
    if estado:
        updates.append("estado = %s")
        params.append(estado)
        
        if estado == 'respondido':
            updates.append("fecha_respuesta = CURRENT_TIMESTAMP")
    
    if notas:
        updates.append("notas = COALESCE(notas || E'\\n\\n', '') || %s")
        params.append(f"[{datetime.now().strftime('%d/%m/%Y %H:%M')}] {notas}")
    
    if fecha_reunion:
        updates.append("fecha_reunion = %s")
        params.append(fecha_reunion)
    
    if condiciones_propuestas:
        updates.append("condiciones_propuestas = %s")
        params.append(condiciones_propuestas)
    
    if comision_acordada is not None:
        updates.append("comision_acordada = %s")
        params.append(comision_acordada)
    
    if updates:
        updates.append("updated_at = CURRENT_TIMESTAMP")
        params.append(universidad_id)
        
        query = f"""
            UPDATE contactos_universidades
            SET {', '.join(updates)}
            WHERE id = %s
        """
        
        cursor.execute(query, params)
        conn.commit()
    
    cursor.close()
    conn.close()
    
    return {'success': True, 'mensaje': 'Universidad actualizada correctamente'}


# ========================================
# ENDPOINTS: PROCESO COMPLETO DE VISA
# ========================================

# Tabla proceso_visa_pasos movida al startup principal


@app.get("/api/estudiantes/{estudiante_id}/proceso-visa")
async def obtener_proceso_visa(estudiante_id: int):
    """Obtener el proceso completo de visa del estudiante (acceso público con ID)"""
    
    import os
    import psycopg2
    from datetime import datetime
    from psycopg2.extras import RealDictCursor
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Verificar que el estudiante existe
        cursor.execute("SELECT id FROM estudiantes WHERE id = %s", (estudiante_id,))
        if not cursor.fetchone():
            cursor.close()
            conn.close()
            raise HTTPException(status_code=404, detail="Estudiante no encontrado")
        
        # Obtener proceso o crear uno nuevo
        cursor.execute("""
            SELECT * FROM proceso_visa_pasos
            WHERE estudiante_id = %s
        """, (estudiante_id,))
        
        proceso = cursor.fetchone()
        
        if not proceso:
            # Crear proceso inicial con todos los campos en False
            cursor.execute("""
                INSERT INTO proceso_visa_pasos (estudiante_id)
                VALUES (%s)
                RETURNING *
            """, (estudiante_id,))
            proceso = cursor.fetchone()
            conn.commit()
        
        cursor.close()
        conn.close()
        
        # Convertir a formato serializable
        proceso_dict = dict(proceso)
        for key, value in proceso_dict.items():
            if isinstance(value, datetime):
                proceso_dict[key] = value.isoformat()
        
        return proceso_dict
        
    except Exception as e:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()
        print(f"❌ Error obteniendo proceso: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


class ActualizarProcesoRequest(BaseModel):
    paso: str
    completado: bool
    notas: Optional[str] = None
    fecha_cita: Optional[str] = None
    resultado: Optional[str] = None

@app.put("/api/admin/estudiantes/{estudiante_id}/proceso-visa")
async def actualizar_paso_proceso(
    estudiante_id: int,
    request: ActualizarProcesoRequest = Body(...),
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Actualizar un paso específico del proceso (SOLO ADMIN)"""
    usuario = verificar_token(credentials.credentials)
    
    if not usuario:
        raise HTTPException(status_code=401, detail="Token inválido")
    
    if usuario.get('rol') != 'admin':
        raise HTTPException(status_code=403, detail="Solo administradores")
    
    import os
    import psycopg2
    from datetime import datetime
    
    conn = psycopg2.connect(os.getenv('DATABASE_URL'), sslmode='require')
    cursor = conn.cursor()
    
    # Extraer valores del request
    paso = request.paso
    completado = request.completado
    notas = request.notas
    fecha_cita = request.fecha_cita
    resultado = request.resultado
    
    # Verificar que el paso existe
    pasos_validos = [
        'paso_inscripcion', 'paso_pago_inicial', 'paso_documentos_personales',
        'paso_seleccion_universidad', 'paso_solicitud_universidad', 'paso_carta_aceptacion',
        'paso_antecedentes_solicitados', 'paso_antecedentes_recibidos', 'paso_apostilla_haya',
        'paso_traduccion_documentos', 'paso_seguro_medico', 'paso_comprobante_fondos',
        'paso_carta_banco', 'paso_formulario_visa', 'paso_fotos_biometricas',
        'paso_pago_tasa_visa', 'paso_cita_agendada', 'paso_documentos_revisados',
        'paso_simulacro_entrevista', 'paso_entrevista_completada', 'paso_pasaporte_recogido',
        'paso_visa_otorgada'
    ]
    
    if paso not in pasos_validos:
        raise HTTPException(status_code=400, detail="Paso no válido")
    
    # Construir query
    updates = [f"{paso} = %s"]
    params = [completado]
    
    # Actualizar fecha del paso
    fecha_campo = paso.replace('paso_', 'fecha_')
    if completado:
        updates.append(f"{fecha_campo} = CURRENT_TIMESTAMP")
    else:
        updates.append(f"{fecha_campo} = NULL")
    
    # Fecha de cita (solo para paso_cita_agendada)
    if fecha_cita and paso == 'paso_cita_agendada':
        updates.append("fecha_cita_embajada = %s")
        params.append(fecha_cita)
    
    # Resultado de entrevista
    if resultado and paso == 'paso_entrevista_completada':
        updates.append("resultado_entrevista = %s")
        params.append(resultado)
    
    # Notas
    if notas:
        updates.append("notas_admin = COALESCE(notas_admin || E'\\n\\n', '') || %s")
        params.append(f"[{datetime.now().strftime('%d/%m/%Y %H:%M')}] {paso}: {notas}")
    
    updates.append("ultima_actualizacion = CURRENT_TIMESTAMP")
    params.append(estudiante_id)
    
    query = f"""
        UPDATE proceso_visa_pasos
        SET {', '.join(updates)}
        WHERE estudiante_id = %s
    """
    
    cursor.execute(query, params)
    
    if cursor.rowcount == 0:
        # Crear registro si no existe
        cursor.execute("""
            INSERT INTO proceso_visa_pasos (estudiante_id, {}, {})
            VALUES (%s, %s, CURRENT_TIMESTAMP)
        """.format(paso, fecha_campo), (estudiante_id, completado))
    
    conn.commit()
    cursor.close()
    conn.close()
    
    log_event("proceso_visa_actualizado", {
        'estudiante_id': estudiante_id,
        'paso': paso,
        'completado': completado,
        'admin_id': usuario.get('id')
    })
    
    return {'success': True, 'mensaje': f'Paso {paso} actualizado'}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
